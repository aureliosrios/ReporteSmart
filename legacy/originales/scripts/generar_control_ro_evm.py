import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generar_simulacion_ro_semana1():
    # 1. Cargar presupuesto base JSON
    with open("presupuesto_con_apu.json", "r", encoding="utf-8") as f:
        presupuesto = json.load(f)

    # 2. Entrada Sintética de Datos Diarios de la Semana 1 (Acumulado 6 días de trabajo)
    # Avances reales de campo (Ing. Campo)
    avances_campo_semana1 = {
        "01.01": 1.0,         # 100% Obras Preliminares
        "01.02.01": 600.0,    # 600 m de Trazo y Replanteo (Planificado: 500 m) -> Adelantado
        "01.02.02": 800.0,    # 800 m3 Excavación Zanja (Planificado: 900 m3) -> Retrasado
        "01.02.03": 400.0,    # 400 m Cama de Arena (Planificado: 450 m)
        "01.02.04": 200.0,    # 200 m Tubería PVC 200mm (Planificado: 240 m)
        "01.02.05": 300.0,    # 300 m3 Relleno Compactado
        "01.02.06": 4.0,      # 4 Buzones prefabricados
        "01.02.07": 0.0,
        "01.03.01": 0.0,
        "01.03.02": 0.0,
        "01.04.01": 0.0
    }

    # Planificado a la Semana 1 (PV por Partida)
    planificado_semana1 = {
        "01.01": 1.0,
        "01.02.01": 500.0,
        "01.02.02": 900.0,
        "01.02.03": 450.0,
        "01.02.04": 240.0,
        "01.02.05": 350.0,
        "01.02.06": 4.0,
        "01.02.07": 0.0,
        "01.03.01": 0.0,
        "01.03.02": 0.0,
        "01.04.01": 0.0
    }

    # Costos Reales Incurridos en la Semana 1 (AC agrupado por partida y elemento)
    # Fuente: Bildin (MO) + Almacén (MAT) + Partes de Equipo/Subcontrato (EQP/SUB)
    costos_reales_semana1 = {
        "01.01": {"MO": 4800.00, "MAT": 3600.00, "EQP": 2400.00, "SUB": 4000.00},     # Total AC: 14,800 (PV: 14,500)
        "01.02.01": {"MO": 5100.00, "MAT": 750.00, "EQP": 1800.00, "SUB": 0.00},      # Total AC: 7,650 (EV: 2,100... espera, P.U. es 3.50 -> EV: 2,100)
        "01.02.02": {"MO": 2100.00, "MAT": 0.00, "EQP": 12800.00, "SUB": 0.00},       # Total AC: 14,900 (EV: 11,600)
        "01.02.03": {"MO": 2100.00, "MAT": 2900.00, "EQP": 450.00, "SUB": 0.00},       # Total AC: 5,450 (EV: 5,120)
        "01.02.04": {"MO": 3000.00, "MAT": 8900.00, "EQP": 220.00, "SUB": 0.00},       # Total AC: 12,120 (EV: 11,600)
        "01.02.05": {"MO": 1800.00, "MAT": 2100.00, "EQP": 1400.00, "SUB": 200.00},    # Total AC: 5,500 (EV: 4,800)
        "01.02.06": {"MO": 2100.00, "MAT": 4200.00, "EQP": 400.00, "SUB": 1000.00},   # Total AC: 7,700 (EV: 7,400)
        "01.02.07": {"MO": 0.00, "MAT": 0.00, "EQP": 0.00, "SUB": 0.00},
        "01.03.01": {"MO": 0.00, "MAT": 0.00, "EQP": 0.00, "SUB": 0.00},
        "01.03.02": {"MO": 0.00, "MAT": 0.00, "EQP": 0.00, "SUB": 0.00},
        "01.04.01": {"MO": 0.00, "MAT": 0.00, "EQP": 0.00, "SUB": 0.00}
    }

    # 3. Procesar Tabla EVM por Partida
    tabla_evm = []
    totales = {"PV": 0.0, "EV": 0.0, "AC": 0.0, "BAC": 0.0}

    for apu in presupuesto["analisis_precios_unitarios"]:
        item = apu["item"]
        desc = apu["descripcion"]
        unit = apu["unidad"]
        metrado_tot = apu["metrado"]
        pu_meta = apu["precio_unitario_directo"]
        bac = apu["costo_total_partida_directo"]

        met_pv = planificado_semana1.get(item, 0.0)
        met_ev = avances_campo_semana1.get(item, 0.0)
        c_real = costos_reales_semana1.get(item, {"MO":0,"MAT":0,"EQP":0,"SUB":0})

        pv = round(met_pv * pu_meta, 2)
        ev = round(met_ev * pu_meta, 2)
        ac = round(c_real["MO"] + c_real["MAT"] + c_real["EQP"] + c_real["SUB"], 2)

        cv = round(ev - ac, 2)
        sv = round(ev - pv, 2)
        cpi = round(ev / ac, 2) if ac > 0 else 1.0
        spi = round(ev / pv, 2) if pv > 0 else 1.0

        # Proyección EAC a Término
        eac = round(ac + (bac - ev) / cpi, 2) if cpi > 0 else bac
        var_eac = round(bac - eac, 2)

        totales["BAC"] += bac
        totales["PV"] += pv
        totales["EV"] += ev
        totales["AC"] += ac

        tabla_evm.append({
            "item": item,
            "descripcion": desc,
            "unidad": unit,
            "metrado_total": metrado_tot,
            "pu_meta": pu_meta,
            "bac": bac,
            "metrado_pv": met_pv,
            "pv": pv,
            "metrado_ev": met_ev,
            "ev": ev,
            "ac": ac,
            "cv": cv,
            "sv": sv,
            "cpi": cpi,
            "spi": spi,
            "eac": eac,
            "variacion_eac": var_eac
        })

    # Resumen Global KPI
    tot_pv = totales["PV"]
    tot_ev = totales["EV"]
    tot_ac = totales["AC"]
    tot_bac = totales["BAC"]
    cpi_global = round(tot_ev / tot_ac, 2) if tot_ac > 0 else 1.0
    spi_global = round(tot_ev / tot_pv, 2) if tot_pv > 0 else 1.0
    eac_global = round(tot_ac + (tot_bac - tot_ev) / cpi_global, 2)
    desvio_global = round(tot_bac - eac_global, 2)

    # 4. Generar Libro Excel de Control Semanal RO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control Semanal RO - Sem 1"
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name="Arial", size=14, bold=True, color="1F4E78")
    font_subtitle = Font(name="Arial", size=11, bold=True, color="2F5597")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)
    
    fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_blue_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_subtotal = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_total = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    fill_alert_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_alert_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.append(["INFORME SEMANAL DE RESULTADO OPERATIVO Y VALOR GANADO (EVM) - SEMANA 1"])
    ws["A1"].font = font_title
    ws.append(["Proyecto: Redes Sanitarias de Agua Potable y Alcantarillado - Habilitación Urbana Los Cedros"])
    ws["A2"].font = font_subtitle
    ws.append(["Periodo Evaluado: Semana 1 (Día 1 al Día 6) | Corte: Viernes 5:00 PM"])
    ws.append([])

    headers = [
        "Item", "Descripción Partida", "Und.", "Metrado Total", "P.U. Meta (S/)", "BAC Meta (S/)",
        "PV Plan. (S/)", "EV Ganado (S/)", "AC Real (S/)", "CV Costo (S/)", "SV Plazo (S/)",
        "CPI (Costo)", "SPI (Plazo)", "EAC Proyectado (S/)", "Desvío EAC (S/)"
    ]

    current_row = 5
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = Alignment(horizontal="center" if col_idx in [1,3,12,13] else "right" if col_idx >= 4 else "left")
    current_row += 1

    start_data_row = current_row
    for r in tabla_evm:
        ws.cell(row=current_row, column=1, value=r["item"]).font = font_regular
        ws.cell(row=current_row, column=2, value=r["descripcion"]).font = font_regular
        ws.cell(row=current_row, column=3, value=r["unidad"]).font = font_regular
        
        ws.cell(row=current_row, column=4, value=r["metrado_total"]).number_format = "#,##0.00"
        ws.cell(row=current_row, column=5, value=r["pu_meta"]).number_format = "#,##0.00"
        ws.cell(row=current_row, column=6, value=r["bac"]).number_format = "#,##0.00"

        ws.cell(row=current_row, column=7, value=r["pv"]).number_format = "#,##0.00"
        ws.cell(row=current_row, column=8, value=r["ev"]).number_format = "#,##0.00"
        ws.cell(row=current_row, column=9, value=r["ac"]).number_format = "#,##0.00"

        # CV = EV - AC
        c_cv = ws.cell(row=current_row, column=10, value=f"=H{current_row}-I{current_row}")
        c_cv.number_format = "#,##0.00"

        # SV = EV - PV
        c_sv = ws.cell(row=current_row, column=11, value=f"=H{current_row}-G{current_row}")
        c_sv.number_format = "#,##0.00"

        # CPI = EV / AC
        c_cpi = ws.cell(row=current_row, column=12, value=f"=IF(I{current_row}>0, H{current_row}/I{current_row}, 1.0)")
        c_cpi.number_format = "0.00"
        if r["ac"] > 0 and r["cpi"] < 1.0:
            c_cpi.fill = fill_alert_red
        elif r["ac"] > 0 and r["cpi"] >= 1.0:
            c_cpi.fill = fill_alert_green

        # SPI = EV / PV
        c_spi = ws.cell(row=current_row, column=13, value=f"=IF(G{current_row}>0, H{current_row}/G{current_row}, 1.0)")
        c_spi.number_format = "0.00"

        # EAC = AC + (BAC - EV) / CPI
        c_eac = ws.cell(row=current_row, column=14, value=f"=IF(L{current_row}>0, I{current_row}+(F{current_row}-H{current_row})/L{current_row}, F{current_row})")
        c_eac.number_format = "#,##0.00"

        # Desvío = BAC - EAC
        c_desv = ws.cell(row=current_row, column=15, value=f"=F{current_row}-N{current_row}")
        c_desv.number_format = "#,##0.00"

        current_row += 1

    end_data_row = current_row - 1

    # TOTALES SEMANA 1
    ws.cell(row=current_row, column=2, value="TOTALES COSTO DIRECTO ACUMULADO").font = font_bold
    ws.cell(row=current_row, column=6, value=f"=SUM(F{start_data_row}:F{end_data_row})").number_format = "#,##0.00"
    ws.cell(row=current_row, column=7, value=f"=SUM(G{start_data_row}:G{end_data_row})").number_format = "#,##0.00"
    ws.cell(row=current_row, column=8, value=f"=SUM(H{start_data_row}:H{end_data_row})").number_format = "#,##0.00"
    ws.cell(row=current_row, column=9, value=f"=SUM(I{start_data_row}:I{end_data_row})").number_format = "#,##0.00"
    
    ws.cell(row=current_row, column=10, value=f"=H{current_row}-I{current_row}").number_format = "#,##0.00"
    ws.cell(row=current_row, column=11, value=f"=H{current_row}-G{current_row}").number_format = "#,##0.00"

    ws.cell(row=current_row, column=12, value=f"=H{current_row}/I{current_row}").number_format = "0.00"
    ws.cell(row=current_row, column=13, value=f"=H{current_row}/G{current_row}").number_format = "0.00"

    ws.cell(row=current_row, column=14, value=f"=I{current_row}+(F{current_row}-H{current_row})/L{current_row}").number_format = "#,##0.00"
    ws.cell(row=current_row, column=15, value=f"=F{current_row}-N{current_row}").number_format = "#,##0.00"

    for col in range(1, 16):
        c = ws.cell(row=current_row, column=col)
        c.font = font_bold
        c.fill = fill_total

    # Guardar Excel de reporte
    out_excel = "d:\\Agentes de IA\\Habilitación urbana\\Reporte_Semanal_RO_Semana1.xlsx"
    wb.save(out_excel)
    print(f"Reporte Semanal de RO generado en: {out_excel}")

    # Guardar JSON de reporte
    out_json = "d:\\Agentes de IA\\Habilitación urbana\\reporte_ro_semana1.json"
    data_report = {
        "semana": 1,
        "indicadores_kpi_globales": {
            "pv_planificado": tot_pv,
            "ev_ganado": tot_ev,
            "ac_costo_real": tot_ac,
            "bac_presupuesto_total": tot_bac,
            "cpi_indice_costo": cpi_global,
            "spi_indice_plazo": spi_global,
            "eac_proyeccion_cierre": eac_global,
            "variacion_margen_esperado": desvio_global
        },
        "desglose_partidas": tabla_evm
    }
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(data_report, f, indent=2, ensure_ascii=False)
    print(f"JSON de reporte semanal guardado en: {out_json}")

if __name__ == "__main__":
    generar_simulacion_ro_semana1()
