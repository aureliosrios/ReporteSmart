import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_ro_excel_database_traceable():
    excel_filename = "Base_de_Datos_RO_Reportabilidad.xlsx"

    with open('presupuesto_con_apu.json', 'r', encoding='utf-8') as f:
        presupuesto_apu = json.load(f)

    with open('base_datos_ro_diaria.json', 'r', encoding='utf-8') as f:
        db_diaria = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_family = "Calibri"
    navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    blue_header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    green_header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
    purple_header_fill = PatternFill(start_color="7E22CE", end_color="7E22CE", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_bold = Font(name=font_family, size=11, bold=True)
    font_regular = Font(name=font_family, size=11)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    fmt_currency = 'S/ #,##0.00'
    fmt_num = '#,##0.00'

    # 1. MAESTRO_RECURSOS_Y_APU
    ws_rec = wb.create_sheet(title="MAESTRO_RECURSOS_Y_APU")
    ws_rec.views.sheetView[0].showGridLines = True

    ws_rec.merge_cells("A1:E1")
    ws_rec["A1"] = "CATÁLOGO MAESTRO COMPLETO DE RECURSOS Y APU META (DICCIONARIO UNIFICADO)"
    ws_rec["A1"].font = font_title
    ws_rec["A1"].fill = navy_fill
    ws_rec["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers_rec = ["Categoría Tipo", "Código Recurso/Partida", "Descripción Detallada del Recurso / APU", "Unidad", "Precio Unitario Meta (S/)"]
    ws_rec.append([])
    ws_rec.append(headers_rec)

    for col_idx in range(1, 6):
        cell = ws_rec.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = purple_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rec_row = 4
    for cat_key, cat_name in [("mano_obra", "Mano de Obra"), ("materiales", "Materiales"), ("equipos", "Equipos"), ("subcontratos", "Subcontratos")]:
        for code, info in presupuesto_apu["precios_recursos_maestros"][cat_key].items():
            ws_rec.cell(row=rec_row, column=1, value=cat_name)
            ws_rec.cell(row=rec_row, column=2, value=code).alignment = Alignment(horizontal="center")
            ws_rec.cell(row=rec_row, column=3, value=info["desc"])
            ws_rec.cell(row=rec_row, column=4, value=info["unit"]).alignment = Alignment(horizontal="center")
            ws_rec.cell(row=rec_row, column=5, value=info["precio"]).number_format = fmt_currency
            rec_row += 1

    for apu in presupuesto_apu["analisis_precios_unitarios"]:
        ws_rec.cell(row=rec_row, column=1, value="Partida APU (EV)")
        ws_rec.cell(row=rec_row, column=2, value=apu["item"]).alignment = Alignment(horizontal="center")
        ws_rec.cell(row=rec_row, column=3, value=f"Partida {apu['item']} - {apu['descripcion']}")
        ws_rec.cell(row=rec_row, column=4, value=apu["unidad"]).alignment = Alignment(horizontal="center")
        ws_rec.cell(row=rec_row, column=5, value=apu["precio_unitario_directo"]).number_format = fmt_currency
        rec_row += 1

    for r in range(4, rec_row):
        for c in range(1, 6):
            cell = ws_rec.cell(row=r, column=c)
            cell.font = font_regular
            cell.border = thin_border

    # 2. DB_REGISTROS_DIARIOS
    ws_logs = wb.create_sheet(title="DB_REGISTROS_DIARIOS")
    ws_logs.views.sheetView[0].showGridLines = True

    ws_logs.merge_cells("A1:K1")
    ws_logs["A1"] = "BASE DE DATOS REGISTROS DIARIOS CON TRAZABILIDAD Y BUSQUEDA DE TARIFAS META"
    ws_logs["A1"].font = font_title
    ws_logs["A1"].fill = navy_fill
    ws_logs["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers_logs = ["ID Registro", "Fecha", "Emisor / Rol", "Código WBS", "Código Recurso/Partida", "Descripción Detallada", "Cantidad", "Unidad", "Costo Unitario Formula (S/)", "Costo Total Formula (S/)", "Categoría EVM"]
    ws_logs.append([])
    ws_logs.append(headers_logs)

    for col_idx in range(1, 12):
        cell = ws_logs.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = green_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    last_rec_row = rec_row - 1
    logs_list = db_diaria.get('registros_diarios_logs', [])

    for idx, log in enumerate(logs_list):
        r = 4 + idx
        id_reg = log.get('id_registro') or f"LOG-{idx+1}"
        fecha = log.get('fecha') or ''
        rol = log.get('emisor_rol') or log.get('rol') or 'Tareador (Bildin)'
        wbs = log.get('wbs_codigo') or log.get('wbs') or 'WBS-200'
        code_rec = log.get('codigo_recurso_partida') or log.get('codigoRecurso') or 'MO_OPERARIO'
        desc = log.get('descripcion') or log.get('detalle') or ''
        cant = float(log.get('cantidad', 0))
        unidad = log.get('unidad') or ''
        cat_evm = log.get('categoria_evm') or log.get('tipo') or 'AC_MO'

        ws_logs.cell(row=r, column=1, value=id_reg).alignment = Alignment(horizontal="center")
        ws_logs.cell(row=r, column=2, value=fecha).alignment = Alignment(horizontal="center")
        ws_logs.cell(row=r, column=3, value=rol)
        ws_logs.cell(row=r, column=4, value=wbs).alignment = Alignment(horizontal="center")
        ws_logs.cell(row=r, column=5, value=code_rec).alignment = Alignment(horizontal="center")
        ws_logs.cell(row=r, column=6, value=desc)
        ws_logs.cell(row=r, column=7, value=cant).number_format = fmt_num
        ws_logs.cell(row=r, column=8, value=unidad).alignment = Alignment(horizontal="center")

        formula_pu = f'=VLOOKUP(E{r}, MAESTRO_RECURSOS_Y_APU!$B$4:$E${last_rec_row}, 4, FALSE)'
        ws_logs.cell(row=r, column=9, value=formula_pu).number_format = fmt_currency

        formula_costo = f'=G{r}*I{r}'
        ws_logs.cell(row=r, column=10, value=formula_costo).number_format = fmt_currency
        ws_logs.cell(row=r, column=11, value=cat_evm).alignment = Alignment(horizontal="center")

        for c in range(1, 12):
            cell = ws_logs.cell(row=r, column=c)
            cell.font = font_regular
            cell.border = thin_border
            if idx % 2 == 1:
                cell.fill = zebra_fill

    # 3. DASHBOARD_EVM_WBS
    ws_dash = wb.create_sheet(title="DASHBOARD_EVM_WBS")
    ws_dash.views.sheetView[0].showGridLines = True

    ws_dash.merge_cells("A1:J1")
    ws_dash["A1"] = "DASHBOARD Y CONTROL DE RESULTADO OPERATIVO (CUANTIFICACIÓN DINÁMICA CON FORMULAS SUMIFS)"
    ws_dash["A1"].font = font_title
    ws_dash["A1"].fill = navy_fill
    ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers_dash = ["Código WBS", "Descripción del Frente WBS", "Presupuesto BAC (S/)", "Planificado PV (S/)", "Valor Ganado EV (S/)", "Costo Real AC (S/)", "Variación Costo CV (S/)", "CPI (Costo)", "SPI (Plazo)", "EAC Proyectado (S/)"]
    ws_dash.append([])
    ws_dash.append(headers_dash)

    for col_idx in range(1, 11):
        cell = ws_dash.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = blue_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_start = 4
    for idx, wbs in enumerate(db_diaria['nodos_wbs_estructura']):
        curr_row = row_start + idx
        code = wbs['codigo']

        ws_dash.cell(row=curr_row, column=1, value=code).alignment = Alignment(horizontal="center")
        ws_dash.cell(row=curr_row, column=2, value=wbs['nombre'])
        ws_dash.cell(row=curr_row, column=3, value=wbs['bac_pen']).number_format = fmt_currency
        ws_dash.cell(row=curr_row, column=4, value=wbs['pv_acumulado_pen']).number_format = fmt_currency
        ws_dash.cell(row=curr_row, column=5, value=f'=SUMIFS(DB_REGISTROS_DIARIOS!J:J, DB_REGISTROS_DIARIOS!D:D, A{curr_row}, DB_REGISTROS_DIARIOS!K:K, "EV_PRODUCCION")').number_format = fmt_currency
        ws_dash.cell(row=curr_row, column=6, value=f'=SUMIFS(DB_REGISTROS_DIARIOS!J:J, DB_REGISTROS_DIARIOS!D:D, A{curr_row}, DB_REGISTROS_DIARIOS!K:K, "AC_*")').number_format = fmt_currency
        ws_dash.cell(row=curr_row, column=7, value=f'=E{curr_row}-F{curr_row}').number_format = fmt_currency
        ws_dash.cell(row=curr_row, column=8, value=f'=IF(F{curr_row}>0, E{curr_row}/F{curr_row}, 1.0)').number_format = fmt_num
        ws_dash.cell(row=curr_row, column=9, value=f'=IF(D{curr_row}>0, E{curr_row}/D{curr_row}, 1.0)').number_format = fmt_num
        ws_dash.cell(row=curr_row, column=10, value=f'=IF(H{curr_row}>0, F{curr_row}+(C{curr_row}-E{curr_row})/H{curr_row}, C{curr_row})').number_format = fmt_currency

        for c in range(1, 11):
            cell = ws_dash.cell(row=curr_row, column=c)
            cell.font = font_regular
            cell.border = thin_border

    tot_row = row_start + len(db_diaria['nodos_wbs_estructura'])
    ws_dash.cell(row=tot_row, column=1, value="TOTAL PROYECTO").font = font_bold
    ws_dash.cell(row=tot_row, column=3, value=f"=SUM(C4:C{tot_row-1})").number_format = fmt_currency
    ws_dash.cell(row=tot_row, column=4, value=f"=SUM(D4:D{tot_row-1})").number_format = fmt_currency
    ws_dash.cell(row=tot_row, column=5, value=f"=SUM(E4:E{tot_row-1})").number_format = fmt_currency
    ws_dash.cell(row=tot_row, column=6, value=f"=SUM(F4:F{tot_row-1})").number_format = fmt_currency
    ws_dash.cell(row=tot_row, column=7, value=f"=E{tot_row}-F{tot_row}").number_format = fmt_currency
    ws_dash.cell(row=tot_row, column=8, value=f"=IF(F{tot_row}>0, E{tot_row}/F{tot_row}, 1.0)").number_format = fmt_num
    ws_dash.cell(row=tot_row, column=9, value=f"=IF(D{tot_row}>0, E{tot_row}/D{tot_row}, 1.0)").number_format = fmt_num
    ws_dash.cell(row=tot_row, column=10, value=f"=SUM(J4:J{tot_row-1})").number_format = fmt_currency

    for c in range(1, 11):
        cell = ws_dash.cell(row=tot_row, column=c)
        cell.font = font_bold
        cell.border = thin_border
        cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.row in [1, 2]: continue
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    try:
        wb.save(excel_filename)
        print(f"Base de datos Excel generada exitosamente: {excel_filename}")
        return True
    except PermissionError:
        print(f"AVISO: {excel_filename} esta abierto en Microsoft Excel.")
        return False

if __name__ == "__main__":
    generate_ro_excel_database_traceable()
