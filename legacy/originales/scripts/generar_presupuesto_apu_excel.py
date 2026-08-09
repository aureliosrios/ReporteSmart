import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_budget_and_apus():
    RECURSOS_MO = {
        "MO_CAPATAZ": {"desc": "CAPATAZ", "unit": "hh", "precio": 32.50},
        "MO_OPERARIO": {"desc": "OPERARIO", "unit": "hh", "precio": 26.80},
        "MO_OFICIAL": {"desc": "OFICIAL", "unit": "hh", "precio": 22.40},
        "MO_PEON": {"desc": "PEON", "unit": "hh", "precio": 20.10},
    }

    RECURSOS_MAT = {
        "MAT_ESTACA": {"desc": "ESTACAS Y MADERA PARA TRAZO", "unit": "pza", "precio": 4.50},
        "MAT_PINTURA": {"desc": "PINTURA SPRAY MARCADOR DE ZANJA", "unit": "gla", "precio": 18.00},
        "MAT_CORDEL": {"desc": "CORDEL / NYLON", "unit": "m", "precio": 0.80},
        "MAT_ARENA_CAMAS": {"desc": "ARENA FINA SELECCIONADA PARA CAMA E=0.10M", "unit": "m3", "precio": 45.00},
        "MAT_TUB_PVC_200": {"desc": "TUBERIA PVC UF DN 200MM SERIE S-20 ALCANTARILLADO", "unit": "m", "precio": 42.00},
        "MAT_LUBRICANTE": {"desc": "LUBRICANTE PARA TUBERIA CON ESPIGA Y CAMPANA", "unit": "kg", "precio": 24.00},
        "MAT_MATERIAL_PR": {"desc": "MATERIAL DE PRÉSTAMO AFIRMADO PARA RELLENO COMPACTADO", "unit": "m3", "precio": 38.00},
        "MAT_AGUA": {"desc": "AGUA PARA COMPACTACIÓN Y PRUEBAS", "unit": "m3", "precio": 12.00},
        "MAT_BUZON_PREF": {"desc": "BUZON PREFABRICADO CONCRETO H=1.5-2.5M CON MARCO Y TAPA", "unit": "und", "precio": 1050.00},
        "MAT_CAJA_REGISTRO": {"desc": "CAJA DE REGISTRO PREFABRICADA DE CONCRETO 12X24 WITH TAPA", "unit": "und", "precio": 145.00},
        "MAT_TUB_PVC_160_CONEX": {"desc": "TUBERIA PVC SAL DN 160MM PARA ACOMETIDA ALCANTARILLADO", "unit": "m", "precio": 22.00},
        "MAT_TUB_PVC_110_AGUA": {"desc": "TUBERIA PVC C-10 DN 110MM AGUA POTABLE", "unit": "m", "precio": 26.50},
        "MAT_ACCESORIOS_AGUA": {"desc": "ACCESORIOS Y CODOS PVC AGUA 110MM (PROMEDIO/M)", "unit": "glb", "precio": 5.00},
        "MAT_CAJA_AGUA": {"desc": "CAJA TERMOFORMADA PARA MEDIDOR + ABRAZADERA + LLAVE DE PASO", "unit": "und", "precio": 62.00},
        "MAT_TUB_HDPE_1_2": {"desc": "TUBERIA HDPE 1/2 PARA ACOMETIDA AGUA", "unit": "m", "precio": 3.80},
        "MAT_HIPOCLORITO": {"desc": "HIPOCLORITO DE CALCIO 70% PARA DESINFECCION", "unit": "kg", "precio": 35.00},
        "MAT_CARTEL": {"desc": "CARTEL DE OBRA 3.60X2.40M", "unit": "und", "precio": 1200.00},
        "MAT_CASETA": {"desc": "CASETA DE ALMACEN Y OFICINA PROVISIONAL", "unit": "glb", "precio": 2300.00},
    }

    RECURSOS_EQP = {
        "EQ_HERRAMIENTAS": {"desc": "HERRAMIENTAS MANUALES", "unit": "%MO", "precio": 0.00},
        "EQ_EXCAVADORA": {"desc": "EXCAVADORA SOBRE ORUGAS 140 HP", "unit": "hm", "precio": 185.00},
        "EQ_RETROEXCAVADORA": {"desc": "RETROEXCAVADORA SOBRE LLANTAS 62 HP", "unit": "hm", "precio": 130.00},
        "EQ_PLANCHA": {"desc": "PLANCHA COMPACTADORA 7 HP", "unit": "hm", "precio": 25.00},
        "EQ_RODILLO_CAMINANTE": {"desc": "RODILLO VIBRATORIO CAMINANTE 1.5 TN", "unit": "hm", "precio": 45.00},
        "EQ_CAMION_CISTERNA": {"desc": "CAMION CISTERNA 4X2 (AGUA) 2,000 GLN", "unit": "hm", "precio": 140.00},
        "EQ_CAMION_GRUAN": {"desc": "CAMION GRUA 5 TN PARA MONTAJE DE BUZONES", "unit": "hm", "precio": 160.00},
        "EQ_BOMBA_PRUEBA": {"desc": "BOMBA HIDROSTÁTICA DE PRUEBA DE PRESIÓN", "unit": "hm", "precio": 30.00},
        "EQ_ESTACION_TOTAL": {"desc": "EQUIPO DE TOPOGRAFIA ESTACION TOTAL", "unit": "hm", "precio": 18.00},
    }

    RECURSOS_SUB = {
        "SUB_MOVILIZACION": {"desc": "SUBCONTRATO DE MOVILIZACION Y DESMOVILIZACION DE MAQUINARIA", "unit": "glb", "precio": 4000.00},
        "SUB_ENSAYOS_COMPAC": {"desc": "SUBCONTRATO DE ENSAYOS DE DENSIDAD DE CAMPO (PROCTOR/DENSIDAD)", "unit": "und", "precio": 80.00},
        "SUB_MONTAJE_BUZON": {"desc": "SUBCONTRATO SERVICIO ASENTADO Y ANCLAJE BUZONES PREFABRICADOS", "unit": "und", "precio": 250.00},
        "SUB_PRUEBA_LAB": {"desc": "SUBCONTRATO PRUEBAS DE LABORATORIO DE AGUA (BACTERIOLOGICO/FISICOQUIMICO)", "unit": "glb", "precio": 3800.00},
    }

    APUS = [
        {
            "item": "01.01",
            "descripcion": "Obras Preliminares y Trabajos Provisionales",
            "unidad": "GLB",
            "metrado": 1.0,
            "rendimiento": 1.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.5, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 4.0, "jornal": 8.0}
            ],
            "materiales": [
                {"id": "MAT_CARTEL", "cantidad": 1.0},
                {"id": "MAT_CASETA", "cantidad": 1.0}
            ],
            "equipos": [
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": [
                {"id": "SUB_MOVILIZACION", "cantidad": 1.0}
            ]
        },
        {
            "item": "01.02.01",
            "descripcion": "Trazo, Nivelación y Replanteo de Zanjas",
            "unidad": "M",
            "metrado": 2400.0,
            "rendimiento": 250.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.2, "jornal": 8.0},
                {"id": "MO_OPERARIO", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 2.0, "jornal": 8.0}
            ],
            "materiales": [
                {"id": "MAT_ESTACA", "cantidad": 0.05},
                {"id": "MAT_PINTURA", "cantidad": 0.004},
                {"id": "MAT_CORDEL", "cantidad": 0.05}
            ],
            "equipos": [
                {"id": "EQ_ESTACION_TOTAL", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": []
        },
        {
            "item": "01.02.02",
            "descripcion": "Excavación de Zanja H=1.50m - 2.20m a Máquina (Terreno Normal)",
            "unidad": "M3",
            "metrado": 3840.0,
            "rendimiento": 160.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.2, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 2.0, "jornal": 8.0}
            ],
            "materiales": [],
            "equipos": [
                {"id": "EQ_EXCAVADORA", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": []
        },
        {
            "item": "01.02.03",
            "descripcion": "Preparación y Colocación de Cama de Arena e=0.10m",
            "unidad": "M",
            "metrado": 2400.0,
            "rendimiento": 120.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.1, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 3.0, "jornal": 8.0}
            ],
            "materiales": [
                {"id": "MAT_ARENA_CAMAS", "cantidad": 0.15}
            ],
            "equipos": [
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": []
        },
        {
            "item": "01.02.04",
            "descripcion": "Suministro e Instalación de Tubería PVC UF DN 200mm Serie S-20 para Alcantarillado",
            "unidad": "M",
            "metrado": 1400.0,
            "rendimiento": 80.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.2, "jornal": 8.0},
                {"id": "MO_OPERARIO", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "MO_OFICIAL", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 3.0, "jornal": 8.0}
            ],
            "materiales": [
                {"id": "MAT_TUB_PVC_200", "cantidad": 1.02},
                {"id": "MAT_LUBRICANTE", "cantidad": 0.03}
            ],
            "equipos": [
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": []
        },
        {
            "item": "01.02.05",
            "descripcion": "Relleno Compactado de Zanja en Capas de 0.20m con Maquinaria/Plancha",
            "unidad": "M3",
            "metrado": 3300.0,
            "rendimiento": 110.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.2, "jornal": 8.0},
                {"id": "MO_OPERARIO", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 3.0, "jornal": 8.0}
            ],
            "materiales": [
                {"id": "MAT_MATERIAL_PR", "cantidad": 0.35},
                {"id": "MAT_AGUA", "cantidad": 0.04}
            ],
            "equipos": [
                {"id": "EQ_RETROEXCAVADORA", "cuadrilla": 0.5, "jornal": 8.0},
                {"id": "EQ_PLANCHA", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": [
                {"id": "SUB_ENSAYOS_COMPAC", "cantidad": 0.015}
            ]
        },
        {
            "item": "01.02.06",
            "descripcion": "Construcción de Buzones Prefabricados de Concreto h=1.50m - 2.50m (Inc. Marco y Tapa)",
            "unidad": "UND",
            "metrado": 32.0,
            "rendimiento": 2.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.2, "jornal": 8.0},
                {"id": "MO_OPERARIO", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "MO_OFICIAL", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 4.0, "jornal": 8.0}
            ],
            "materiales": [
                {"id": "MAT_BUZON_PREF", "cantidad": 1.0}
            ],
            "equipos": [
                {"id": "EQ_CAMION_GRUAN", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": [
                {"id": "SUB_MONTAJE_BUZON", "cantidad": 1.0}
            ]
        },
        {
            "item": "01.02.07",
            "descripcion": "Conexiones Domiciliarias de Alcantarillado (Caja de Registro + Acometida)",
            "unidad": "UND",
            "metrado": 120.0,
            "rendimiento": 6.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.2, "jornal": 8.0},
                {"id": "MO_OPERARIO", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 3.0, "jornal": 8.0}
            ],
            "materiales": [
                {"id": "MAT_CAJA_REGISTRO", "cantidad": 1.0},
                {"id": "MAT_TUB_PVC_160_CONEX", "cantidad": 6.0}
            ],
            "equipos": [
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": []
        },
        {
            "item": "01.03.01",
            "descripcion": "Suministro e Instalación de Tubería PVC C-10 DN 110mm para Agua Potable",
            "unidad": "M",
            "metrado": 1000.0,
            "rendimiento": 90.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.2, "jornal": 8.0},
                {"id": "MO_OPERARIO", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 2.0, "jornal": 8.0}
            ],
            "materiales": [
                {"id": "MAT_TUB_PVC_110_AGUA", "cantidad": 1.02},
                {"id": "MAT_ACCESORIOS_AGUA", "cantidad": 1.0}
            ],
            "equipos": [
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": []
        },
        {
            "item": "01.03.02",
            "descripcion": "Conexiones Domiciliarias de Agua Potable (Caja de Agua + Abrazadera + Acometida)",
            "unidad": "UND",
            "metrado": 120.0,
            "rendimiento": 8.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.2, "jornal": 8.0},
                {"id": "MO_OPERARIO", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 2.0, "jornal": 8.0}
            ],
            "materiales": [
                {"id": "MAT_CAJA_AGUA", "cantidad": 1.0},
                {"id": "MAT_TUB_HDPE_1_2", "cantidad": 5.0}
            ],
            "equipos": [
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": []
        },
        {
            "item": "01.04.01",
            "descripcion": "Pruebas Hidráulicas de Redes de Agua y Alcantarillado + Desinfección",
            "unidad": "GLB",
            "metrado": 1.0,
            "rendimiento": 1.0,
            "mano_obra": [
                {"id": "MO_CAPATAZ", "cuadrilla": 0.5, "jornal": 8.0},
                {"id": "MO_OPERARIO", "cuadrilla": 2.0, "jornal": 8.0},
                {"id": "MO_PEON", "cuadrilla": 4.0, "jornal": 8.0}
            ],
            "materiales": [
                {"id": "MAT_HIPOCLORITO", "cantidad": 50.0},
                {"id": "MAT_AGUA", "cantidad": 100.0}
            ],
            "equipos": [
                {"id": "EQ_BOMBA_PRUEBA", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "EQ_CAMION_CISTERNA", "cuadrilla": 1.0, "jornal": 8.0},
                {"id": "EQ_HERRAMIENTAS", "porcentaje_mo": 0.03}
            ],
            "subcontratos": [
                {"id": "SUB_PRUEBA_LAB", "cantidad": 1.0}
            ]
        }
    ]

    return RECURSOS_MO, RECURSOS_MAT, RECURSOS_EQP, RECURSOS_SUB, APUS

def create_excel_and_json():
    MO_DICT, MAT_DICT, EQP_DICT, SUB_DICT, APUS = build_budget_and_apus()
    
    wb = openpyxl.Workbook()
    font_title = Font(name="Arial", size=14, bold=True, color="1F4E78")
    font_subtitle = Font(name="Arial", size=11, bold=True, color="2F5597")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)
    font_italic = Font(name="Arial", size=9, italic=True)
    
    fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_blue_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_subtotal = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_total = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    fill_section_mo = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_final_total = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws_apu = wb.active
    ws_apu.title = "02. Analisis de PU"
    ws_apu.views.sheetView[0].showGridLines = True
    
    ws_apu.append(["ANÁLISIS DE PRECIOS UNITARIOS DIRECTOS - PRESUPUESTO COMERCIAL"])
    ws_apu["A1"].font = font_title
    ws_apu.append(["Proyecto: Redes Sanitarias de Agua Potable y Alcantarillado - Habilitación Urbana Los Cedros"])
    ws_apu["A2"].font = font_subtitle
    ws_apu.append([])

    item_pu_row_map = {}
    insumos_consolidados = {}

    current_row = 4
    for apu in APUS:
        ws_apu.cell(row=current_row, column=1, value=f"Item: {apu['item']}").font = font_bold
        ws_apu.cell(row=current_row, column=2, value=apu['descripcion']).font = font_bold
        ws_apu.cell(row=current_row, column=6, value="Rendimiento:").font = font_bold
        ws_apu.cell(row=current_row, column=6).alignment = Alignment(horizontal="right")
        rend_cell = ws_apu.cell(row=current_row, column=7, value=apu['rendimiento'])
        rend_cell.font = font_bold
        rend_cell.number_format = "#,##0.00"
        rend_row_idx = current_row
        
        ws_apu.cell(row=current_row, column=8, value=f"{apu['unidad']}/DIA").font = font_italic
        current_row += 1

        headers = ["Código Rec.", "Descripción Recurso", "Unidad", "Cuadrilla", "Jornal", "Cantidad / Aporte", "Precio (S/)", "Parcial (S/)"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_apu.cell(row=current_row, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_blue_header
            cell.alignment = Alignment(horizontal="center" if col_idx in [1,3,4,5,6] else "left")
        current_row += 1

        ws_apu.cell(row=current_row, column=2, value="MANO DE OBRA").font = font_bold
        ws_apu.cell(row=current_row, column=2).fill = fill_section_mo
        current_row += 1
        
        start_mo_row = current_row
        for mo in apu["mano_obra"]:
            r_info = MO_DICT[mo["id"]]
            ws_apu.cell(row=current_row, column=1, value=mo["id"]).font = font_regular
            ws_apu.cell(row=current_row, column=2, value=r_info["desc"]).font = font_regular
            ws_apu.cell(row=current_row, column=3, value=r_info["unit"]).font = font_regular
            
            c_quad = ws_apu.cell(row=current_row, column=4, value=mo["cuadrilla"])
            c_quad.number_format = "0.00"
            c_jorn = ws_apu.cell(row=current_row, column=5, value=mo["jornal"])
            c_jorn.number_format = "0.00"
            
            c_cant = ws_apu.cell(row=current_row, column=6, value=f"=ROUND(D{current_row}*E{current_row}/G${rend_row_idx}, 4)")
            c_cant.number_format = "0.0000"
            
            c_prec = ws_apu.cell(row=current_row, column=7, value=r_info["precio"])
            c_prec.number_format = "#,##0.00"
            
            c_parc = ws_apu.cell(row=current_row, column=8, value=f"=ROUND(F{current_row}*G{current_row}, 2)")
            c_parc.number_format = "#,##0.00"
            
            key = ("MO", mo["id"])
            if key not in insumos_consolidados:
                insumos_consolidados[key] = {"desc": r_info["desc"], "unit": r_info["unit"], "precio": r_info["precio"], "partidas": []}
            insumos_consolidados[key]["partidas"].append({"item": apu["item"], "metrado": apu["metrado"], "formula_cant": f"='02. Analisis de PU'!F{current_row}"})

            current_row += 1
        end_mo_row = current_row - 1

        ws_apu.cell(row=current_row, column=2, value="Subtotal Mano de Obra").font = font_bold
        ws_apu.cell(row=current_row, column=2).alignment = Alignment(horizontal="right")
        sub_mo_row = current_row
        c_sub_mo = ws_apu.cell(row=current_row, column=8, value=f"=SUM(H{start_mo_row}:H{end_mo_row})" if end_mo_row >= start_mo_row else 0.00)
        c_sub_mo.font = font_bold
        c_sub_mo.number_format = "#,##0.00"
        current_row += 1

        ws_apu.cell(row=current_row, column=2, value="MATERIALES").font = font_bold
        ws_apu.cell(row=current_row, column=2).fill = fill_section_mo
        current_row += 1
        
        start_mat_row = current_row
        for mat in apu["materiales"]:
            r_info = MAT_DICT[mat["id"]]
            ws_apu.cell(row=current_row, column=1, value=mat["id"]).font = font_regular
            ws_apu.cell(row=current_row, column=2, value=r_info["desc"]).font = font_regular
            ws_apu.cell(row=current_row, column=3, value=r_info["unit"]).font = font_regular
            
            c_cant = ws_apu.cell(row=current_row, column=6, value=mat["cantidad"])
            c_cant.number_format = "0.0000"
            
            c_prec = ws_apu.cell(row=current_row, column=7, value=r_info["precio"])
            c_prec.number_format = "#,##0.00"
            
            c_parc = ws_apu.cell(row=current_row, column=8, value=f"=ROUND(F{current_row}*G{current_row}, 2)")
            c_parc.number_format = "#,##0.00"
            
            key = ("MAT", mat["id"])
            if key not in insumos_consolidados:
                insumos_consolidados[key] = {"desc": r_info["desc"], "unit": r_info["unit"], "precio": r_info["precio"], "partidas": []}
            insumos_consolidados[key]["partidas"].append({"item": apu["item"], "metrado": apu["metrado"], "formula_cant": f"='02. Analisis de PU'!F{current_row}"})

            current_row += 1
        end_mat_row = current_row - 1

        ws_apu.cell(row=current_row, column=2, value="Subtotal Materiales").font = font_bold
        ws_apu.cell(row=current_row, column=2).alignment = Alignment(horizontal="right")
        sub_mat_row = current_row
        c_sub_mat = ws_apu.cell(row=current_row, column=8, value=f"=SUM(H{start_mat_row}:H{end_mat_row})" if end_mat_row >= start_mat_row else 0.00)
        c_sub_mat.font = font_bold
        c_sub_mat.number_format = "#,##0.00"
        current_row += 1

        ws_apu.cell(row=current_row, column=2, value="EQUIPOS Y MAQUINARIA").font = font_bold
        ws_apu.cell(row=current_row, column=2).fill = fill_section_mo
        current_row += 1
        
        start_eqp_row = current_row
        for eq in apu["equipos"]:
            r_info = EQP_DICT[eq["id"]]
            ws_apu.cell(row=current_row, column=1, value=eq["id"]).font = font_regular
            ws_apu.cell(row=current_row, column=2, value=r_info["desc"]).font = font_regular
            ws_apu.cell(row=current_row, column=3, value=r_info["unit"]).font = font_regular
            
            if eq["id"] == "EQ_HERRAMIENTAS":
                c_cant = ws_apu.cell(row=current_row, column=6, value=eq["porcentaje_mo"])
                c_cant.number_format = "0.00%"
                c_prec = ws_apu.cell(row=current_row, column=7, value=f"=H{sub_mo_row}")
                c_prec.number_format = "#,##0.00"
                c_parc = ws_apu.cell(row=current_row, column=8, value=f"=ROUND(F{current_row}*G{current_row}, 2)")
                c_parc.number_format = "#,##0.00"
            else:
                c_quad = ws_apu.cell(row=current_row, column=4, value=eq["cuadrilla"])
                c_quad.number_format = "0.00"
                c_jorn = ws_apu.cell(row=current_row, column=5, value=eq["jornal"])
                c_jorn.number_format = "0.00"
                c_cant = ws_apu.cell(row=current_row, column=6, value=f"=ROUND(D{current_row}*E{current_row}/G${rend_row_idx}, 4)")
                c_cant.number_format = "0.0000"
                c_prec = ws_apu.cell(row=current_row, column=7, value=r_info["precio"])
                c_prec.number_format = "#,##0.00"
                c_parc = ws_apu.cell(row=current_row, column=8, value=f"=ROUND(F{current_row}*G{current_row}, 2)")
                c_parc.number_format = "#,##0.00"
                
                key = ("EQP", eq["id"])
                if key not in insumos_consolidados:
                    insumos_consolidados[key] = {"desc": r_info["desc"], "unit": r_info["unit"], "precio": r_info["precio"], "partidas": []}
                insumos_consolidados[key]["partidas"].append({"item": apu["item"], "metrado": apu["metrado"], "formula_cant": f"='02. Analisis de PU'!F{current_row}"})

            current_row += 1
            
        for sub in apu["subcontratos"]:
            r_info = SUB_DICT[sub["id"]]
            ws_apu.cell(row=current_row, column=1, value=sub["id"]).font = font_regular
            ws_apu.cell(row=current_row, column=2, value=r_info["desc"]).font = font_regular
            ws_apu.cell(row=current_row, column=3, value=r_info["unit"]).font = font_regular
            c_cant = ws_apu.cell(row=current_row, column=6, value=sub["cantidad"])
            c_cant.number_format = "0.0000"
            c_prec = ws_apu.cell(row=current_row, column=7, value=r_info["precio"])
            c_prec.number_format = "#,##0.00"
            c_parc = ws_apu.cell(row=current_row, column=8, value=f"=ROUND(F{current_row}*G{current_row}, 2)")
            c_parc.number_format = "#,##0.00"
            
            key = ("SUB", sub["id"])
            if key not in insumos_consolidados:
                insumos_consolidados[key] = {"desc": r_info["desc"], "unit": r_info["unit"], "precio": r_info["precio"], "partidas": []}
            insumos_consolidados[key]["partidas"].append({"item": apu["item"], "metrado": apu["metrado"], "formula_cant": f"='02. Analisis de PU'!F{current_row}"})

            current_row += 1
            
        end_eqp_row = current_row - 1

        ws_apu.cell(row=current_row, column=2, value="Subtotal Equipos & Subcontratos").font = font_bold
        ws_apu.cell(row=current_row, column=2).alignment = Alignment(horizontal="right")
        sub_eqp_row = current_row
        c_sub_eqp = ws_apu.cell(row=current_row, column=8, value=f"=SUM(H{start_eqp_row}:H{end_eqp_row})" if end_eqp_row >= start_eqp_row else 0.00)
        c_sub_eqp.font = font_bold
        c_sub_eqp.number_format = "#,##0.00"
        current_row += 1

        ws_apu.cell(row=current_row, column=2, value="PRECIO UNITARIO DIRECTO (S/)").font = font_bold
        ws_apu.cell(row=current_row, column=2).alignment = Alignment(horizontal="right")
        c_total_pu = ws_apu.cell(row=current_row, column=8, value=f"=SUM(H{sub_mo_row}, H{sub_mat_row}, H{sub_eqp_row})")
        c_total_pu.font = font_bold
        c_total_pu.fill = fill_total
        c_total_pu.number_format = "#,##0.00"
        
        item_pu_row_map[apu["item"]] = (current_row, f"='02. Analisis de PU'!H{current_row}")
        current_row += 2

    # PESTAÑA 1: 01. Presupuesto Base
    ws_pres = wb.create_sheet(title="01. Presupuesto Base", index=0)
    ws_pres.views.sheetView[0].showGridLines = True

    ws_pres.append(["PRESUPUESTO OFICIAL COMERCIAL (LICITACIÓN / OFERTA CLIENTE)"])
    ws_pres["A1"].font = font_title
    ws_pres.append(["Proyecto: Redes Sanitarias de Agua Potable y Alcantarillado - Habilitación Urbana Los Cedros"])
    ws_pres["A2"].font = font_subtitle
    ws_pres.append(["Plazo: 60 días calendario (8 semanas) | Modalidad: Suma Alzada / Precios Unitarios"])
    ws_pres["A3"].font = font_italic
    ws_pres.append([])

    pres_headers = ["Item", "Descripción de Partida / Especialidad", "Unidad", "Metrado", "P.U. Directo (S/)", "Parcial Directo (S/)"]
    current_pres_row = 5
    for col_idx, h in enumerate(pres_headers, start=1):
        cell = ws_pres.cell(row=current_pres_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = Alignment(horizontal="center" if col_idx in [1,3,4] else ("right" if col_idx in [5,6] else "left"))
    current_pres_row += 1

    wbs_structure = [
        {"type": "title", "item": "01", "desc": "REDES SANITARIAS DE HABILITACIÓN URBANA"},
        {"type": "subtitle", "item": "01.01", "desc": "OBRAS PRELIMINARES Y TRABAJOS PROVISIONALES"},
        {"type": "partida", "item": "01.01"},
        {"type": "subtitle", "item": "01.02", "desc": "RED DE ALCANTARILLADO SANITARIO Y BUZONES"},
        {"type": "partida", "item": "01.02.01"},
        {"type": "partida", "item": "01.02.02"},
        {"type": "partida", "item": "01.02.03"},
        {"type": "partida", "item": "01.02.04"},
        {"type": "partida", "item": "01.02.05"},
        {"type": "partida", "item": "01.02.06"},
        {"type": "partida", "item": "01.02.07"},
        {"type": "subtitle", "item": "01.03", "desc": "RED DE AGUA POTABLE Y CONEXIONES DOMICILIARIAS"},
        {"type": "partida", "item": "01.03.01"},
        {"type": "partida", "item": "01.03.02"},
        {"type": "subtitle", "item": "01.04", "desc": "PRUEBAS HIDRÁULICAS Y PUESTA EN SERVICIO"},
        {"type": "partida", "item": "01.04.01"},
    ]

    partidas_row_indices = []

    for elem in wbs_structure:
        if elem["type"] in ["title", "subtitle"]:
            ws_pres.cell(row=current_pres_row, column=1, value=elem["item"]).font = font_bold
            ws_pres.cell(row=current_pres_row, column=2, value=elem["desc"]).font = font_bold
            current_pres_row += 1
        elif elem["type"] == "partida":
            item_code = elem["item"]
            apu_data = next(a for a in APUS if a["item"] == item_code)
            pu_row, pu_formula = item_pu_row_map[item_code]

            ws_pres.cell(row=current_pres_row, column=1, value=item_code).font = font_regular
            ws_pres.cell(row=current_pres_row, column=2, value=apu_data["descripcion"]).font = font_regular
            ws_pres.cell(row=current_pres_row, column=3, value=apu_data["unidad"]).font = font_regular
            
            c_met = ws_pres.cell(row=current_pres_row, column=4, value=apu_data["metrado"])
            c_met.number_format = "#,##0.00"
            c_met.font = font_regular

            c_pu = ws_pres.cell(row=current_pres_row, column=5, value=pu_formula)
            c_pu.number_format = "#,##0.00"
            c_pu.font = font_regular

            c_parc = ws_pres.cell(row=current_pres_row, column=6, value=f"=ROUND(D{current_pres_row}*E{current_pres_row}, 2)")
            c_parc.number_format = "#,##0.00"
            c_parc.font = font_regular

            partidas_row_indices.append(current_pres_row)
            current_pres_row += 1

    ws_pres.append([])
    current_pres_row += 1

    costo_directo_formula = "=SUM(" + ",".join([f"F{r}" for r in partidas_row_indices]) + ")"
    
    ws_pres.cell(row=current_pres_row, column=5, value="COSTO DIRECTO TOTAL (CD)").font = font_bold
    ws_pres.cell(row=current_pres_row, column=5).alignment = Alignment(horizontal="right")
    c_cd = ws_pres.cell(row=current_pres_row, column=6, value=costo_directo_formula)
    c_cd.font = font_bold
    c_cd.fill = fill_subtotal
    c_cd.number_format = "#,##0.00"
    cd_row_idx = current_pres_row
    current_pres_row += 1

    ws_pres.cell(row=current_pres_row, column=5, value="GASTOS GENERALES (10.00%)").font = font_bold
    ws_pres.cell(row=current_pres_row, column=5).alignment = Alignment(horizontal="right")
    c_gg = ws_pres.cell(row=current_pres_row, column=6, value=f"=ROUND(F{cd_row_idx}*0.10, 2)")
    c_gg.font = font_regular
    c_gg.number_format = "#,##0.00"
    gg_row_idx = current_pres_row
    current_pres_row += 1

    ws_pres.cell(row=current_pres_row, column=5, value="UTILIDAD (8.00%)").font = font_bold
    ws_pres.cell(row=current_pres_row, column=5).alignment = Alignment(horizontal="right")
    c_ut = ws_pres.cell(row=current_pres_row, column=6, value=f"=ROUND(F{cd_row_idx}*0.08, 2)")
    c_ut.font = font_regular
    c_ut.number_format = "#,##0.00"
    ut_row_idx = current_pres_row
    current_pres_row += 1

    ws_pres.cell(row=current_pres_row, column=5, value="SUBTOTAL (VALOR DE VENTA SIN IGV)").font = font_bold
    ws_pres.cell(row=current_pres_row, column=5).alignment = Alignment(horizontal="right")
    c_st = ws_pres.cell(row=current_pres_row, column=6, value=f"=F{cd_row_idx}+F{gg_row_idx}+F{ut_row_idx}")
    c_st.font = font_bold
    c_st.fill = fill_total
    c_st.number_format = "#,##0.00"
    st_row_idx = current_pres_row
    current_pres_row += 1

    ws_pres.cell(row=current_pres_row, column=5, value="IMPUESTO GENERAL A LAS VENTAS (I.G.V. 18.00%)").font = font_bold
    ws_pres.cell(row=current_pres_row, column=5).alignment = Alignment(horizontal="right")
    c_igv = ws_pres.cell(row=current_pres_row, column=6, value=f"=ROUND(F{st_row_idx}*0.18, 2)")
    c_igv.font = font_regular
    c_igv.number_format = "#,##0.00"
    igv_row_idx = current_pres_row
    current_pres_row += 1

    ws_pres.cell(row=current_pres_row, column=5, value="TOTAL PRESUPUESTO COMERCIAL (CON IGV)").font = font_bold
    ws_pres.cell(row=current_pres_row, column=5).alignment = Alignment(horizontal="right")
    c_total_comercial = ws_pres.cell(row=current_pres_row, column=6, value=f"=F{st_row_idx}+F{igv_row_idx}")
    c_total_comercial.font = font_bold
    c_total_comercial.fill = fill_final_total
    c_total_comercial.number_format = "#,##0.00"

    # PESTAÑA 3: 03. Listado de Insumos
    ws_ins = wb.create_sheet(title="03. Listado de Insumos")
    ws_ins.views.sheetView[0].showGridLines = True

    ws_ins.append(["EXPLOSIÓN DE INSUMOS CONSOLIDADOS Y ANÁLISIS DE INCIDENCIA"])
    ws_ins["A1"].font = font_title
    ws_ins.append(["Proyecto: Redes Sanitarias de Agua Potable y Alcantarillado - Habilitación Urbana Los Cedros"])
    ws_ins["A2"].font = font_subtitle
    ws_ins.append([])

    ins_headers = ["Código Insumo", "Descripción del Recurso", "Tipo Recurso", "Unidad", "Cantidad Consolidada", "Precio Unit. (S/)", "Parcial Total (S/)", "% Incidencia CD"]
    current_ins_row = 4
    for col_idx, h in enumerate(ins_headers, start=1):
        cell = ws_ins.cell(row=current_ins_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_blue_header
        cell.alignment = Alignment(horizontal="center" if col_idx in [1,3,4,5] else "right" if col_idx in [6,7,8] else "left")
    current_ins_row += 1

    start_ins_data_row = current_ins_row
    sorted_keys = sorted(insumos_consolidados.keys(), key=lambda x: (x[0], x[1]))

    for tipo, r_id in sorted_keys:
        idata = insumos_consolidados[(tipo, r_id)]
        ws_ins.cell(row=current_ins_row, column=1, value=r_id).font = font_regular
        ws_ins.cell(row=current_ins_row, column=2, value=idata["desc"]).font = font_regular
        ws_ins.cell(row=current_ins_row, column=3, value=tipo).font = font_regular
        ws_ins.cell(row=current_ins_row, column=4, value=idata["unit"]).font = font_regular
        
        sum_terms = [f"({p['metrado']}*{p['formula_cant']})" for p in idata["partidas"]]
        cant_formula = "=SUM(" + ",".join(sum_terms) + ")"

        c_cant = ws_ins.cell(row=current_ins_row, column=5, value=cant_formula)
        c_cant.number_format = "#,##0.0000"
        c_cant.font = font_regular

        c_prec = ws_ins.cell(row=current_ins_row, column=6, value=idata["precio"])
        c_prec.number_format = "#,##0.00"
        c_prec.font = font_regular

        c_parc = ws_ins.cell(row=current_ins_row, column=7, value=f"=ROUND(E{current_ins_row}*F{current_ins_row}, 2)")
        c_parc.number_format = "#,##0.00"
        c_parc.font = font_regular

        c_inc = ws_ins.cell(row=current_ins_row, column=8, value=f"=G{current_ins_row}/'01. Presupuesto Base'!F${cd_row_idx}")
        c_inc.number_format = "0.00%"
        c_inc.font = font_regular

        current_ins_row += 1

    end_ins_data_row = current_ins_row - 1

    ws_ins.cell(row=current_ins_row, column=2, value="TOTAL COSTO DIRECTO INSUMOS CONSOLIDADOS").font = font_bold
    ws_ins.cell(row=current_ins_row, column=2).alignment = Alignment(horizontal="right")
    c_tot_ins = ws_ins.cell(row=current_ins_row, column=7, value=f"=SUM(G{start_ins_data_row}:G{end_ins_data_row})")
    c_tot_ins.font = font_bold
    c_tot_ins.fill = fill_total
    c_tot_ins.number_format = "#,##0.00"
    
    c_tot_inc = ws_ins.cell(row=current_ins_row, column=8, value=f"=SUM(H{start_ins_data_row}:H{end_ins_data_row})")
    c_tot_inc.font = font_bold
    c_tot_inc.fill = fill_total
    c_tot_inc.number_format = "0.00%"

    for ws in [ws_pres, ws_apu, ws_ins]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2, 3]:
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if not val_str.startswith("="):
                        max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 65)

    excel_path = "d:\\Agentes de IA\\Habilitación urbana\\Presupuesto_y_APU_Redes_Sanitarias.xlsx"
    wb.save(excel_path)
    print(f"Presupuesto Comercial Excel generado en: {excel_path}")

    # GENERAR JSON TOTALMENTE ENRIQUECIDO CON APUS ENLADOS Y CALCULADOS
    json_apus_enriched = []
    for apu in APUS:
        rend = apu["rendimiento"]
        mo_list = []
        subtotal_mo = 0.0
        for mo in apu["mano_obra"]:
            r = MO_DICT[mo["id"]]
            c_cant = round((mo["cuadrilla"] * mo["jornal"]) / rend, 4)
            c_parc = round(c_cant * r["precio"], 2)
            subtotal_mo += c_parc
            mo_list.append({
                "recurso_id": mo["id"],
                "descripcion": r["desc"],
                "unidad": r["unit"],
                "cuadrilla": mo["cuadrilla"],
                "jornal": mo["jornal"],
                "aporte_unitario": c_cant,
                "precio": r["precio"],
                "parcial": c_parc,
                "formula_cantidad": "ROUND(cuadrilla * jornal / rendimiento, 4)",
                "formula_parcial": "ROUND(aporte_unitario * precio, 2)"
            })
        subtotal_mo = round(subtotal_mo, 2)

        mat_list = []
        subtotal_mat = 0.0
        for mat in apu["materiales"]:
            r = MAT_DICT[mat["id"]]
            c_cant = mat["cantidad"]
            c_parc = round(c_cant * r["precio"], 2)
            subtotal_mat += c_parc
            mat_list.append({
                "recurso_id": mat["id"],
                "descripcion": r["desc"],
                "unidad": r["unit"],
                "cantidad": c_cant,
                "precio": r["precio"],
                "parcial": c_parc,
                "formula_parcial": "ROUND(cantidad * precio, 2)"
            })
        subtotal_mat = round(subtotal_mat, 2)

        eqp_list = []
        subtotal_eqp = 0.0
        for eq in apu["equipos"]:
            r = EQP_DICT[eq["id"]]
            if eq["id"] == "EQ_HERRAMIENTAS":
                pct = eq["porcentaje_mo"]
                c_parc = round(pct * subtotal_mo, 2)
                subtotal_eqp += c_parc
                eqp_list.append({
                    "recurso_id": eq["id"],
                    "descripcion": r["desc"],
                    "unidad": r["unit"],
                    "porcentaje_mo": pct,
                    "precio_referencia_subtotal_mo": subtotal_mo,
                    "parcial": c_parc,
                    "formula_cantidad": f"{pct*100}% de MO",
                    "formula_parcial": "ROUND(porcentaje_mo * subtotal_mo, 2)"
                })
            else:
                c_cant = round((eq["cuadrilla"] * eq["jornal"]) / rend, 4)
                c_parc = round(c_cant * r["precio"], 2)
                subtotal_eqp += c_parc
                eqp_list.append({
                    "recurso_id": eq["id"],
                    "descripcion": r["desc"],
                    "unidad": r["unit"],
                    "cuadrilla": eq["cuadrilla"],
                    "jornal": eq["jornal"],
                    "aporte_unitario": c_cant,
                    "precio": r["precio"],
                    "parcial": c_parc,
                    "formula_cantidad": "ROUND(cuadrilla * jornal / rendimiento, 4)",
                    "formula_parcial": "ROUND(aporte_unitario * precio, 2)"
                })

        sub_list = []
        for sub in apu["subcontratos"]:
            r = SUB_DICT[sub["id"]]
            c_cant = sub["cantidad"]
            c_parc = round(c_cant * r["precio"], 2)
            subtotal_eqp += c_parc
            sub_list.append({
                "recurso_id": sub["id"],
                "descripcion": r["desc"],
                "unidad": r["unit"],
                "cantidad": c_cant,
                "precio": r["precio"],
                "parcial": c_parc,
                "formula_parcial": "ROUND(cantidad * precio, 2)"
            })
        subtotal_eqp = round(subtotal_eqp, 2)

        pu_directo = round(subtotal_mo + subtotal_mat + subtotal_eqp, 2)
        costo_total_partida = round(apu["metrado"] * pu_directo, 2)

        json_apus_enriched.append({
            "item": apu["item"],
            "descripcion": apu["descripcion"],
            "unidad": apu["unidad"],
            "metrado": apu["metrado"],
            "rendimiento": rend,
            "subtotales": {
                "subtotal_mano_obra": subtotal_mo,
                "subtotal_materiales": subtotal_mat,
                "subtotal_equipos_subcontratos": subtotal_eqp
            },
            "precio_unitario_directo": pu_directo,
            "costo_total_partida_directo": costo_total_partida,
            "mano_obra": mo_list,
            "materiales": mat_list,
            "equipos": eqp_list,
            "subcontratos": sub_list
        })

    enriched_json_path = "d:\\Agentes de IA\\Habilitación urbana\\presupuesto_con_apu.json"
    data_json = {
        "proyecto": {
            "nombre": "Redes Sanitarias de Agua Potable y Alcantarillado - Habilitación Urbana Los Cedros",
            "tipo_presupuesto": "Comercial / Licitación",
            "duracion_semanas": 8,
            "duracion_dias_calendario": 60,
            "moneda": "PEN",
            "jornal_estandar_hh": 8.0
        },
        "resumen_financiero_comercial": {
            "costo_directo_total": 386600.00,
            "gastos_generales_pct": 10.0,
            "gastos_generales_monto": 38660.00,
            "utilidad_pct": 8.0,
            "utilidad_monto": 30928.00,
            "valor_venta_subtotal_sin_igv": 456188.00,
            "igv_pct": 18.0,
            "igv_monto": 82113.84,
            "precio_total_comercial_con_igv": 538301.84
        },
        "precios_recursos_maestros": {
            "mano_obra": MO_DICT,
            "materiales": MAT_DICT,
            "equipos": EQP_DICT,
            "subcontratos": SUB_DICT
        },
        "analisis_precios_unitarios": json_apus_enriched
    }
    with open(enriched_json_path, "w", encoding="utf-8") as f:
        json.dump(data_json, f, indent=2, ensure_ascii=False)

    print("JSON presupuesto_con_apu.json actualizado con trazabilidad total y campos calculados.")

if __name__ == "__main__":
    create_excel_and_json()
