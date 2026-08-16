import openpyxl
from datetime import datetime, timedelta

def generate_database():
    wb = openpyxl.load_workbook('Base_Datos_Proyecto_Sheets_Viva.xlsx', data_only=False)
    
    # 1. LIMPIAR PESTAÑA 4 (04_LOG_FIELD_ENTRIES)
    sheet4 = wb['04_LOG_FIELD_ENTRIES']
    # Mantener el encabezado (Fila 1) y borrar todas las demás filas
    if sheet4.max_row > 1:
        sheet4.delete_rows(2, sheet4.max_row)
        
    print("Pestaña 04_LOG_FIELD_ENTRIES vaciada. Generando datos sintéticos...")

    # Recursos maestros de Mano de Obra (MO), Almacén (MAT), Equipos (EQP) y Subcontratos (SUB)
    resource_info = {
        "MO_CAPATAZ": {"unidad": "hh", "tipo": "AC_MO", "rol": "Tareador (Bildin)", "orig": "tareador.html", "desc": "CAPATAZ"},
        "MO_OPERARIO": {"unidad": "hh", "tipo": "AC_MO", "rol": "Tareador (Bildin)", "orig": "tareador.html", "desc": "OPERARIO"},
        "MO_OFICIAL": {"unidad": "hh", "tipo": "AC_MO", "rol": "Tareador (Bildin)", "orig": "tareador.html", "desc": "OFICIAL"},
        "MO_PEON": {"unidad": "hh", "tipo": "AC_MO", "rol": "Tareador (Bildin)", "orig": "tareador.html", "desc": "PEON"},
        
        "MAT_ESTACA": {"unidad": "pza", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "ESTACAS Y MADERA PARA TRAZO"},
        "MAT_PINTURA": {"unidad": "gla", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "PINTURA SPRAY MARCADOR ZANJA"},
        "MAT_CORDEL": {"unidad": "m", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "CORDEL / NYLON"},
        "MAT_ARENA_CAMAS": {"unidad": "m3", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "ARENA FINA SELECCIONADA PARA CAMA"},
        "MAT_TUB_PVC_200": {"unidad": "m", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "TUBERIA PVC UF DN 200MM S-20 ALCANTARILLADO"},
        "MAT_LUBRICANTE": {"unidad": "kg", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "LUBRICANTE PARA TUBERIA"},
        "MAT_BUZON_PREF": {"unidad": "und", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "BUZON PREFABRICADO CONCRETO"},
        "MAT_CAJA_REGISTRO": {"unidad": "und", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "CAJA REGISTRO CONCRETO 12X24"},
        "MAT_TUB_PVC_160_CONEX": {"unidad": "m", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "TUBERIA PVC SAL 160MM ACOMETIDA"},
        "MAT_TUB_PVC_110_AGUA": {"unidad": "m", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "TUBERIA PVC C-10 DN 110MM AGUA"},
        "MAT_ACCESORIOS_AGUA": {"unidad": "glb", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "ACCESORIOS PVC AGUA 110MM"},
        "MAT_CAJA_AGUA": {"unidad": "und", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "CAJA TERMOFORMADA AGUA + MEDIDOR"},
        "MAT_TUB_HDPE_1_2": {"unidad": "m", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "TUBERIA HDPE 1/2 ACOMETIDA AGUA"},
        "MAT_HIPOCLORITO": {"unidad": "kg", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "HIPOCLORITO DE CALCIO 70%"},
        "MAT_AGUA": {"unidad": "m3", "tipo": "AC_MAT", "rol": "Almacenero", "orig": "almacenero.html", "desc": "AGUA PARA COMPACTACION/PRUEBAS"},
        
        "EQ_ESTACION_TOTAL": {"unidad": "hm", "tipo": "AC_EQP", "rol": "Administradora", "orig": "administradora.html", "desc": "ESTACION TOTAL TOPOGRAFICA"},
        "EQ_EXCAVADORA": {"unidad": "hm", "tipo": "AC_EQP", "rol": "Administradora", "orig": "administradora.html", "desc": "EXCAVADORA ORUGAS 140 HP"},
        "EQ_RETROEXCAVADORA": {"unidad": "hm", "tipo": "AC_EQP", "rol": "Administradora", "orig": "administradora.html", "desc": "RETROEXCAVADORA LLANTAS 62 HP"},
        "EQ_PLANCHA": {"unidad": "hm", "tipo": "AC_EQP", "rol": "Administradora", "orig": "administradora.html", "desc": "PLANCHA COMPACTADORA 7 HP"},
        "EQ_RODILLO_CAMINANTE": {"unidad": "hm", "tipo": "AC_EQP", "rol": "Administradora", "orig": "administradora.html", "desc": "RODILLO VIBRATORIO 1.5 TN"},
        "EQ_CAMION_CISTERNA": {"unidad": "hm", "tipo": "AC_EQP", "rol": "Administradora", "orig": "administradora.html", "desc": "CAMION CISTERNA AGUA 2000 GLN"},
        "EQ_CAMION_GRUAN": {"unidad": "hm", "tipo": "AC_EQP", "rol": "Administradora", "orig": "administradora.html", "desc": "CAMION GRUA 5 TN MONTAJE BUZONES"},
        "EQ_BOMBA_PRUEBA": {"unidad": "hm", "tipo": "AC_EQP", "rol": "Administradora", "orig": "administradora.html", "desc": "BOMBA HIDROSTATICA PRUEBA PRESION"},
        
        "SUB_MOVILIZACION": {"unidad": "glb", "tipo": "AC_SUB", "rol": "Administradora", "orig": "administradora.html", "desc": "MOVILIZACION MAQUINARIA"},
        "SUB_MONTAJE_BUZON": {"unidad": "und", "tipo": "AC_SUB", "rol": "Administradora", "orig": "administradora.html", "desc": "SUBCONTRATO MONTAJE BUZONES"},
        "SUB_ENSAYOS_COMPAC": {"unidad": "und", "tipo": "AC_SUB", "rol": "Administradora", "orig": "administradora.html", "desc": "SUBCONTRATO DENSIDADES PROCTOR"},
        "SUB_PRUEBA_LAB": {"unidad": "glb", "tipo": "AC_SUB", "rol": "Administradora", "orig": "administradora.html", "desc": "SUBCONTRATO ANALISIS DE AGUA"}
    }

    partidas_info = {
        "01.01": {"desc": "OBRAS PRELIMINARES Y TRABAJOS PROVISIONALES", "unidad": "GLB"},
        "01.02.01": {"desc": "TRAZO, NIVELACION Y REPLANTEO DE ZANJAS", "unidad": "M"},
        "01.02.02": {"desc": "EXCAVACION DE ZANJA C/MAQUINA EN T. SEMIROCOSO", "unidad": "M3"},
        "01.02.03": {"desc": "CAMA DE APOYO H=0.10 M CON ARENA SELECCIONADA", "unidad": "M"},
        "01.02.04": {"desc": "SUMINISTRO E INSTALACION DE TUBERIA PVC DN 200MM S-20", "unidad": "M"},
        "01.02.05": {"desc": "RELLENO COMPACTADO DE ZANJA C/MATERIAL PRESTAMO", "unidad": "M"},
        "01.02.06": {"desc": "CONSTRUCCION DE BUZONES PREFABRICADOS H=1.5 A 2.5M", "unidad": "UND"},
        "01.02.07": {"desc": "CONEXIONES DOMICILIARIAS ALCANTARILLADO DN 160MM", "unidad": "UND"},
        
        "01.03.01": {"desc": "INSTALACION TUBERIA PVC C-10 DN 110MM AGUA POTABLE", "unidad": "M"},
        "01.03.02": {"desc": "CONEXIONES DOMICILIARIAS AGUA POTABLE", "unidad": "UND"},
        
        "01.04.01": {"desc": "PRUEBA HIDRAULICA Y DESINFECCION DE RED DE AGUA", "unidad": "GLB"}
    }

    start_date = datetime(2026, 8, 1)
    logs_generated = []
    log_counter = 1

    for day in range(15):
        current_date = start_date + timedelta(days=day)
        date_str = current_date.strftime("%Y-%m-%d")
        day_num = day + 1
        
        # --- WBS-100: OBRAS PRELIMINARES (Días 1 a 3) ---
        if 1 <= day_num <= 3:
            wbs = "WBS-100"
            # MO
            logs_generated.append((date_str, wbs, "MO_CAPATAZ", 8))
            logs_generated.append((date_str, wbs, "MO_OPERARIO", 16))
            logs_generated.append((date_str, wbs, "MO_PEON", 32))
            # MAT
            if day_num == 1:
                logs_generated.append((date_str, wbs, "MAT_ESTACA", 150))
                logs_generated.append((date_str, wbs, "MAT_PINTURA", 10))
                logs_generated.append((date_str, wbs, "MAT_CORDEL", 100))
                logs_generated.append((date_str, wbs, "SUB_MOVILIZACION", 1))
            # EQP
            logs_generated.append((date_str, wbs, "EQ_ESTACION_TOTAL", 8))
            # EV_PRODUCCION
            logs_generated.append((date_str, wbs, "01.01", 0.33, True)) # 0.33 GLB por día
            logs_generated.append((date_str, wbs, "01.02.01", 800, True)) # 800 M por día

        # --- WBS-200: ALCANTARILLADO Y BUZONES (Días 2 a 15) ---
        if 2 <= day_num <= 15:
            wbs = "WBS-200"
            # MO
            logs_generated.append((date_str, wbs, "MO_CAPATAZ", 8))
            logs_generated.append((date_str, wbs, "MO_OPERARIO", 24))
            logs_generated.append((date_str, wbs, "MO_OFICIAL", 16))
            logs_generated.append((date_str, wbs, "MO_PEON", 64))
            # MAT
            logs_generated.append((date_str, wbs, "MAT_ARENA_CAMAS", 10))
            logs_generated.append((date_str, wbs, "MAT_TUB_PVC_200", 120))
            logs_generated.append((date_str, wbs, "MAT_LUBRICANTE", 2))
            
            # Buzones prefabricados (días impares a partir del día 3)
            is_buzon_day = (day_num in [3, 5, 7, 9, 11, 13, 15])
            if is_buzon_day:
                logs_generated.append((date_str, wbs, "MAT_BUZON_PREF", 1))
                logs_generated.append((date_str, wbs, "EQ_CAMION_GRUAN", 4))
                logs_generated.append((date_str, wbs, "SUB_MONTAJE_BUZON", 1))
                logs_generated.append((date_str, wbs, "01.02.06", 1, True)) # 1 buzón ejecutado

            # Conexiones domiciliarias de alcantarillado (días pares/impares a partir del día 4)
            if day_num >= 4:
                logs_generated.append((date_str, wbs, "MAT_CAJA_REGISTRO", 3))
                logs_generated.append((date_str, wbs, "MAT_TUB_PVC_160_CONEX", 18))
                logs_generated.append((date_str, wbs, "SUB_ENSAYOS_COMPAC", 2))
                logs_generated.append((date_str, wbs, "01.02.07", 3, True)) # 3 conexiones ejecutadas
                
            # EQP
            logs_generated.append((date_str, wbs, "EQ_EXCAVADORA", 8))
            logs_generated.append((date_str, wbs, "EQ_PLANCHA", 8))
            logs_generated.append((date_str, wbs, "EQ_RODILLO_CAMINANTE", 8))
            
            # EV_PRODUCCION
            logs_generated.append((date_str, wbs, "01.02.02", 150, True)) # Excavación M3
            logs_generated.append((date_str, wbs, "01.02.03", 120, True)) # Cama M
            logs_generated.append((date_str, wbs, "01.02.04", 120, True)) # Instalación M
            logs_generated.append((date_str, wbs, "01.02.05", 120, True)) # Relleno M

        # --- WBS-300: AGUA POTABLE (Días 5 a 15) ---
        if 5 <= day_num <= 15:
            wbs = "WBS-300"
            # MO
            logs_generated.append((date_str, wbs, "MO_CAPATAZ", 4))
            logs_generated.append((date_str, wbs, "MO_OPERARIO", 16))
            logs_generated.append((date_str, wbs, "MO_PEON", 32))
            # MAT
            logs_generated.append((date_str, wbs, "MAT_TUB_PVC_110_AGUA", 150))
            logs_generated.append((date_str, wbs, "MAT_ACCESORIOS_AGUA", 1))
            if day_num >= 6:
                logs_generated.append((date_str, wbs, "MAT_CAJA_AGUA", 5))
                logs_generated.append((date_str, wbs, "MAT_TUB_HDPE_1_2", 30))
                logs_generated.append((date_str, wbs, "01.03.02", 5, True)) # 5 conexiones agua ejecutadas
            # EQP
            logs_generated.append((date_str, wbs, "EQ_RETROEXCAVADORA", 8))
            logs_generated.append((date_str, wbs, "EQ_PLANCHA", 8))
            logs_generated.append((date_str, wbs, "EQ_RODILLO_CAMINANTE", 8))
            # EV_PRODUCCION
            logs_generated.append((date_str, wbs, "01.03.01", 150, True)) # Instalación tuberías M

        # --- WBS-400: PRUEBAS E INSTALACIONES (Días 10 a 15) ---
        if 10 <= day_num <= 15:
            wbs = "WBS-400"
            # MO
            logs_generated.append((date_str, wbs, "MO_OPERARIO", 8))
            logs_generated.append((date_str, wbs, "MO_OFICIAL", 8))
            logs_generated.append((date_str, wbs, "MO_PEON", 16))
            # MAT
            logs_generated.append((date_str, wbs, "MAT_HIPOCLORITO", 5))
            logs_generated.append((date_str, wbs, "MAT_AGUA", 20))
            if day_num == 15:
                logs_generated.append((date_str, wbs, "SUB_PRUEBA_LAB", 1))
            # EQP
            logs_generated.append((date_str, wbs, "EQ_BOMBA_PRUEBA", 8))
            logs_generated.append((date_str, wbs, "EQ_CAMION_CISTERNA", 8))
            # EV_PRODUCCION
            logs_generated.append((date_str, wbs, "01.04.01", 0.16, True)) # 0.16 GLB por día (total ~1.0 GLB)

    # 2. ESCRIBIR LOS REGISTROS EN LA PESTAÑA 4
    for idx, log in enumerate(logs_generated):
        r = idx + 2 # Fila en excel (inicia en 2)
        date_str, wbs, code, qty = log[0], log[1], log[2], log[3]
        is_partida = len(log) > 4 and log[4]
        
        # Obtener mapeo del recurso o de la partida
        if is_partida:
            mapping = partidas_info[code]
            category = "EV_PRODUCCION"
            rol = "Ing. de Campo"
            orig = "ing_campo.html"
        else:
            mapping = resource_info[code]
            category = mapping["tipo"]
            rol = mapping["rol"]
            orig = mapping["orig"]
            
        desc = mapping["desc"]
        unit = mapping["unidad"]
        
        # Generar ID Registro secuencial
        num_str = str(log_counter).zfill(3)
        id_val = f"LOG-{date_str.replace('-', '')}-{num_str}"
        log_counter += 1

        # Celdas
        sheet4.cell(row=r, column=1, value=id_val) # ID Registro
        sheet4.cell(row=r, column=2, value=date_str) # Fecha
        sheet4.cell(row=r, column=3, value=rol) # Rol
        sheet4.cell(row=r, column=4, value=wbs) # WBS
        sheet4.cell(row=r, column=5, value=code) # Código
        sheet4.cell(row=r, column=6, value=desc) # Descripción
        sheet4.cell(row=r, column=7, value=qty) # Cantidad
        sheet4.cell(row=r, column=8, value=unit) # Unidad
        
        # Fórmula P.U.: Columna I (9)
        sheet4.cell(row=r, column=9, value=(
            f"=IFERROR(VLOOKUP(E{r}, '05_MAESTRO_RECURSOS'!A:D, 4, FALSE), "
            f"IFERROR(VLOOKUP(E{r}, '06_MAESTRO_PARTIDAS_EV'!B:F, 5, FALSE), 0))"
        ))
        
        # Fórmula Subtotal: Columna J (10)
        sheet4.cell(row=r, column=10, value=f"=ROUND(G{r} * I{r}, 2)")
        
        # Categoria y origen
        sheet4.cell(row=r, column=11, value=category)
        sheet4.cell(row=r, column=12, value=orig)
        
        # Alinear al centro columnas A, D, E, H
        sheet4.cell(row=r, column=1).alignment = openpyxl.styles.Alignment(horizontal="center")
        sheet4.cell(row=r, column=4).alignment = openpyxl.styles.Alignment(horizontal="center")
        sheet4.cell(row=r, column=5).alignment = openpyxl.styles.Alignment(horizontal="center")
        sheet4.cell(row=r, column=8).alignment = openpyxl.styles.Alignment(horizontal="center")

    wb.save('Base_Datos_Proyecto_Sheets_Viva.xlsx')
    print(f"[OK] Se han generado {len(logs_generated)} registros de base sintetica desde el 01/08 hasta el 15/08.")

if __name__ == '__main__':
    generate_database()
