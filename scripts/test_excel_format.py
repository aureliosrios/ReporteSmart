import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

# Test standard numeric values with double-quoted currency format
ws['A1'] = 8296.40
ws['A1'].number_format = '"S/" #,##0.00'

ws['B1'] = 3.35
ws['B1'].number_format = '"S/" #,##0.00'

ws['C1'] = '=A1*B1'
ws['C1'].number_format = '"S/" #,##0.00'

wb.save("test_format.xlsx")
print("[OK] Excel de prueba creado con formato \"S/\" #,##0.00")
