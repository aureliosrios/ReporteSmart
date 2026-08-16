import openpyxl

def update_formulas():
    wb = openpyxl.load_workbook('Base_Datos_Proyecto_Sheets_Viva.xlsx', data_only=False)
    sheet = wb['03_CONSOLIDADO_DIARIO_EVM_WBS']
    
    print("Modificando fórmulas en la pestaña 03_CONSOLIDADO_DIARIO_EVM_WBS...")
    
    # Recorrer las filas de datos (de la 2 a la 241)
    for r in range(2, 242):
        # Column H: EV Ejecutado Día
        sheet.cell(row=r, column=8).value = (
            f'=IF(B{r} > MAX(\'04_LOG_FIELD_ENTRIES\'!B:B), "", '
            f'SUMIFS(\'04_LOG_FIELD_ENTRIES\'!J:J, \'04_LOG_FIELD_ENTRIES\'!B:B, B{r}, '
            f'\'04_LOG_FIELD_ENTRIES\'!D:D, C{r}, \'04_LOG_FIELD_ENTRIES\'!K:K, "EV_PRODUCCION"))'
        )
        
        # Column I: EV Acumulado
        sheet.cell(row=r, column=9).value = (
            f'=IF(B{r} > MAX(\'04_LOG_FIELD_ENTRIES\'!B:B), "", '
            f'SUMIFS(H$2:H{r}, C$2:C{r}, C{r}))'
        )
        
        # Column J: AC Costo Real Día
        sheet.cell(row=r, column=10).value = (
            f'=IF(B{r} > MAX(\'04_LOG_FIELD_ENTRIES\'!B:B), "", '
            f'SUMIFS(\'04_LOG_FIELD_ENTRIES\'!J:J, \'04_LOG_FIELD_ENTRIES\'!B:B, B{r}, '
            f'\'04_LOG_FIELD_ENTRIES\'!D:D, C{r}, \'04_LOG_FIELD_ENTRIES\'!K:K, "AC_*"))'
        )
        
        # Column K: AC Acumulado
        sheet.cell(row=r, column=11).value = (
            f'=IF(B{r} > MAX(\'04_LOG_FIELD_ENTRIES\'!B:B), "", '
            f'SUMIFS(J$2:J{r}, C$2:C{r}, C{r}))'
        )
        
        # Column L: Variación Costo CV
        sheet.cell(row=r, column=12).value = f'=IF(K{r}="", "", I{r}-K{r})'
        
        # Column M: Variación Plazo SV
        sheet.cell(row=r, column=13).value = f'=IF(I{r}="", "", I{r}-G{r})'
        
        # Column N: SPI Plazo
        sheet.cell(row=r, column=14).value = f'=IF(I{r}="", "", IF(G{r}>0, I{r}/G{r}, 1))'
        
        # Column O: CPI Costo
        sheet.cell(row=r, column=15).value = f'=IF(I{r}="", "", IF(K{r}>0, I{r}/K{r}, 1))'
        
        # Column P: Estado Alerta WBS
        sheet.cell(row=r, column=16).value = (
            f'=IF(I{r}="", "", IF(AND(N{r}>=0.95, O{r}>=0.95), "SALUDABLE", '
            f'IF(N{r}<0.9, "ALERTA RETRASO", "ALERTA SOBRECOSTO")))'
        )

    wb.save('Base_Datos_Proyecto_Sheets_Viva.xlsx')
    print("[OK] Formulas actualizadas y guardadas con exito en 'Base_Datos_Proyecto_Sheets_Viva.xlsx'.")

if __name__ == '__main__':
    update_formulas()
