import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True)
font_regular = Font(name="Calibri", size=10)

fill_dark = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
fill_blue = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
fill_green = PatternFill(start_color="15803D", end_color="15803D", fill_type="solid")
fill_orange = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")
fill_purple = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
fill_kpi = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
fill_ok = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Soft green
fill_adj = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft red

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)
double_bottom_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='double', color='1E3A8A')
)

align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------
# PESTAÑA 1: AUDITORIA_ALUMNOS (RÚBRICA Y EVALUACIÓN POR ALUMNO)
# ---------------------------------------------------------
ws_aud = wb.create_sheet(title="01_AUDITORIA_ALUMNOS")
ws_aud.views.sheetView[0].showGridLines = True

headers_aud = [
    "N°", "Nombre del Alumno", "Portal HTML Evaluado", "Rol de Campo", 
    "Estado Alineación Data", "Campo Fecha (ISO)", "Campo WBS (Select)", 
    "Campo Código Recurso/Partida", "Campo CANTIDAD (Solo Número)", 
    "Campo Unidad Medida", "Campo Descripción / Detalle", "JSON POST Payload Aceptado", 
    "Adaptación / Acción Requerida en HTML-JS"
]

ws_aud.row_dimensions[1].height = 36

for c_idx, h in enumerate(headers_aud, start=1):
    cell = ws_aud.cell(row=1, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_dark
    cell.alignment = align_header

# Datos de muestra de evaluación de alumnos para la clase
alumnos_sample = [
    (1, "Juan Pérez", "tareador.html", "Tareador (Bildin)", "CONFORME", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "Ninguna. HTML alineado 100% con Pestaña 04."),
    (2, "María Gómez", "almacen.html", "Almacenero", "AJUSTE REQUERIDO", "OK", "OK", "NO (Usa texto)", "OK", "OK", "OK", "NO (Falta campo)", "Cambiar input de material a select con 'codigoRecurso' del maestro."),
    (3, "Carlos Mendoza", "partes_maquinaria.html", "Administradora", "CONFORME", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "Ninguna. Cumple contrato JSON."),
    (4, "Ana Torres", "avance_campo.html", "Ing. de Campo", "AJUSTE REQUERIDO", "OK", "NO (Escribe libre)", "OK", "OK", "OK", "OK", "NO (Envía P.U.)", "Restringir WBS a lista desplegable y eliminar envío de P.U. desde el cliente."),
    (5, "Luis Fernández", "tareador_v2.html", "Tareador (Bildin)", "CONFORME", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "Ninguna. Excelente estructura.")
]

for r_idx, (num, nom, html_file, rol, st, f_fec, f_wbs, f_cod, f_cant, f_und, f_det, f_post, obs) in enumerate(alumnos_sample, start=2):
    ws_aud.row_dimensions[r_idx].height = 22
    
    ws_aud.cell(row=r_idx, column=1, value=num).alignment = align_center
    ws_aud.cell(row=r_idx, column=2, value=nom).alignment = align_left
    ws_aud.cell(row=r_idx, column=2).font = font_bold
    
    ws_aud.cell(row=r_idx, column=3, value=html_file).alignment = align_left
    ws_aud.cell(row=r_idx, column=4, value=rol).alignment = align_left
    
    cell_st = ws_aud.cell(row=r_idx, column=5, value=st)
    cell_st.alignment = align_center
    cell_st.font = font_bold
    cell_st.fill = fill_ok if st == "CONFORME" else fill_adj
    
    for c_i, val_c in enumerate([f_fec, f_wbs, f_cod, f_cant, f_und, f_det, f_post], start=6):
        cell_check = ws_aud.cell(row=r_idx, column=c_i, value=val_c)
        cell_check.alignment = align_center
        if val_c == "OK":
            cell_check.font = font_bold
        else:
            cell_check.font = font_bold
            cell_check.fill = fill_adj
            
    ws_aud.cell(row=r_idx, column=13, value=obs).alignment = align_left
    
    for c in range(1, 14):
        ws_aud.cell(row=r_idx, column=c).border = thin_border

widths_aud = [6, 22, 22, 18, 18, 12, 14, 16, 16, 12, 18, 16, 38]
for c_i, w in enumerate(widths_aud, start=1):
    ws_aud.column_dimensions[get_column_letter(c_i)].width = w


# ---------------------------------------------------------
# PESTAÑA 2: CONTRATO_SCHEMA_SHEETS (ESTÁNDAR DE INTEGRACIÓN)
# ---------------------------------------------------------
ws_sch = wb.create_sheet(title="02_CONTRATO_SCHEMA_SHEETS")
ws_sch.views.sheetView[0].showGridLines = True

headers_sch = [
    "Campo JSON en HTML (POST)", "Columna Destino en Pestaña 04", 
    "Tipo de Dato Esperado", "Regla de Validación / Obligatoriedad", 
    "Origen del Valor", "Ejemplo de Payload Válido"
]

ws_sch.row_dimensions[1].height = 36

for c_idx, h in enumerate(headers_sch, start=1):
    cell = ws_sch.cell(row=1, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = align_header

schema_rules = [
    ("id", "Columna A (ID Registro)", "Texto / String", "Obligatorio. Generado auto por JS (ej. REG-17232000)", "JavaScript cliente", "LOG-20260803-020"),
    ("fecha", "Columna B (Fecha)", "Fecha ISO (YYYY-MM-DD)", "Obligatorio. Formato Estándar de fecha", "Input date HTML", "2026-08-03"),
    ("rol", "Columna C (Rol Responsable)", "Texto Enum", "Obligatorio. Tareador, Almacenero, Administradora, Ing. Campo", "Definido por archivo HTML", "Almacenero"),
    ("wbs", "Columna D (Código WBS)", "Texto Enum", "Obligatorio. WBS-100, WBS-200, WBS-300 o WBS-400", "Select option HTML", "WBS-200"),
    ("codigoRecurso", "Columna E (Código Recurso/Partida)", "Texto Enum", "Obligatorio. Debe coincidir con Pestaña 05 o 06", "Select option HTML", "MAT_TUB_PVC_200"),
    ("detalle", "Columna F (Descripción / Detalle)", "Texto libre", "Opcional. Ubicación, tramo o nota aclaratoria", "Input text HTML", "Tubería PVC 200mm Tramo 1"),
    ("cantidad", "Columna G (Cantidad Campo)", "Número Decimal", "Obligatorio. Mayor a cero (> 0)", "Input number HTML", "280.00"),
    ("unidad", "Columna H (Unidad Medida)", "Texto corto", "Obligatorio. hh, hm, m, m3, und, glb, pza", "Auto por opción elegida", "m"),
    ("pu", "Columna I (P.U. Maestro)", "Fórmula Viva Sheets", "NO enviar desde HTML. Sheets lo busca automáticamente con VLOOKUP", "Fórmula en Pestaña 04", "=VLOOKUP(...) (Auto)"),
    ("costo", "Columna J (Subtotal Monto)", "Fórmula Viva Sheets", "NO enviar desde HTML. Sheets multiplica ROUND(G*I, 2)", "Fórmula en Pestaña 04", "=ROUND(G*I, 2) (Auto)"),
    ("tipo", "Columna K (Categoría EVM)", "Texto Enum", "Obligatorio. AC_MO, AC_MAT, AC_EQP, AC_SUB o EV_PRODUCCION", "Asignado por rol HTML", "AC_MAT"),
    ("origen_html", "Columna L (Origen HTML)", "Texto String", "Obligatorio. Nombre del archivo HTML emisor", "Nombre de archivo", "almacenero.html")
]

for r_s, (j_key, col_sheets, dt_type, req_rule, orig, ex_val) in enumerate(schema_rules, start=2):
    ws_sch.row_dimensions[r_s].height = 22
    
    ws_sch.cell(row=r_s, column=1, value=j_key).font = font_bold
    ws_sch.cell(row=r_s, column=1).alignment = align_center
    
    ws_sch.cell(row=r_s, column=2, value=col_sheets).font = font_regular
    ws_sch.cell(row=r_s, column=2).alignment = align_left
    
    ws_sch.cell(row=r_s, column=3, value=dt_type).font = font_regular
    ws_sch.cell(row=r_s, column=3).alignment = align_center
    
    ws_sch.cell(row=r_s, column=4, value=req_rule).font = font_regular
    ws_sch.cell(row=r_s, column=4).alignment = align_left
    
    ws_sch.cell(row=r_s, column=5, value=orig).font = font_regular
    ws_sch.cell(row=r_s, column=5).alignment = align_left
    
    ws_sch.cell(row=r_s, column=6, value=ex_val).font = font_bold
    ws_sch.cell(row=r_s, column=6).alignment = align_left

    for c in range(1, 7):
        ws_sch.cell(row=r_s, column=c).border = thin_border

widths_sch = [20, 24, 16, 36, 18, 22]
for c_i, w in enumerate(widths_sch, start=1):
    ws_sch.column_dimensions[get_column_letter(c_i)].width = w

os.makedirs("outputs", exist_ok=True)
excel_checklist_path = "outputs/Lista_Chequeo_Auditoria_HTMLs_Alumnos.xlsx"
wb.save(excel_checklist_path)

print(f"[OK] Archivo Excel de Lista de Chequeo de Alumnos creado en: {excel_checklist_path}")
