import json
import os
import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------
# 1. CATALOGOS Y LÍNEA BASE DEL PROYECTO (60 DÍAS COMPLETOS)
# ---------------------------------------------------------

metadata = {
    "proyecto": "Redes Sanitarias de Agua Potable y Alcantarillado - Habilitación Urbana Los Cedros",
    "codigo_proyecto": "HU-CEDROS-2026",
    "version_esquema": "6.0.0",
    "moneda": "PEN",
    "duracion_dias_calendario": 60,
    "fecha_inicio_obra": "2026-08-01",
    "fecha_fin_programada": "2026-09-29",
    "descripcion": "Base de datos viva oficial para Google Sheets. Incluye el CRONOGRAMA VALORIZADO DIARIO COMPLETO PARA LOS 60 DÍAS DE LA OBRA (PV por día y WBS). El campo envía las cantidades diarias reales. Las fórmulas de Sheets calculan AC, EV, CPI y SPI en tiempo real a cualquier fecha de corte."
}

# 1.1 Maestro de Recursos (Mano de Obra, Materiales, Equipos, Subcontratos)
maestro_recursos = [
    # MANO DE OBRA
    {"codigo": "MO_CAPATAZ", "descripcion": "CAPATAZ", "unidad": "hh", "precio_unitario": 32.50, "categoria": "AC_MO"},
    {"codigo": "MO_OPERARIO", "descripcion": "OPERARIO", "unidad": "hh", "precio_unitario": 26.80, "categoria": "AC_MO"},
    {"codigo": "MO_OFICIAL", "descripcion": "OFICIAL", "unidad": "hh", "precio_unitario": 22.40, "categoria": "AC_MO"},
    {"codigo": "MO_PEON", "descripcion": "PEON", "unidad": "hh", "precio_unitario": 20.10, "categoria": "AC_MO"},
    # MATERIALES
    {"codigo": "MAT_ESTACA", "descripcion": "ESTACAS Y MADERA PARA TRAZO", "unidad": "pza", "precio_unitario": 4.50, "categoria": "AC_MAT"},
    {"codigo": "MAT_PINTURA", "descripcion": "PINTURA SPRAY MARCADOR DE ZANJA", "unidad": "gla", "precio_unitario": 18.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_CORDEL", "descripcion": "CORDEL / NYLON", "unidad": "m", "precio_unitario": 0.80, "categoria": "AC_MAT"},
    {"codigo": "MAT_ARENA_CAMAS", "descripcion": "ARENA FINA SELECCIONADA PARA CAMA E=0.10M", "unidad": "m3", "precio_unitario": 45.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_TUB_PVC_200", "descripcion": "TUBERIA PVC UF DN 200MM SERIE S-20 ALCANTARILLADO", "unidad": "m", "precio_unitario": 42.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_LUBRICANTE", "descripcion": "LUBRICANTE PARA TUBERIA CON ESPIGA Y CAMPANA", "unidad": "kg", "precio_unitario": 24.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_MATERIAL_PR", "descripcion": "MATERIAL DE PRÉSTAMO AFIRMADO PARA RELLENO COMPACTADO", "unidad": "m3", "precio_unitario": 38.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_AGUA", "descripcion": "AGUA PARA COMPACTACIÓN Y PRUEBAS", "unidad": "m3", "precio_unitario": 12.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_BUZON_PREF", "descripcion": "BUZON PREFABRICADO CONCRETO H=1.5-2.5M CON MARCO Y TAPA", "unidad": "und", "precio_unitario": 1050.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_CAJA_REGISTRO", "descripcion": "CAJA DE REGISTRO PREFABRICADA DE CONCRETO 12X24 WITH TAPA", "unidad": "und", "precio_unitario": 145.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_TUB_PVC_160_CONEX", "descripcion": "TUBERIA PVC SAL DN 160MM PARA ACOMETIDA ALCANTARILLADO", "unidad": "m", "precio_unitario": 22.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_TUB_PVC_110_AGUA", "descripcion": "TUBERIA PVC C-10 DN 110MM AGUA POTABLE", "unidad": "m", "precio_unitario": 26.50, "categoria": "AC_MAT"},
    {"codigo": "MAT_CAJA_AGUA", "descripcion": "CAJA TERMOFORMADA PARA MEDIDOR + ABRAZADERA + LLAVE DE PASO", "unidad": "und", "precio_unitario": 62.00, "categoria": "AC_MAT"},
    # EQUIPOS Y MAQUINARIA
    {"codigo": "EQ_EXCAVADORA", "descripcion": "EXCAVADORA SOBRE ORUGAS 140 HP", "unidad": "hm", "precio_unitario": 185.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_RETROEXCAVADORA", "descripcion": "RETROEXCAVADORA SOBRE LLANTAS 62 HP", "unidad": "hm", "precio_unitario": 130.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_PLANCHA", "descripcion": "PLANCHA COMPACTADORA 7 HP", "unidad": "hm", "precio_unitario": 25.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_RODILLO_CAMINANTE", "descripcion": "RODILLO VIBRATORIO CAMINANTE 1.5 TN", "unidad": "hm", "precio_unitario": 45.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_CAMION_CISTERNA", "descripcion": "CAMION CISTERNA 4X2 (AGUA) 2,000 GLN", "unidad": "hm", "precio_unitario": 140.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_CAMION_GRUAN", "descripcion": "CAMION GRUA 5 TN PARA MONTAJE DE BUZONES", "unidad": "hm", "precio_unitario": 160.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_ESTACION_TOTAL", "descripcion": "EQUIPO DE TOPOGRAFIA ESTACION TOTAL", "unidad": "hm", "precio_unitario": 18.00, "categoria": "AC_EQP"},
    # SUBCONTRATOS
    {"codigo": "SUB_MOVILIZACION", "descripcion": "SUBCONTRATO DE MOVILIZACION Y DESMOVILIZACION DE MAQUINARIA", "unidad": "glb", "precio_unitario": 4000.00, "categoria": "AC_SUB"},
    {"codigo": "SUB_ENSAYOS_COMPAC", "descripcion": "SUBCONTRATO DE ENSAYOS DE DENSIDAD DE CAMPO (PROCTOR/DENSIDAD)", "unidad": "und", "precio_unitario": 80.00, "categoria": "AC_SUB"},
    {"codigo": "SUB_MONTAJE_BUZON", "descripcion": "SUBCONTRATO SERVICIO ASENTADO Y ANCLAJE BUZONES PREFABRICADOS", "unidad": "und", "precio_unitario": 250.00, "categoria": "AC_SUB"}
]

# 1.2 Maestro de Partidas Presupuestales (para avance de Producción EV)
maestro_partidas_ev = [
    {"codigo": "01.01", "descripcion": "Obras Preliminares y Trabajos Provisionales", "unidad": "GLB", "precio_unitario": 8296.40, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.01", "descripcion": "Trazo, Nivelación y Replanteo de Zanjas", "unidad": "M", "precio_unitario": 3.35, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.02", "descripcion": "Excavación de Zanja H=1.50m - 2.20m a Máquina (Terreno Normal)", "unidad": "M3", "precio_unitario": 11.66, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.03", "descripcion": "Preparación y Colocación de Cama de Arena e=0.10m", "unidad": "M", "precio_unitario": 11.12, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.04", "descripcion": "Suministro e Instalación de Tubería PVC UF DN 200mm Serie S-20 para Alcantarillado", "unidad": "M", "precio_unitario": 55.51, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.05", "descripcion": "Relleno Compactado de Zanja en Capas de 0.20m con Maquinaria/Plancha", "unidad": "M3", "precio_unitario": 28.54, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.06", "descripcion": "Construcción de Buzones Prefabricados de Concreto h=1.50m - 2.50m (Inc. Marco y Tapa)", "unidad": "UND", "precio_unitario": 2500.73, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.07", "descripcion": "Conexiones Domiciliarias de Alcantarillado (Caja de Registro + Acometida)", "unidad": "UND", "precio_unitario": 405.54, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.03.01", "descripcion": "Suministro e Instalación de Tubería PVC C-10 DN 110mm para Agua Potable", "unidad": "M", "precio_unitario": 38.76, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.03.02", "descripcion": "Conexiones Domiciliarias de Agua Potable (Caja de Agua + Abrazadera + Acometida)", "unidad": "UND", "precio_unitario": 156.71, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.04.01", "descripcion": "Pruebas Hidráulicas de Redes de Agua y Alcantarillado + Desinfección", "unidad": "GLB", "precio_unitario": 9348.06, "categoria": "EV_PRODUCCION"}
]

# 1.3 Presupuesto Meta Total por WBS (BAC)
presupuesto_wbs_bac = [
    {"codigo_wbs": "WBS-100", "nombre_frente": "Obras Preliminares, Trazo y Movilización", "bac_meta_pen": 8296.40},
    {"codigo_wbs": "WBS-200", "nombre_frente": "Red de Alcantarillado, Zanjas y Buzones", "bac_meta_pen": 380086.56},
    {"codigo_wbs": "WBS-300", "nombre_frente": "Red de Agua Potable y Conexiones Domiciliarias", "bac_meta_pen": 57565.20},
    {"codigo_wbs": "WBS-400", "nombre_frente": "Pruebas Hidráulicas, Desinfección y Entrega", "bac_meta_pen": 9348.06}
]

# ---------------------------------------------------------
# GENERADOR AUTOMÁTICO DE CRONOGRAMA PV DIARIO COMPLETO (60 DÍAS)
# ---------------------------------------------------------

start_date = datetime.strptime("2026-08-01", "%Y-%m-%d")
total_days = 60

# Definición de ventanas de ejecución activa por WBS (Día Inicio, Día Fin de cada frente)
wbs_schedule_plan = {
    "WBS-100": {"day_start": 1, "day_end": 14, "bac": 8296.40},
    "WBS-200": {"day_start": 10, "day_end": 43, "bac": 380086.56},
    "WBS-300": {"day_start": 22, "day_end": 51, "bac": 57565.20},
    "WBS-400": {"day_start": 48, "day_end": 60, "bac": 9348.06}
}

# Calcular la tasa diaria constante para cada WBS en su ventana activa
for code, plan in wbs_schedule_plan.items():
    days_duration = (plan["day_end"] - plan["day_start"]) + 1
    plan["daily_rate"] = round(plan["bac"] / days_duration, 2)

cronograma_pv_diario = []
pv_acum_wbs = {code: 0.0 for code in wbs_schedule_plan}

for day_idx in range(1, total_days + 1):
    current_dt = start_date + timedelta(days=day_idx - 1)
    date_str = current_dt.strftime("%Y-%m-%d")
    
    for code, plan in wbs_schedule_plan.items():
        if plan["day_start"] <= day_idx <= plan["day_end"]:
            # Si es el último día activo de ese WBS, ajustar redondeo exacto al BAC
            if day_idx == plan["day_end"]:
                pv_day = round(plan["bac"] - pv_acum_wbs[code], 2)
            else:
                pv_day = plan["daily_rate"]
        else:
            pv_day = 0.0
            
        pv_acum_wbs[code] = round(pv_acum_wbs[code] + pv_day, 2)
        
        cronograma_pv_diario.append({
            "dia_numero": day_idx,
            "fecha": date_str,
            "codigo_wbs": code,
            "pv_diario_pen": pv_day,
            "pv_acumulado_pen": pv_acum_wbs[code]
        })

# ---------------------------------------------------------
# 2. ENTRADAS DE CAMPO SIMULADAS (MUESTRA DE 5 DÍAS REALES)
# ---------------------------------------------------------

field_entries_raw = [
    # DÍA 1 (2026-08-01)
    {"id": "LOG-20260801-001", "fecha": "2026-08-01", "rol": "Tareador (Bildin)", "wbs": "WBS-100", "recurso": "MO_CAPATAZ", "detalle": "CAPATAZ (1 pers. x 8h)", "cantidad": 8.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260801-002", "fecha": "2026-08-01", "rol": "Tareador (Bildin)", "wbs": "WBS-100", "recurso": "MO_OPERARIO", "detalle": "OPERARIO (2 pers. x 8h)", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260801-003", "fecha": "2026-08-01", "rol": "Tareador (Bildin)", "wbs": "WBS-100", "recurso": "MO_PEON", "detalle": "PEON (4 pers. x 8h)", "cantidad": 32.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260801-004", "fecha": "2026-08-01", "rol": "Almacenero", "wbs": "WBS-100", "recurso": "MAT_ESTACA", "detalle": "ESTACAS Y MADERA PARA TRAZO", "cantidad": 150.0, "unidad": "pza", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260801-005", "fecha": "2026-08-01", "rol": "Almacenero", "wbs": "WBS-100", "recurso": "MAT_PINTURA", "detalle": "PINTURA SPRAY MARCADOR ZANJA", "cantidad": 10.0, "unidad": "gla", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260801-006", "fecha": "2026-08-01", "rol": "Administradora", "wbs": "WBS-100", "recurso": "SUB_MOVILIZACION", "detalle": "MOVILIZACION DE MAQUINARIA", "cantidad": 1.0, "unidad": "glb", "cat": "AC_SUB", "html": "administradora.html"},
    {"id": "LOG-20260801-007", "fecha": "2026-08-01", "rol": "Administradora", "wbs": "WBS-100", "recurso": "EQ_ESTACION_TOTAL", "detalle": "ESTACION TOTAL TOPOGRAFICA", "cantidad": 8.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260801-008", "fecha": "2026-08-01", "rol": "Ing. de Campo", "wbs": "WBS-100", "recurso": "01.01", "detalle": "OBRAS PRELIMINARES Y PROVISIONALES (50%)", "cantidad": 0.5, "unidad": "GLB", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260801-009", "fecha": "2026-08-01", "rol": "Ing. de Campo", "wbs": "WBS-100", "recurso": "01.02.01", "detalle": "TRAZO Y REPLANTEO INICIAL TRAMO T-1", "cantidad": 500.0, "unidad": "M", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},

    # DÍA 2 (2026-08-02)
    {"id": "LOG-20260802-010", "fecha": "2026-08-02", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_CAPATAZ", "detalle": "CAPATAZ CONTROL ZANJA", "cantidad": 8.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260802-011", "fecha": "2026-08-02", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_OPERARIO", "detalle": "OPERARIO ENTIBADO Y CAMA", "cantidad": 24.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260802-012", "fecha": "2026-08-02", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_PEON", "detalle": "PEON REPELE Y CAMA", "cantidad": 48.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260802-013", "fecha": "2026-08-02", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_ARENA_CAMAS", "detalle": "ARENA FINA SELECCIONADA PARA CAMA", "cantidad": 35.0, "unidad": "m3", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260802-014", "fecha": "2026-08-02", "rol": "Administradora", "wbs": "WBS-200", "recurso": "EQ_EXCAVADORA", "detalle": "EXCAVADORA SOBRE ORUGAS 140 HP", "cantidad": 8.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260802-015", "fecha": "2026-08-02", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.02", "detalle": "EXCAVACION ZANJA H=1.5-2.2M A MAQUINA", "cantidad": 450.0, "unidad": "M3", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260802-016", "fecha": "2026-08-02", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.03", "detalle": "COLOCACION CAMA DE ARENA E=0.10M", "cantidad": 300.0, "unidad": "M", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},

    # DÍA 3 (2026-08-03)
    {"id": "LOG-20260803-017", "fecha": "2026-08-03", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_OPERARIO", "detalle": "OPERARIO TUBERO ALCANTARILLADO", "cantidad": 24.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260803-018", "fecha": "2026-08-03", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_OFICIAL", "detalle": "OFICIAL COMPACTACION", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260803-019", "fecha": "2026-08-03", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_PEON", "detalle": "PEON APOYO MONTAJE TUBERIA", "cantidad": 40.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260803-020", "fecha": "2026-08-03", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_TUB_PVC_200", "detalle": "TUBERIA PVC UF DN 200MM S-20", "cantidad": 280.0, "unidad": "m", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260803-021", "fecha": "2026-08-03", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_LUBRICANTE", "detalle": "LUBRICANTE PARA TUBERIA PVC", "cantidad": 12.0, "unidad": "kg", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260803-022", "fecha": "2026-08-03", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_MATERIAL_PR", "detalle": "MATERIAL AFIRMADO RELLENO COMPACTADO", "cantidad": 60.0, "unidad": "m3", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260803-023", "fecha": "2026-08-03", "rol": "Administradora", "wbs": "WBS-200", "recurso": "EQ_RETROEXCAVADORA", "detalle": "RETROEXCAVADORA RELLENO Y LIMPIEZA", "cantidad": 8.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260803-024", "fecha": "2026-08-03", "rol": "Administradora", "wbs": "WBS-200", "recurso": "EQ_PLANCHA", "detalle": "PLANCHA COMPACTADORA 7 HP", "cantidad": 8.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260803-025", "fecha": "2026-08-03", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.04", "detalle": "INSTALACION TUBERIA PVC 200MM ALCANTARILLADO", "cantidad": 250.0, "unidad": "M", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260803-026", "fecha": "2026-08-03", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.05", "detalle": "RELLENO COMPACTADO ZANJA CAPAS 0.20M", "cantidad": 200.0, "unidad": "M3", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},

    # DÍA 4 (2026-08-04)
    {"id": "LOG-20260804-027", "fecha": "2026-08-04", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_OPERARIO", "detalle": "OPERARIO MONTAJE BUZONES", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260804-028", "fecha": "2026-08-04", "rol": "Tareador (Bildin)", "wbs": "WBS-300", "recurso": "MO_OPERARIO", "detalle": "OPERARIO TUBERO AGUA POTABLE", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260804-029", "fecha": "2026-08-04", "rol": "Tareador (Bildin)", "wbs": "WBS-300", "recurso": "MO_PEON", "detalle": "PEON APERTURA ZANJA AGUA", "cantidad": 32.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260804-030", "fecha": "2026-08-04", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_BUZON_PREF", "detalle": "BUZON PREFABRICADO DE CONCRETO H=2.0M", "cantidad": 4.0, "unidad": "und", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260804-031", "fecha": "2026-08-04", "rol": "Almacenero", "wbs": "WBS-300", "recurso": "MAT_TUB_PVC_110_AGUA", "detalle": "TUBERIA PVC C-10 DN 110MM AGUA", "cantidad": 200.0, "unidad": "m", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260804-032", "fecha": "2026-08-04", "rol": "Administradora", "wbs": "WBS-200", "recurso": "EQ_CAMION_GRUAN", "detalle": "CAMION GRUA 5 TN MONTAJE BUZONES", "cantidad": 6.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260804-033", "fecha": "2026-08-04", "rol": "Administradora", "wbs": "WBS-200", "recurso": "SUB_MONTAJE_BUZON", "detalle": "SUBCONTRATO ASENTADO Y SELLADO BUZONES", "cantidad": 4.0, "unidad": "und", "cat": "AC_SUB", "html": "administradora.html"},
    {"id": "LOG-20260804-034", "fecha": "2026-08-04", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.06", "detalle": "CONSTRUCCION BUZONES PREFABRICADOS CONCRETO", "cantidad": 4.0, "unidad": "UND", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260804-035", "fecha": "2026-08-04", "rol": "Ing. de Campo", "wbs": "WBS-300", "recurso": "01.03.01", "detalle": "INSTALACION TUBERIA PVC C-10 DN 110MM AGUA", "cantidad": 180.0, "unidad": "M", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},

    # DÍA 5 (2026-08-05)
    {"id": "LOG-20260805-036", "fecha": "2026-08-05", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_OPERARIO", "detalle": "OPERARIO CONEXION DOMICILIARIA ALCANTARILLADO", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260805-037", "fecha": "2026-08-05", "rol": "Tareador (Bildin)", "wbs": "WBS-300", "recurso": "MO_OPERARIO", "detalle": "OPERARIO CONEXION DOMICILIARIA AGUA", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260805-038", "fecha": "2026-08-05", "rol": "Tareador (Bildin)", "wbs": "WBS-400", "recurso": "MO_OFICIAL", "detalle": "OFICIAL PRUEBAS HIDRAULICAS DE RED", "cantidad": 8.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260805-039", "fecha": "2026-08-05", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_CAJA_REGISTRO", "detalle": "CAJA DE REGISTRO PREFABRICADA 12X24", "cantidad": 15.0, "unidad": "und", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260805-040", "fecha": "2026-08-05", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_TUB_PVC_160_CONEX", "detalle": "TUBERIA PVC SAL DN 160MM ACOMETIDA", "cantidad": 90.0, "unidad": "m", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260805-041", "fecha": "2026-08-05", "rol": "Almacenero", "wbs": "WBS-300", "recurso": "MAT_CAJA_AGUA", "detalle": "CAJA TERMOFORMADA MEDIDOR + ABRAZADERA", "cantidad": 15.0, "unidad": "und", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260805-042", "fecha": "2026-08-05", "rol": "Administradora", "wbs": "WBS-200", "recurso": "SUB_ENSAYOS_COMPAC", "detalle": "SUBCONTRATO ENSAYOS DENSIDAD DE CAMPO", "cantidad": 10.0, "unidad": "und", "cat": "AC_SUB", "html": "administradora.html"},
    {"id": "LOG-20260805-043", "fecha": "2026-08-05", "rol": "Administradora", "wbs": "WBS-400", "recurso": "EQ_CAMION_CISTERNA", "detalle": "CAMION CISTERNA AGUA PRUEBAS HIDRAULICAS", "cantidad": 8.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260805-044", "fecha": "2026-08-05", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.07", "detalle": "CONEXIONES DOMICILIARIAS ALCANTARILLADO (15 ACOMETIDAS)", "cantidad": 15.0, "unidad": "UND", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260805-045", "fecha": "2026-08-05", "rol": "Ing. de Campo", "wbs": "WBS-300", "recurso": "01.03.02", "detalle": "CONEXIONES DOMICILIARIAS AGUA POTABLE (15 ACOMETIDAS)", "cantidad": 15.0, "unidad": "UND", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260805-046", "fecha": "2026-08-05", "rol": "Ing. de Campo", "wbs": "WBS-400", "recurso": "01.04.01", "detalle": "PRUEBA HIDRAULICA TRAMO 1 REDES (25%)", "cantidad": 0.25, "unidad": "GLB", "cat": "EV_PRODUCCION", "html": "ing_campo.html"}
]

# ---------------------------------------------------------
# 3. GENERACIÓN DEL JSON ESTRUCTURADO MASTER CON 60 DÍAS
# ---------------------------------------------------------

db_sheets_json = {
    "metadata": metadata,
    "linea_base_proyecto": {
        "maestro_recursos": maestro_recursos,
        "maestro_partidas_ev": maestro_partidas_ev,
        "presupuesto_wbs_bac": presupuesto_wbs_bac,
        "cronograma_pv_diario_60dias": cronograma_pv_diario
    },
    "especificacion_formulas_vivas_sheets": {
        "formula_busqueda_pu_meta": "=IFERROR(VLOOKUP(E{row}, MAESTRO_RECURSOS!A:D, 4, FALSE), IFERROR(VLOOKUP(E{row}, PARTIDAS_EV!A:D, 4, FALSE), 0))",
        "formula_subtotal_pen": "=ROUND(G{row} * I{row}, 2)",
        "formula_pv_diario": "=SUMIFS(CRONOGRAMA_PV_DIARIO!C:C, CRONOGRAMA_PV_DIARIO!A:A, FECHA_CORTE, CRONOGRAMA_PV_DIARIO!B:B, WBS_CODE)",
        "formula_pv_acumulado_a_la_fecha": "=SUMIFS(CRONOGRAMA_PV_DIARIO!C:C, CRONOGRAMA_PV_DIARIO!A:A, \"<=\"&FECHA_CORTE, CRONOGRAMA_PV_DIARIO!B:B, WBS_CODE)",
        "formula_ev_acumulado": "=SUMIFS(LOG_FIELD_ENTRIES!J:J, LOG_FIELD_ENTRIES!D:D, WBS_CODE, LOG_FIELD_ENTRIES!K:K, \"EV_PRODUCCION\")",
        "formula_ac_acumulado": "=SUMIFS(LOG_FIELD_ENTRIES!J:J, LOG_FIELD_ENTRIES!D:D, WBS_CODE, LOG_FIELD_ENTRIES!K:K, \"AC_*\")",
        "formula_spi": "=IF(PV_ACUM>0, EV_ACUM/PV_ACUM, 1)",
        "formula_cpi": "=IF(AC_ACUM>0, EV_ACUM/AC_ACUM, 1)"
    },
    "registros_diarios_campo_muestra": field_entries_raw
}

os.makedirs("docs/data", exist_ok=True)
json_sheets_path = "docs/data/base_datos_master_sheets.json"
with open(json_sheets_path, "w", encoding="utf-8") as f:
    json.dump(db_sheets_json, f, ensure_ascii=False, indent=2)

print(f"[OK] Archivo JSON Master con 60 Días de PV creado en: {json_sheets_path}")


# ---------------------------------------------------------
# 4. EXPORTACIÓN A EXCEL CON EL CRONOGRAMA PV DE 60 DÍAS
# ---------------------------------------------------------

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Estilos de Excel
font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
font_section = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True)
font_regular = Font(name="Calibri", size=10)

fill_dark = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
fill_blue = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
fill_green = PatternFill(start_color="15803D", end_color="15803D", fill_type="solid")
fill_orange = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")
fill_purple = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
fill_kpi = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)
double_bottom_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='double', color='1E3A8A')
)


# ---------------------------------------------------------
# PESTAÑA 1: RESULTADO_OPERATIVO (DASHBOARD RO & KPIS DE CORTE)
# ---------------------------------------------------------
ws_ro = wb.create_sheet(title="RESULTADO_OPERATIVO")
ws_ro.views.sheetView[0].showGridLines = True

ws_ro.merge_cells("A1:I2")
ws_ro["A1"] = "RESULTADO OPERATIVO (RO) Y CONTROL EVM - CRONOGRAMA COMPLETO DE 60 DÍAS OBRA"
ws_ro["A1"].font = font_title
ws_ro["A1"].fill = fill_dark
ws_ro["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws_ro["A3"] = "Habilitación Urbana Los Cedros | Línea Base PV de 60 Días vs Registros Reales de Campo"
ws_ro["A3"].font = font_section

# KPIs Fila 5-6 (Corte al Día 5)
kpi_titles = ["Presupuesto BAC (60d)", "Planificado PV (Día 5)", "Valor Ganado EV (Día 5)", "Costo Real AC (Día 5)", "Variación Costo CV", "CPI Costo (EV/AC)", "SPI Plazo (EV/PV)", "EAC Proyectado", "Desvío (EAC - BAC)"]
kpi_forms = [
    "=SUM(CRONOGRAMA_PV_DIARIO!F2:F241)/60",
    "=SUMIFS(CRONOGRAMA_PV_DIARIO!C:C, CRONOGRAMA_PV_DIARIO!A:A, \"<=2026-08-05\")",
    "=SUMIFS(LOG_FIELD_ENTRIES!J:J, LOG_FIELD_ENTRIES!K:K, \"EV_PRODUCCION\")",
    "=SUMIFS(LOG_FIELD_ENTRIES!J:J, LOG_FIELD_ENTRIES!K:K, \"AC_*\")",
    "=C6-D6",
    "=IF(D6>0, C6/D6, 1)",
    "=IF(B6>0, C6/B6, 1)",
    "=IF(F6>0, A6/F6, A6)",
    "=H6-A6"
]
kpi_fmts = ['S/ #,##0.00', 'S/ #,##0.00', 'S/ #,##0.00', 'S/ #,##0.00', 'S/ #,##0.00', '0.00', '0.00', 'S/ #,##0.00', 'S/ #,##0.00']

for c_idx, (t, f, fmt) in enumerate(zip(kpi_titles, kpi_forms, kpi_fmts), start=1):
    ws_ro.cell(row=5, column=c_idx, value=t).font = font_header
    ws_ro.cell(row=5, column=c_idx).fill = fill_blue
    ws_ro.cell(row=5, column=c_idx).alignment = Alignment(horizontal="center")
    
    cell_v = ws_ro.cell(row=6, column=c_idx, value=f)
    cell_v.font = Font(name="Calibri", size=11, bold=True)
    cell_v.fill = fill_kpi
    cell_v.number_format = fmt
    cell_v.alignment = Alignment(horizontal="center")
    cell_v.border = thin_border

# SECCIÓN 2: COMPARATIVA DIARIA EVM (DÍAS 1 AL 5 CON DATA REAL) (Filas 9 a 17)
ws_ro["A9"] = "SEGUIMIENTO DIARIO DE CAMPO: PV PROGRAMADO VS EV EJECUTADO VS AC COSTO REAL"
ws_ro["A9"].font = font_section

headers_daily_evm = ["Fecha", "PV Programado Día (S/)", "PV Acumulado (S/)", "EV Ejecutado Día (S/)", "EV Acumulado (S/)", "AC Costo Real Día (S/)", "AC Acumulado (S/)", "SPI Diario (EV/PV)", "CPI Diario (EV/AC)"]
for c_idx, h in enumerate(headers_daily_evm, start=1):
    cell = ws_ro.cell(row=10, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_purple
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

fechas_list = ["2026-08-01", "2026-08-02", "2026-08-03", "2026-08-04", "2026-08-05"]
for r_d, f_str in enumerate(fechas_list, start=11):
    ws_ro.cell(row=r_d, column=1, value=f_str).font = font_bold
    
    # PV Programado Día
    ws_ro.cell(row=r_d, column=2, value=f"=SUMIF(CRONOGRAMA_PV_DIARIO!A:A, A{r_d}, CRONOGRAMA_PV_DIARIO!C:C)").number_format = 'S/ #,##0.00'
    # PV Acumulado
    ws_ro.cell(row=r_d, column=3, value=f"=SUM(B$11:B{r_d})").number_format = 'S/ #,##0.00'
    
    # EV Ejecutado Día
    ws_ro.cell(row=r_d, column=4, value=f"=SUMIFS(LOG_FIELD_ENTRIES!J:J, LOG_FIELD_ENTRIES!B:B, A{r_d}, LOG_FIELD_ENTRIES!K:K, \"EV_PRODUCCION\")").number_format = 'S/ #,##0.00'
    # EV Acumulado
    ws_ro.cell(row=r_d, column=5, value=f"=SUM(D$11:D{r_d})").number_format = 'S/ #,##0.00'
    
    # AC Costo Real Día
    ws_ro.cell(row=r_d, column=6, value=f"=SUMIFS(LOG_FIELD_ENTRIES!J:J, LOG_FIELD_ENTRIES!B:B, A{r_d}, LOG_FIELD_ENTRIES!K:K, \"AC_*\")").number_format = 'S/ #,##0.00'
    # AC Acumulado
    ws_ro.cell(row=r_d, column=7, value=f"=SUM(F$11:F{r_d})").number_format = 'S/ #,##0.00'
    
    # SPI y CPI Diario Acumulado
    ws_ro.cell(row=r_d, column=8, value=f"=IF(C{r_d}>0, E{r_d}/C{r_d}, 1)").number_format = '0.00'
    ws_ro.cell(row=r_d, column=9, value=f"=IF(G{r_d}>0, E{r_d}/G{r_d}, 1)").number_format = '0.00'

    for c in range(1, 10):
        ws_ro.cell(row=r_d, column=c).border = thin_border

# SECCIÓN 3: RESUMEN EVM POR WBS (AL CORTE DE HOY) (Filas 19+)
ws_ro["A18"] = "DESEMPEÑO ACUMULADO POR FRENTE DE TRABAJO (WBS) AL CORTE DEL DÍA 5"
ws_ro["A18"].font = font_section

headers_ro_wbs = ["Código WBS", "Nombre Frente WBS", "Presupuesto BAC (S/)", "PV Acum al Día 5 (S/)", "AC Costo Real (S/)", "EV Valor Ganado (S/)", "Diferencia (EV - AC)", "CPI Costo", "SPI Plazo", "EAC Proyectado (S/)"]
for c_idx, h in enumerate(headers_ro_wbs, start=1):
    cell = ws_ro.cell(row=19, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_dark
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for r_w, wbs in enumerate(presupuesto_wbs_bac, start=20):
    code = wbs["codigo_wbs"]
    ws_ro.cell(row=r_w, column=1, value=code).font = font_bold
    ws_ro.cell(row=r_w, column=2, value=wbs["nombre_frente"]).font = font_regular
    ws_ro.cell(row=r_w, column=3, value=wbs["bac_meta_pen"]).number_format = 'S/ #,##0.00'
    
    # PV Acumulado al Día 5 (2026-08-05) desde CRONOGRAMA_PV_DIARIO
    ws_ro.cell(row=r_w, column=4, value=f"=SUMIFS(CRONOGRAMA_PV_DIARIO!C:C, CRONOGRAMA_PV_DIARIO!A:A, \"<=2026-08-05\", CRONOGRAMA_PV_DIARIO!B:B, A{r_w})").number_format = 'S/ #,##0.00'
    
    # AC Costo Real
    ws_ro.cell(row=r_w, column=5, value=f"=SUMIFS(LOG_FIELD_ENTRIES!J:J, LOG_FIELD_ENTRIES!D:D, A{r_w}, LOG_FIELD_ENTRIES!K:K, \"AC_*\")").number_format = 'S/ #,##0.00'
    
    # EV Valor Ganado
    ws_ro.cell(row=r_w, column=6, value=f"=SUMIFS(LOG_FIELD_ENTRIES!J:J, LOG_FIELD_ENTRIES!D:D, A{r_w}, LOG_FIELD_ENTRIES!K:K, \"EV_PRODUCCION\")").number_format = 'S/ #,##0.00'
    
    # EV - AC
    ws_ro.cell(row=r_w, column=7, value=f"=F{r_w}-E{r_w}").number_format = 'S/ #,##0.00'
    
    # CPI y SPI
    ws_ro.cell(row=r_w, column=8, value=f"=IF(E{r_w}>0, F{r_w}/E{r_w}, 1)").number_format = '0.00'
    ws_ro.cell(row=r_w, column=9, value=f"=IF(D{r_w}>0, F{r_w}/D{r_w}, 1)").number_format = '0.00'
    
    # EAC Proyectado
    ws_ro.cell(row=r_w, column=10, value=f"=IF(H{r_w}>0, C{r_w}/H{r_w}, C{r_w})").number_format = 'S/ #,##0.00'
    
    for c in range(1, 11):
        ws_ro.cell(row=r_w, column=c).border = thin_border

# Totales WBS
ws_ro.cell(row=24, column=1, value="TOTAL OBRA").font = font_bold
ws_ro.cell(row=24, column=2, value="Consolidado Habilitación Urbana").font = font_regular
for c_i, col_let in enumerate(["C", "D", "E", "F", "G"], start=3):
    cell = ws_ro.cell(row=24, column=c_i, value=f"=SUM({col_let}20:{col_let}23)")
    cell.font = font_bold
    cell.number_format = 'S/ #,##0.00'

ws_ro.cell(row=24, column=8, value="=IF(E24>0, F24/E24, 1)").font = font_bold
ws_ro.cell(row=24, column=8).number_format = '0.00'

ws_ro.cell(row=24, column=9, value="=IF(D24>0, F24/D24, 1)").font = font_bold
ws_ro.cell(row=24, column=9).number_format = '0.00'

ws_ro.cell(row=24, column=10, value="=IF(H24>0, C24/H24, C24)").font = font_bold
ws_ro.cell(row=24, column=10).number_format = 'S/ #,##0.00'

for c in range(1, 11):
    ws_ro.cell(row=24, column=c).border = double_bottom_border


# ---------------------------------------------------------
# PESTAÑA 2: CRONOGRAMA_PV_DIARIO (LÍNEA BASE COMPLETA 60 DÍAS)
# ---------------------------------------------------------
ws_pv_d = wb.create_sheet(title="CRONOGRAMA_PV_DIARIO")
ws_pv_d.views.sheetView[0].showGridLines = True

headers_pv_d = ["Fecha", "Código WBS", "PV Diario Programado (S/)", "PV Acumulado Programado (S/)", "N° Día Obra", "Nombre Frente WBS", "Presupuesto BAC Total (S/)"]
for c_idx, h in enumerate(headers_pv_d, start=1):
    cell = ws_pv_d.cell(row=1, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_dark
    cell.alignment = Alignment(horizontal="center")

bac_map = {b["codigo_wbs"]: (b["nombre_frente"], b["bac_meta_pen"]) for b in presupuesto_wbs_bac}

for r_idx, row in enumerate(cronograma_pv_diario, start=2):
    w_code = row["codigo_wbs"]
    w_name, w_bac = bac_map.get(w_code, ("", 0.0))
    
    ws_pv_d.cell(row=r_idx, column=1, value=row["fecha"]).font = font_regular
    ws_pv_d.cell(row=r_idx, column=2, value=w_code).font = font_bold
    
    cell_pv_d = ws_pv_d.cell(row=r_idx, column=3, value=row["pv_diario_pen"])
    cell_pv_d.font = font_bold
    cell_pv_d.number_format = 'S/ #,##0.00'
    
    cell_pv_a = ws_pv_d.cell(row=r_idx, column=4, value=row["pv_acumulado_pen"])
    cell_pv_a.font = font_regular
    cell_pv_a.number_format = 'S/ #,##0.00'
    
    ws_pv_d.cell(row=r_idx, column=5, value=row["dia_numero"]).font = font_regular
    ws_pv_d.cell(row=r_idx, column=6, value=w_name).font = font_regular
    
    cell_bac = ws_pv_d.cell(row=r_idx, column=7, value=w_bac)
    cell_bac.font = font_regular
    cell_bac.number_format = 'S/ #,##0.00'

    for c in range(1, 8):
        ws_pv_d.cell(row=r_idx, column=c).border = thin_border


# ---------------------------------------------------------
# PESTAÑA 3: LOG_FIELD_ENTRIES (REPORTE DIARIO DESDE CELULARES)
# ---------------------------------------------------------
ws_log = wb.create_sheet(title="LOG_FIELD_ENTRIES")
ws_log.views.sheetView[0].showGridLines = True

headers_log = [
    "ID Registro", "Fecha", "Rol Responsable", "Código WBS", 
    "Código Recurso/Partida", "Descripción / Detalle", "Cantidad Campo", 
    "Unidad", "P.U. (Busca en Maestro)", "Subtotal Monto (S/)", 
    "Categoría EVM", "Origen HTML"
]

for c_idx, h in enumerate(headers_log, start=1):
    cell = ws_log.cell(row=1, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = Alignment(horizontal="center", vertical="center")

for r_idx, reg in enumerate(field_entries_raw, start=2):
    ws_log.cell(row=r_idx, column=1, value=reg["id"]).font = font_regular
    ws_log.cell(row=r_idx, column=2, value=reg["fecha"]).font = font_regular
    ws_log.cell(row=r_idx, column=3, value=reg["rol"]).font = font_regular
    ws_log.cell(row=r_idx, column=4, value=reg["wbs"]).font = font_bold
    ws_log.cell(row=r_idx, column=5, value=reg["recurso"]).font = font_bold
    ws_log.cell(row=r_idx, column=6, value=reg["detalle"]).font = font_regular
    
    cell_q = ws_log.cell(row=r_idx, column=7, value=reg["cantidad"])
    cell_q.font = font_bold
    cell_q.number_format = '#,##0.00'
    
    ws_log.cell(row=r_idx, column=8, value=reg["unidad"]).font = font_regular
    
    # FÓRMULA VIVA DE BÚSQUEDA DE P.U. EN MAESTRO DE RECURSOS O PARTIDAS EV
    formula_pu = f"=IFERROR(VLOOKUP(E{r_idx}, MAESTRO_RECURSOS!A:D, 4, FALSE), IFERROR(VLOOKUP(E{r_idx}, PARTIDAS_EV!A:D, 4, FALSE), 0))"
    cell_pu = ws_log.cell(row=r_idx, column=9, value=formula_pu)
    cell_pu.font = font_regular
    cell_pu.number_format = 'S/ #,##0.00'
    
    # FÓRMULA VIVA DE MONTO (Cantidad Campo * PU Maestro)
    formula_subtotal = f"=ROUND(G{r_idx} * I{r_idx}, 2)"
    cell_sub = ws_log.cell(row=r_idx, column=10, value=formula_subtotal)
    cell_sub.font = font_bold
    cell_sub.number_format = 'S/ #,##0.00'
    
    ws_log.cell(row=r_idx, column=11, value=reg["cat"]).font = font_bold
    ws_log.cell(row=r_idx, column=12, value=reg["html"]).font = font_regular

    for c in range(1, 13):
        ws_log.cell(row=r_idx, column=c).border = thin_border


# ---------------------------------------------------------
# PESTAÑA 4: MAESTRO_RECURSOS (DICCIONARIO DE PRECIOS UNITARIOS DE OBRA)
# ---------------------------------------------------------
ws_rec = wb.create_sheet(title="MAESTRO_RECURSOS")
ws_rec.views.sheetView[0].showGridLines = True

ws_rec.append(["Código Recurso", "Descripción del Recurso / Insumo", "Unidad", "P.U. Meta Oficial (S/)", "Categoría EVM"])
for cell in ws_rec[1]:
    cell.font = font_header
    cell.fill = fill_purple
    cell.alignment = Alignment(horizontal="center")

for row in maestro_recursos:
    ws_rec.append([row["codigo"], row["descripcion"], row["unidad"], row["precio_unitario"], row["categoria"]])

for r in range(2, len(maestro_recursos) + 2):
    ws_rec.cell(row=r, column=4).number_format = 'S/ #,##0.00'
    for c in range(1, 6):
        ws_rec.cell(row=r, column=c).border = thin_border


# ---------------------------------------------------------
# PESTAÑA 5: PARTIDAS_EV (DICCIONARIO DE PARTIDAS Y P.U. DE PRODUCCIÓN)
# ---------------------------------------------------------
ws_part = wb.create_sheet(title="PARTIDAS_EV")
ws_part.views.sheetView[0].showGridLines = True

ws_part.append(["Código Partida", "Descripción de la Partida", "Unidad", "P.U. Directo Meta (S/)", "Categoría EVM"])
for cell in ws_part[1]:
    cell.font = font_header
    cell.fill = fill_green
    cell.alignment = Alignment(horizontal="center")

for row in maestro_partidas_ev:
    ws_part.append([row["codigo"], row["descripcion"], row["unidad"], row["precio_unitario"], row["categoria"]])

for r in range(2, len(maestro_partidas_ev) + 2):
    ws_part.cell(row=r, column=4).number_format = 'S/ #,##0.00'
    for c in range(1, 6):
        ws_part.cell(row=r, column=c).border = thin_border


# ---------------------------------------------------------
# AJUSTE DE ANCHO DE COLUMNAS EN TODAS LAS HOJAS
# ---------------------------------------------------------
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in sheet.merged_cells:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 55)

os.makedirs("outputs", exist_ok=True)
excel_sheets_path = "outputs/Base_Datos_Proyecto_Sheets_Viva.xlsx"
wb.save(excel_sheets_path)

print(f"[OK] Archivo Excel de Base de Datos Viva con Cronograma Completo de 60 Días creado en: {excel_sheets_path}")
