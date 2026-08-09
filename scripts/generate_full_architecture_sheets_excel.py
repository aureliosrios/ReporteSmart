import json
import os
import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------
# 1. CATALOGOS Y LÍNEA BASE DEL PROYECTO
# ---------------------------------------------------------

metadata = {
    "proyecto": "Redes Sanitarias de Agua Potable y Alcantarillado - Habilitación Urbana Los Cedros",
    "codigo_proyecto": "HU-CEDROS-2026",
    "version_esquema": "7.0.0",
    "moneda": "PEN",
    "duracion_dias_calendario": 60,
    "fecha_inicio_obra": "2026-08-01",
    "fecha_fin_programada": "2026-09-29",
    "descripcion": "Base de datos viva completa en Google Sheets. Pestaña de Presupuesto y Cronograma Base de referencia, PV diario armado con Fórmulas Vivas, y una Base de Datos Diaria Independiente de EVM por WBS para visualización diaria del proyecto."
}

# Presupuesto Total y Cronograma Base de Referencia
referencia_base_presupuesto = [
    {"wbs": "WBS-100", "nombre": "Obras Preliminares, Trazo y Movilización", "bac_pen": 8296.40, "dia_inicio": 1, "dia_fin": 14},
    {"wbs": "WBS-200", "nombre": "Red de Alcantarillado, Zanjas y Buzones", "bac_pen": 380086.56, "dia_inicio": 10, "dia_fin": 43},
    {"wbs": "WBS-300", "nombre": "Red de Agua Potable y Conexiones Domiciliarias", "bac_pen": 57565.20, "dia_inicio": 22, "dia_fin": 51},
    {"wbs": "WBS-400", "nombre": "Pruebas Hidráulicas, Desinfección y Entrega", "bac_pen": 9348.06, "dia_inicio": 48, "dia_fin": 60}
]

# Maestro de Recursos (Precios Unitarios Oficiales)
maestro_recursos = [
    {"codigo": "MO_CAPATAZ", "descripcion": "CAPATAZ", "unidad": "hh", "precio_unitario": 32.50, "categoria": "AC_MO"},
    {"codigo": "MO_OPERARIO", "descripcion": "OPERARIO", "unidad": "hh", "precio_unitario": 26.80, "categoria": "AC_MO"},
    {"codigo": "MO_OFICIAL", "descripcion": "OFICIAL", "unidad": "hh", "precio_unitario": 22.40, "categoria": "AC_MO"},
    {"codigo": "MO_PEON", "descripcion": "PEON", "unidad": "hh", "precio_unitario": 20.10, "categoria": "AC_MO"},
    {"codigo": "MAT_ESTACA", "descripcion": "ESTACAS Y MADERA PARA TRAZO", "unidad": "pza", "precio_unitario": 4.50, "categoria": "AC_MAT"},
    {"codigo": "MAT_PINTURA", "descripcion": "PINTURA SPRAY MARCADOR DE ZANJA", "unidad": "gla", "precio_unitario": 18.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_CORDEL", "descripcion": "CORDEL / NYLON", "unidad": "m", "precio_unitario": 0.80, "categoria": "AC_MAT"},
    {"codigo": "MAT_ARENA_CAMAS", "descripcion": "ARENA FINA SELECCIONADA PARA CAMA E=0.10M", "unidad": "m3", "precio_unitario": 45.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_TUB_PVC_200", "descripcion": "TUBERIA PVC UF DN 200MM SERIE S-20 ALCANTARILLADO", "unidad": "m", "precio_unitario": 42.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_LUBRICANTE", "descripcion": "LUBRICANTE PARA TUBERIA CON ESPIGA Y CAMPANA", "unidad": "kg", "precio_unitario": 24.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_MATERIAL_PR", "descripcion": "MATERIAL DE PRÉSTAMO AFIRMADO PARA RELLENO COMPACTADO", "unidad": "m3", "precio_unitario": 38.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_AGUA", "descripcion": "AGUA PARA COMPACTACIÓN Y PRUEBAS", "unidad": "m3", "precio_unitario": 12.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_BUZON_PREF", "descripcion": "BUZON PREFABRICADO CONCRETO H=1.5-2.5M CON MARCO Y TAPA", "unidad": "und", "precio_unitario": 1050.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_CAJA_REGISTRO", "descripcion": "CAJA DE REGISTRO PREFABRICADA DE CONCRETO 12X24 WITH TAPA", "unidad": "und", "precio_unitario": 145.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_TUB_PVC_160_CONEX", "descripcion": "TUBERIA PVC SAL DN 160MM PARA ACOMETIDA ALCANTARILLADO", "unidad": "m", "precio_unitario": 22.00, "categoria": "AC_MAT"},
    {"codigo": "MAT_TUB_PVC_110_AGUA", "descripcion": "TUBERIA PVC C-10 DN 110MM AGUA POTABLE", "unidad": "m", "precio_unitario": 26.50, "categoria": "AC_MAT"},
    {"codigo": "MAT_CAJA_AGUA", "descripcion": "CAJA TERMOFORMADA PARA MEDIDOR + ABRAZADERA + LLAVE DE PASO", "unidad": "und", "precio_unitario": 62.00, "categoria": "AC_MAT"},
    {"codigo": "EQ_EXCAVADORA", "descripcion": "EXCAVADORA SOBRE ORUGAS 140 HP", "unidad": "hm", "precio_unitario": 185.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_RETROEXCAVADORA", "descripcion": "RETROEXCAVADORA SOBRE LLANTAS 62 HP", "unidad": "hm", "precio_unitario": 130.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_PLANCHA", "descripcion": "PLANCHA COMPACTADORA 7 HP", "unidad": "hm", "precio_unitario": 25.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_RODILLO_CAMINANTE", "descripcion": "RODILLO VIBRATORIO CAMINANTE 1.5 TN", "unidad": "hm", "precio_unitario": 45.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_CAMION_CISTERNA", "descripcion": "CAMION CISTERNA 4X2 (AGUA) 2,000 GLN", "unidad": "hm", "precio_unitario": 140.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_CAMION_GRUAN", "descripcion": "CAMION GRUA 5 TN PARA MONTAJE DE BUZONES", "unidad": "hm", "precio_unitario": 160.00, "categoria": "AC_EQP"},
    {"codigo": "EQ_ESTACION_TOTAL", "descripcion": "EQUIPO DE TOPOGRAFIA ESTACION TOTAL", "unidad": "hm", "precio_unitario": 18.00, "categoria": "AC_EQP"},
    {"codigo": "SUB_MOVILIZACION", "descripcion": "SUBCONTRATO DE MOVILIZACION Y DESMOVILIZACION DE MAQUINARIA", "unidad": "glb", "precio_unitario": 4000.00, "categoria": "AC_SUB"},
    {"codigo": "SUB_ENSAYOS_COMPAC", "descripcion": "SUBCONTRATO DE ENSAYOS DE DENSIDAD DE CAMPO (PROCTOR/DENSIDAD)", "unidad": "und", "precio_unitario": 80.00, "categoria": "AC_SUB"},
    {"codigo": "SUB_MONTAJE_BUZON", "descripcion": "SUBCONTRATO SERVICIO ASENTADO Y ANCLAJE BUZONES PREFABRICADOS", "unidad": "und", "precio_unitario": 250.00, "categoria": "AC_SUB"}
]

# Maestro de Partidas Presupuestales (Precios Unitarios Directos Meta)
maestro_partidas_ev = [
    {"codigo": "01.01", "descripcion": "Obras Preliminares y Trabajos Provisionales", "unidad": "GLB", "precio_unitario": 8296.40, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.01", "descripcion": "Trazo, Nivelación y Replanteo de Zanjas", "unidad": "M", "precio_unitario": 3.35, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.02", "descripcion": "Excavación de Zanja H=1.50m - 2.20m a Máquina (Terreno Normal)", "unidad": "M3", "precio_unitario": 11.66, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.03", "descripcion": "Preparación y Colocación de Cama de Arena e=0.10m", "unidad": "M", "precio_unitario": 11.12, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.04", "descripcion": "Suministro e Instalación de Tubería PVC UF DN 200mm Serie S-20 para Alcantarillado", "unidad": "M", "precio_unitario": 55.51, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.05", "descripcion": "Relleno Compactado de Zanja en Capas de 0.20m con Maquinaria/Plancha", "unidad": "M3", "precio_unitario": 28.54, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.06", "descripcion": "Construcción de Buzones Prefabricados de Concreto h=1.50m - 2.50m (Inc. Marco y Tapa)", "unidad": "UND", "precio_unitario": 2500.73, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.02.07", "descripcion": "Conexiones Domiciliarias de Alcantarillado (Caja de Registro + Acometida)", "unidad": "UND", "precio_unitario": 405.54, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.03.01", "descripcion": "Suministro e Instalación de Tubería PVC C-10 DN 110mm para Agua Potable", "unidad": "M", "precio_unitario": 38.76, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.03.02", "descripcion": "Conexiones Domiciliarias de Agua Potable (Caja de Agua + Abrazadera + Acometida)", "unidad": "UND", "precio_unitario": 156.71, "categoria": "EV_PRODUCCION"},
    {"codigo": "01.04.01", "descripcion": "Pruebas Hidráulicas de Redes de Agua y Alcantarillado + Desinfección", "unidad": "GLB", "precio_unitario": 9348.06, "categoria": "EV_PRODUCCION"}
]

# Entradas Diarias de Campo (Muestra de 5 Días enviadas desde Celulares)
field_entries_raw = [
    # DÍA 1 (2026-08-01)
    {"id": "LOG-20260801-001", "fecha": "2026-08-01", "rol": "Tareador (Bildin)", "wbs": "WBS-100", "recurso": "MO_CAPATAZ", "detalle": "CAPATAZ (1 pers. x 8h)", "cantidad": 8.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260801-002", "fecha": "2026-08-01", "rol": "Tareador (Bildin)", "wbs": "WBS-100", "recurso": "MO_OPERARIO", "detalle": "OPERARIO (2 pers. x 8h)", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260801-003", "fecha": "2026-08-01", "rol": "Tareador (Bildin)", "wbs": "WBS-100", "recurso": "MO_PEON", "detalle": "PEON (4 pers. x 8h)", "cantidad": 32.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260801-004", "fecha": "2026-08-01", "rol": "Almacenero", "wbs": "WBS-100", "recurso": "MAT_ESTACA", "detalle": "ESTACAS Y MADERA PARA TRAZO", "cantidad": 150.0, "unidad": "pza", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260801-005", "fecha": "2026-08-01", "rol": "Almacenero", "wbs": "WBS-100", "recurso": "MAT_PINTURA", "detalle": "PINTURA SPRAY MARCADOR ZANJA", "cantidad": 10.0, "unidad": "gla", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260801-006", "fecha": "2026-08-01", "rol": "Administradora", "wbs": "WBS-100", "recurso": "SUB_MOVILIZACION", "detalle": "MOVILIZACION DE MAQUINARIA", "cantidad": 1.0, "unidad": "glb", "cat": "AC_SUB", "html": "administradora.html"},
    {"id": "LOG-20260801-007", "fecha": "2026-08-01", "rol": "Administradora", "wbs": "WBS-100", "recurso": "EQ_ESTACION_TOTAL", "detalle": "ESTACION TOTAL TOPOGRAFICA", "cantidad": 8.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260801-008", "fecha": "2026-08-01", "rol": "Ing. de Campo", "wbs": "WBS-100", "recurso": "01.01", "detalle": "OBRAS PRELIMINARES Y PROVISIONALES (50%)", "cantidad": 0.5, "unidad": "GLB", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260801-009", "fecha": "2026-08-01", "rol": "Ing. de Campo", "wbs": "WBS-100", "recurso": "01.02.01", "detalle": "TRAZO Y REPLANTEO INICIAL TRAMO T-1", "cantidad": 500.0, "unidad": "M", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},

    # DÍA 2 (2026-08-02)
    {"id": "LOG-20260802-010", "fecha": "2026-08-02", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_CAPATAZ", "detalle": "CAPATAZ CONTROL ZANJA", "cantidad": 8.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260802-011", "fecha": "2026-08-02", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_OPERARIO", "detalle": "OPERARIO ENTIBADO Y CAMA", "cantidad": 24.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260802-012", "fecha": "2026-08-02", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_PEON", "detalle": "PEON REPELE Y CAMA", "cantidad": 48.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260802-013", "fecha": "2026-08-02", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_ARENA_CAMAS", "detalle": "ARENA FINA SELECCIONADA PARA CAMA", "cantidad": 35.0, "unidad": "m3", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260802-014", "fecha": "2026-08-02", "rol": "Administradora", "wbs": "WBS-200", "recurso": "EQ_EXCAVADORA", "detalle": "EXCAVADORA SOBRE ORUGAS 140 HP", "cantidad": 8.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260802-015", "fecha": "2026-08-02", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.02", "detalle": "EXCAVACION ZANJA H=1.5-2.2M A MAQUINA", "cantidad": 450.0, "unidad": "M3", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260802-016", "fecha": "2026-08-02", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.03", "detalle": "COLOCACION CAMA DE ARENA E=0.10M", "cantidad": 300.0, "unidad": "M", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},

    # DÍA 3 (2026-08-03)
    {"id": "LOG-20260803-017", "fecha": "2026-08-03", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_OPERARIO", "detalle": "OPERARIO TUBERO ALCANTARILLADO", "cantidad": 24.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260803-018", "fecha": "2026-08-03", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_OFICIAL", "detalle": "OFICIAL COMPACTACION", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260803-019", "fecha": "2026-08-03", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_PEON", "detalle": "PEON APOYO MONTAJE TUBERIA", "cantidad": 40.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260803-020", "fecha": "2026-08-03", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_TUB_PVC_200", "detalle": "TUBERIA PVC UF DN 200MM S-20", "cantidad": 280.0, "unidad": "m", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260803-021", "fecha": "2026-08-03", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_LUBRICANTE", "detalle": "LUBRICANTE PARA TUBERIA PVC", "cantidad": 12.0, "unidad": "kg", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260803-022", "fecha": "2026-08-03", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_MATERIAL_PR", "detalle": "MATERIAL AFIRMADO RELLENO COMPACTADO", "cantidad": 60.0, "unidad": "m3", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260803-023", "fecha": "2026-08-03", "rol": "Administradora", "wbs": "WBS-200", "recurso": "EQ_RETROEXCAVADORA", "detalle": "RETROEXCAVADORA RELLENO Y LIMPIEZA", "cantidad": 8.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260803-024", "fecha": "2026-08-03", "rol": "Administradora", "wbs": "WBS-200", "recurso": "EQ_PLANCHA", "detalle": "PLANCHA COMPACTADORA 7 HP", "cantidad": 8.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260803-025", "fecha": "2026-08-03", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.04", "detalle": "INSTALACION TUBERIA PVC 200MM ALCANTARILLADO", "cantidad": 250.0, "unidad": "M", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260803-026", "fecha": "2026-08-03", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.05", "detalle": "RELLENO COMPACTADO ZANJA CAPAS 0.20M", "cantidad": 200.0, "unidad": "M3", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},

    # DÍA 4 (2026-08-04)
    {"id": "LOG-20260804-027", "fecha": "2026-08-04", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_OPERARIO", "detalle": "OPERARIO MONTAJE BUZONES", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260804-028", "fecha": "2026-08-04", "rol": "Tareador (Bildin)", "wbs": "WBS-300", "recurso": "MO_OPERARIO", "detalle": "OPERARIO TUBERO AGUA POTABLE", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260804-029", "fecha": "2026-08-04", "rol": "Tareador (Bildin)", "wbs": "WBS-300", "recurso": "MO_PEON", "detalle": "PEON APERTURA ZANJA AGUA", "cantidad": 32.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260804-030", "fecha": "2026-08-04", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_BUZON_PREF", "detalle": "BUZON PREFABRICADO DE CONCRETO H=2.0M", "cantidad": 4.0, "unidad": "und", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260804-031", "fecha": "2026-08-04", "rol": "Almacenero", "wbs": "WBS-300", "recurso": "MAT_TUB_PVC_110_AGUA", "detalle": "TUBERIA PVC C-10 DN 110MM AGUA", "cantidad": 200.0, "unidad": "m", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260804-032", "fecha": "2026-08-04", "rol": "Administradora", "wbs": "WBS-200", "recurso": "EQ_CAMION_GRUAN", "detalle": "CAMION GRUA 5 TN MONTAJE BUZONES", "cantidad": 6.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260804-033", "fecha": "2026-08-04", "rol": "Administradora", "wbs": "WBS-200", "recurso": "SUB_MONTAJE_BUZON", "detalle": "SUBCONTRATO ASENTADO Y SELLADO BUZONES", "cantidad": 4.0, "unidad": "und", "cat": "AC_SUB", "html": "administradora.html"},
    {"id": "LOG-20260804-034", "fecha": "2026-08-04", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.06", "detalle": "CONSTRUCCION BUZONES PREFABRICADOS CONCRETO", "cantidad": 4.0, "unidad": "UND", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260804-035", "fecha": "2026-08-04", "rol": "Ing. de Campo", "wbs": "WBS-300", "recurso": "01.03.01", "detalle": "INSTALACION TUBERIA PVC C-10 DN 110MM AGUA", "cantidad": 180.0, "unidad": "M", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},

    # DÍA 5 (2026-08-05)
    {"id": "LOG-20260805-036", "fecha": "2026-08-05", "rol": "Tareador (Bildin)", "wbs": "WBS-200", "recurso": "MO_OPERARIO", "detalle": "OPERARIO CONEXION DOMICILIARIA ALCANTARILLADO", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260805-037", "fecha": "2026-08-05", "rol": "Tareador (Bildin)", "wbs": "WBS-300", "recurso": "MO_OPERARIO", "detalle": "OPERARIO CONEXION DOMICILIARIA AGUA", "cantidad": 16.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260805-038", "fecha": "2026-08-05", "rol": "Tareador (Bildin)", "wbs": "WBS-400", "recurso": "MO_OFICIAL", "detalle": "OFICIAL PRUEBAS HIDRAULICAS DE RED", "cantidad": 8.0, "unidad": "hh", "cat": "AC_MO", "html": "tareador.html"},
    {"id": "LOG-20260805-039", "fecha": "2026-08-05", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_CAJA_REGISTRO", "detalle": "CAJA DE REGISTRO PREFABRICADA 12X24", "cantidad": 15.0, "unidad": "und", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260805-040", "fecha": "2026-08-05", "rol": "Almacenero", "wbs": "WBS-200", "recurso": "MAT_TUB_PVC_160_CONEX", "detalle": "TUBERIA PVC SAL DN 160MM ACOMETIDA", "cantidad": 90.0, "unidad": "m", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260805-041", "fecha": "2026-08-05", "rol": "Almacenero", "wbs": "WBS-300", "recurso": "MAT_CAJA_AGUA", "detalle": "CAJA TERMOFORMADA MEDIDOR + ABRAZADERA", "cantidad": 15.0, "unidad": "und", "cat": "AC_MAT", "html": "almacenero.html"},
    {"id": "LOG-20260805-042", "fecha": "2026-08-05", "rol": "Administradora", "wbs": "WBS-200", "recurso": "SUB_ENSAYOS_COMPAC", "detalle": "SUBCONTRATO ENSAYOS DENSIDAD DE CAMPO", "cantidad": 10.0, "unidad": "und", "cat": "AC_SUB", "html": "administradora.html"},
    {"id": "LOG-20260805-043", "fecha": "2026-08-05", "rol": "Administradora", "wbs": "WBS-400", "recurso": "EQ_CAMION_CISTERNA", "detalle": "CAMION CISTERNA AGUA PRUEBAS HIDRAULICAS", "cantidad": 8.0, "unidad": "hm", "cat": "AC_EQP", "html": "administradora.html"},
    {"id": "LOG-20260805-044", "fecha": "2026-08-05", "rol": "Ing. de Campo", "wbs": "WBS-200", "recurso": "01.02.07", "detalle": "CONEXIONES DOMICILIARIAS ALCANTARILLADO (15 ACOMETIDAS)", "cantidad": 15.0, "unidad": "UND", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260805-045", "fecha": "2026-08-05", "rol": "Ing. de Campo", "wbs": "WBS-300", "recurso": "01.03.02", "detalle": "CONEXIONES DOMICILIARIAS AGUA POTABLE (15 ACOMETIDAS)", "cantidad": 15.0, "unidad": "UND", "cat": "EV_PRODUCCION", "html": "ing_campo.html"},
    {"id": "LOG-20260805-046", "fecha": "2026-08-05", "rol": "Ing. de Campo", "wbs": "WBS-400", "recurso": "01.04.01", "detalle": "PRUEBA HIDRAULICA TRAMO 1 REDES (25%)", "cantidad": 0.25, "unidad": "GLB", "cat": "EV_PRODUCCION", "html": "ing_campo.html"}
]

# ---------------------------------------------------------
# GUARDAR ARCHIVO JSON MASTER DE INTEGRACIÓN
# ---------------------------------------------------------

db_master_json = {
    "metadata": metadata,
    "linea_base_referencia": referencia_base_presupuesto,
    "maestros_precios_unitarios": {
        "recursos": maestro_recursos,
        "partidas_ev": maestro_partidas_ev
    },
    "registros_diarios_campo_muestra": field_entries_raw
}

os.makedirs("docs/data", exist_ok=True)
json_sheets_path = "docs/data/base_datos_master_sheets.json"
with open(json_sheets_path, "w", encoding="utf-8") as f:
    json.dump(db_master_json, f, ensure_ascii=False, indent=2)

print(f"[OK] Archivo JSON Master creado en: {json_sheets_path}")


# ---------------------------------------------------------
# CONSTRUCCIÓN DEL LIBRO EXCEL CON FÓRMULAS VIVAS INTERCONECTADAS
# ---------------------------------------------------------

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Estilos de Excel
font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
font_section = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True)
font_regular = Font(name="Calibri", size=10)

fill_dark = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
fill_blue = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
fill_green = PatternFill(start_color="15803D", end_color="15803D", fill_type="solid")
fill_orange = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")
fill_purple = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
fill_kpi = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)
double_bottom_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='double', color='1E3A8A')
)


# ---------------------------------------------------------
# PESTAÑA 1: 01_PRESUPUESTO_CRONOGRAMA_BASE (REFERENCIA BASE DEL PROYECTO)
# ---------------------------------------------------------
ws_base = wb.create_sheet(title="01_PRESUPUESTO_CRONOGRAMA_BASE")
ws_base.views.sheetView[0].showGridLines = True

ws_base.merge_cells("A1:G2")
ws_base["A1"] = "LINEA BASE DEL PROYECTO: PRESUPUESTO META Y CRONOGRAMA DE EJECUCIÓN"
ws_base["A1"].font = font_title
ws_base["A1"].fill = fill_dark
ws_base["A1"].alignment = Alignment(horizontal="center", vertical="center")

headers_base = ["Código WBS", "Nombre del Frente / WBS", "Presupuesto BAC (S/)", "Día Inicio", "Día Fin", "Duración (Días)", "Tasa PV Diario (S/Día)"]
for c_idx, h in enumerate(headers_base, start=1):
    cell = ws_base.cell(row=4, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = Alignment(horizontal="center")

for r_idx, row in enumerate(referencia_base_presupuesto, start=5):
    ws_base.cell(row=r_idx, column=1, value=row["wbs"]).font = font_bold
    ws_base.cell(row=r_idx, column=2, value=row["nombre"]).font = font_regular
    
    cell_bac = ws_base.cell(row=r_idx, column=3, value=row["bac_pen"])
    cell_bac.font = font_bold
    cell_bac.number_format = 'S/ #,##0.00'
    
    ws_base.cell(row=r_idx, column=4, value=row["dia_inicio"]).font = font_regular
    ws_base.cell(row=r_idx, column=5, value=row["dia_fin"]).font = font_regular
    
    # FÓRMULA VIVA DE DURACIÓN
    ws_base.cell(row=r_idx, column=6, value=f"=(E{r_idx}-D{r_idx})+1").font = font_regular
    
    # FÓRMULA VIVA DE TASA PV DIARIO (BAC / Duración)
    cell_tasa = ws_base.cell(row=r_idx, column=7, value=f"=ROUND(C{r_idx}/F{r_idx}, 2)")
    cell_tasa.font = font_bold
    cell_tasa.number_format = 'S/ #,##0.00'

    for c in range(1, 8):
        ws_base.cell(row=r_idx, column=c).border = thin_border

# Fila Total Base
ws_base.cell(row=9, column=1, value="TOTAL PRESUPUESTO META (BAC)").font = font_bold
ws_base.cell(row=9, column=3, value="=SUM(C5:C8)").font = font_bold
ws_base.cell(row=9, column=3).number_format = 'S/ #,##0.00'
for c in range(1, 8):
    ws_base.cell(row=9, column=c).border = double_bottom_border


# ---------------------------------------------------------
# PESTAÑA 2: 02_CRONOGRAMA_PV_DIARIO (MATRIZ CON FÓRMULAS VIVAS DE PV)
# ---------------------------------------------------------
ws_pv_matriz = wb.create_sheet(title="02_CRONOGRAMA_PV_DIARIO")
ws_pv_matriz.views.sheetView[0].showGridLines = True

headers_pv_m = ["N° Día Obra", "Fecha Obra", "Código WBS", "PV Diario Programado (S/)", "PV Acumulado WBS (S/)", "Nombre Frente WBS"]
for c_idx, h in enumerate(headers_pv_m, start=1):
    cell = ws_pv_matriz.cell(row=1, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_purple
    cell.alignment = Alignment(horizontal="center")

start_date = datetime.strptime("2026-08-01", "%Y-%m-%d")
curr_row_pv = 2

for day_n in range(1, 61):
    f_str = (start_date + timedelta(days=day_n - 1)).strftime("%Y-%m-%d")
    for r_base_idx, wbs_ref in enumerate(referencia_base_presupuesto, start=5):
        w_code = wbs_ref["wbs"]
        
        ws_pv_matriz.cell(row=curr_row_pv, column=1, value=day_n).font = font_regular
        ws_pv_matriz.cell(row=curr_row_pv, column=2, value=f_str).font = font_regular
        ws_pv_matriz.cell(row=curr_row_pv, column=3, value=w_code).font = font_bold
        
        # FÓRMULA VIVA DE PV DIARIO: Si Día está entre Inicio y Fin, aplica Tasa Diario de Referencia Base
        formula_pv_diario_viva = f"=IF(AND(A{curr_row_pv}>='01_PRESUPUESTO_CRONOGRAMA_BASE'!D{r_base_idx}, A{curr_row_pv}<='01_PRESUPUESTO_CRONOGRAMA_BASE'!E{r_base_idx}), '01_PRESUPUESTO_CRONOGRAMA_BASE'!G{r_base_idx}, 0)"
        cell_pv_d = ws_pv_matriz.cell(row=curr_row_pv, column=4, value=formula_pv_diario_viva)
        cell_pv_d.font = font_bold
        cell_pv_d.number_format = 'S/ #,##0.00'
        
        # FÓRMULA VIVA DE PV ACUMULADO POR WBS
        formula_pv_acum_viva = f"=SUMIFS(D$2:D{curr_row_pv}, C$2:C{curr_row_pv}, C{curr_row_pv})"
        cell_pv_a = ws_pv_matriz.cell(row=curr_row_pv, column=5, value=formula_pv_acum_viva)
        cell_pv_a.font = font_regular
        cell_pv_a.number_format = 'S/ #,##0.00'
        
        # Nombre Frente
        ws_pv_matriz.cell(row=curr_row_pv, column=6, value=f"='01_PRESUPUESTO_CRONOGRAMA_BASE'!B{r_base_idx}").font = font_regular

        for c in range(1, 7):
            ws_pv_matriz.cell(row=curr_row_pv, column=c).border = thin_border
            
        curr_row_pv += 1


# ---------------------------------------------------------
# PESTAÑA 3: 03_CONSOLIDADO_DIARIO_EVM_WBS (BASE INDEPENDIENTE DIARIA SEPARADA POR WBS)
# ---------------------------------------------------------
ws_evm_diario = wb.create_sheet(title="03_CONSOLIDADO_DIARIO_EVM_WBS")
ws_evm_diario.views.sheetView[0].showGridLines = True

headers_evm_d = [
    "N° Día Obra", "Fecha Obra", "Código WBS", "Nombre Frente WBS", 
    "Presupuesto BAC (S/)", "PV Programado Día (S/)", "PV Acumulado (S/)", 
    "EV Ejecutado Día (S/)", "EV Acumulado (S/)", "AC Costo Real Día (S/)", 
    "AC Acumulado (S/)", "Variación Costo CV (S/)", "Variación Plazo SV (S/)", 
    "SPI Plazo (EV/PV)", "CPI Costo (EV/AC)", "Estado Alerta WBS"
]

for c_idx, h in enumerate(headers_evm_d, start=1):
    cell = ws_evm_diario.cell(row=1, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_dark
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

curr_row_evm = 2
for day_n in range(1, 61):
    f_str = (start_date + timedelta(days=day_n - 1)).strftime("%Y-%m-%d")
    for r_base_idx, wbs_ref in enumerate(referencia_base_presupuesto, start=5):
        w_code = wbs_ref["wbs"]
        
        ws_evm_diario.cell(row=curr_row_evm, column=1, value=day_n).font = font_regular
        ws_evm_diario.cell(row=curr_row_evm, column=2, value=f_str).font = font_regular
        ws_evm_diario.cell(row=curr_row_evm, column=3, value=w_code).font = font_bold
        ws_evm_diario.cell(row=curr_row_evm, column=4, value=f"='01_PRESUPUESTO_CRONOGRAMA_BASE'!B{r_base_idx}").font = font_regular
        
        # Presupuesto BAC (Ref Base)
        ws_evm_diario.cell(row=curr_row_evm, column=5, value=f"='01_PRESUPUESTO_CRONOGRAMA_BASE'!C{r_base_idx}").number_format = 'S/ #,##0.00'
        
        # PV Programado Día (Fórmula Viva desde 02_CRONOGRAMA_PV_DIARIO)
        ws_evm_diario.cell(row=curr_row_evm, column=6, value=f"=SUMIFS('02_CRONOGRAMA_PV_DIARIO'!D:D, '02_CRONOGRAMA_PV_DIARIO'!B:B, B{curr_row_evm}, '02_CRONOGRAMA_PV_DIARIO'!C:C, C{curr_row_evm})").number_format = 'S/ #,##0.00'
        
        # PV Acumulado (Fórmula Viva)
        ws_evm_diario.cell(row=curr_row_evm, column=7, value=f"=SUMIFS('02_CRONOGRAMA_PV_DIARIO'!D:D, '02_CRONOGRAMA_PV_DIARIO'!B:B, \"<=\"&B{curr_row_evm}, '02_CRONOGRAMA_PV_DIARIO'!C:C, C{curr_row_evm})").number_format = 'S/ #,##0.00'
        
        # EV Ejecutado Día (Fórmula Viva desde 04_LOG_FIELD_ENTRIES)
        ws_evm_diario.cell(row=curr_row_evm, column=8, value=f"=SUMIFS('04_LOG_FIELD_ENTRIES'!J:J, '04_LOG_FIELD_ENTRIES'!B:B, B{curr_row_evm}, '04_LOG_FIELD_ENTRIES'!D:D, C{curr_row_evm}, '04_LOG_FIELD_ENTRIES'!K:K, \"EV_PRODUCCION\")").number_format = 'S/ #,##0.00'
        
        # EV Acumulado (Fórmula Viva)
        ws_evm_diario.cell(row=curr_row_evm, column=9, value=f"=SUMIFS('04_LOG_FIELD_ENTRIES'!J:J, '04_LOG_FIELD_ENTRIES'!B:B, \"<=\"&B{curr_row_evm}, '04_LOG_FIELD_ENTRIES'!D:D, C{curr_row_evm}, '04_LOG_FIELD_ENTRIES'!K:K, \"EV_PRODUCCION\")").number_format = 'S/ #,##0.00'
        
        # AC Costo Real Día (Fórmula Viva desde 04_LOG_FIELD_ENTRIES)
        ws_evm_diario.cell(row=curr_row_evm, column=10, value=f"=SUMIFS('04_LOG_FIELD_ENTRIES'!J:J, '04_LOG_FIELD_ENTRIES'!B:B, B{curr_row_evm}, '04_LOG_FIELD_ENTRIES'!D:D, C{curr_row_evm}, '04_LOG_FIELD_ENTRIES'!K:K, \"AC_*\")").number_format = 'S/ #,##0.00'
        
        # AC Acumulado (Fórmula Viva)
        ws_evm_diario.cell(row=curr_row_evm, column=11, value=f"=SUMIFS('04_LOG_FIELD_ENTRIES'!J:J, '04_LOG_FIELD_ENTRIES'!B:B, \"<=\"&B{curr_row_evm}, '04_LOG_FIELD_ENTRIES'!D:D, C{curr_row_evm}, '04_LOG_FIELD_ENTRIES'!K:K, \"AC_*\")").number_format = 'S/ #,##0.00'
        
        # Variación Costo CV (EV_Acum - AC_Acum)
        ws_evm_diario.cell(row=curr_row_evm, column=12, value=f"=I{curr_row_evm}-K{curr_row_evm}").number_format = 'S/ #,##0.00'
        
        # Variación Plazo SV (EV_Acum - PV_Acum)
        ws_evm_diario.cell(row=curr_row_evm, column=13, value=f"=I{curr_row_evm}-G{curr_row_evm}").number_format = 'S/ #,##0.00'
        
        # SPI Plazo (EV_Acum / PV_Acum)
        ws_evm_diario.cell(row=curr_row_evm, column=14, value=f"=IF(G{curr_row_evm}>0, I{curr_row_evm}/G{curr_row_evm}, 1)").number_format = '0.00'
        
        # CPI Costo (EV_Acum / AC_Acum)
        ws_evm_diario.cell(row=curr_row_evm, column=15, value=f"=IF(K{curr_row_evm}>0, I{curr_row_evm}/K{curr_row_evm}, 1)").number_format = '0.00'
        
        # Estado Alerta WBS
        formula_alerta_viva = f"=IF(AND(N{curr_row_evm}>=0.95, O{curr_row_evm}>=0.95), \"SALUDABLE\", IF(N{curr_row_evm}<0.9, \"ALERTA RETRASO\", \"ALERTA SOBRECOSTO\"))"
        ws_evm_diario.cell(row=curr_row_evm, column=16, value=formula_alerta_viva).font = font_bold

        for c in range(1, 17):
            ws_evm_diario.cell(row=curr_row_evm, column=c).border = thin_border
            
        curr_row_evm += 1


# ---------------------------------------------------------
# PESTAÑA 4: 04_LOG_FIELD_ENTRIES (ENTRADAS ENVIADAS DESDE CELULARES)
# ---------------------------------------------------------
ws_log = wb.create_sheet(title="04_LOG_FIELD_ENTRIES")
ws_log.views.sheetView[0].showGridLines = True

headers_log = [
    "ID Registro", "Fecha", "Rol Responsable", "Código WBS", 
    "Código Recurso/Partida", "Descripción / Detalle", "Cantidad Campo", 
    "Unidad", "P.U. (Busca en Maestro)", "Subtotal Monto (S/)", 
    "Categoría EVM", "Origen HTML"
]

for c_idx, h in enumerate(headers_log, start=1):
    cell = ws_log.cell(row=1, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = Alignment(horizontal="center", vertical="center")

for r_idx, reg in enumerate(field_entries_raw, start=2):
    ws_log.cell(row=r_idx, column=1, value=reg["id"]).font = font_regular
    ws_log.cell(row=r_idx, column=2, value=reg["fecha"]).font = font_regular
    ws_log.cell(row=r_idx, column=3, value=reg["rol"]).font = font_regular
    ws_log.cell(row=r_idx, column=4, value=reg["wbs"]).font = font_bold
    ws_log.cell(row=r_idx, column=5, value=reg["recurso"]).font = font_bold
    ws_log.cell(row=r_idx, column=6, value=reg["detalle"]).font = font_regular
    
    cell_q = ws_log.cell(row=r_idx, column=7, value=reg["cantidad"])
    cell_q.font = font_bold
    cell_q.number_format = '#,##0.00'
    
    ws_log.cell(row=r_idx, column=8, value=reg["unidad"]).font = font_regular
    
    # FÓRMULA VIVA DE BÚSQUEDA DE P.U. EN MAESTROS
    formula_pu = f"=IFERROR(VLOOKUP(E{r_idx}, '05_MAESTRO_RECURSOS'!A:D, 4, FALSE), IFERROR(VLOOKUP(E{r_idx}, '06_MAESTRO_PARTIDAS_EV'!A:D, 4, FALSE), 0))"
    cell_pu = ws_log.cell(row=r_idx, column=9, value=formula_pu)
    cell_pu.font = font_regular
    cell_pu.number_format = 'S/ #,##0.00'
    
    # FÓRMULA VIVA DE MONTO (Cantidad Campo * PU Maestro)
    formula_subtotal = f"=ROUND(G{r_idx} * I{r_idx}, 2)"
    cell_sub = ws_log.cell(row=r_idx, column=10, value=formula_subtotal)
    cell_sub.font = font_bold
    cell_sub.number_format = 'S/ #,##0.00'
    
    ws_log.cell(row=r_idx, column=11, value=reg["cat"]).font = font_bold
    ws_log.cell(row=r_idx, column=12, value=reg["html"]).font = font_regular

    for c in range(1, 13):
        ws_log.cell(row=r_idx, column=c).border = thin_border


# ---------------------------------------------------------
# PESTAÑA 5: 05_MAESTRO_RECURSOS (PRECIOS UNITARIOS DE INSUMOS)
# ---------------------------------------------------------
ws_rec = wb.create_sheet(title="05_MAESTRO_RECURSOS")
ws_rec.views.sheetView[0].showGridLines = True

ws_rec.append(["Código Recurso", "Descripción del Recurso / Insumo", "Unidad", "P.U. Meta Oficial (S/)", "Categoría EVM"])
for cell in ws_rec[1]:
    cell.font = font_header
    cell.fill = fill_purple
    cell.alignment = Alignment(horizontal="center")

for row in maestro_recursos:
    ws_rec.append([row["codigo"], row["descripcion"], row["unidad"], row["precio_unitario"], row["categoria"]])

for r in range(2, len(maestro_recursos) + 2):
    ws_rec.cell(row=r, column=4).number_format = 'S/ #,##0.00'
    for c in range(1, 6):
        ws_rec.cell(row=r, column=c).border = thin_border


# ---------------------------------------------------------
# PESTAÑA 6: 06_MAESTRO_PARTIDAS_EV (PRECIOS UNITARIOS DE PARTIDAS)
# ---------------------------------------------------------
ws_part = wb.create_sheet(title="06_MAESTRO_PARTIDAS_EV")
ws_part.views.sheetView[0].showGridLines = True

ws_part.append(["Código Partida", "Descripción de la Partida", "Unidad", "P.U. Directo Meta (S/)", "Categoría EVM"])
for cell in ws_part[1]:
    cell.font = font_header
    cell.fill = fill_green
    cell.alignment = Alignment(horizontal="center")

for row in maestro_partidas_ev:
    ws_part.append([row["codigo"], row["descripcion"], row["unidad"], row["precio_unitario"], row["categoria"]])

for r in range(2, len(maestro_partidas_ev) + 2):
    ws_part.cell(row=r, column=4).number_format = 'S/ #,##0.00'
    for c in range(1, 6):
        ws_part.cell(row=r, column=c).border = thin_border


# ---------------------------------------------------------
# AJUSTE DE ANCHO DE COLUMNAS EN TODAS LAS HOJAS
# ---------------------------------------------------------
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in sheet.merged_cells:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 55)

os.makedirs("outputs", exist_ok=True)
excel_sheets_path = "outputs/Base_Datos_Proyecto_Sheets_Viva.xlsx"
wb.save(excel_sheets_path)

print(f"[OK] Archivo Excel de Base de Datos Viva Completa con Arquitectura Total creado en: {excel_sheets_path}")
