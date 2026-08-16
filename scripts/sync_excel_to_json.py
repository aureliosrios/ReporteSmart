import openpyxl
import json

def sync():
    # 1. Cargar el Excel local
    wb = openpyxl.load_workbook('Base_Datos_Proyecto_Sheets_Viva.xlsx', data_only=True)
    
    # 2. Cargar los catálogos de precios para calcular unitarios y costos reales en Python
    sh5 = wb['05_MAESTRO_RECURSOS']
    sh6 = wb['06_MAESTRO_PARTIDAS_EV']
    
    recursos_precios = {}
    for r in range(2, sh5.max_row + 1):
        code = sh5.cell(r, 1).value
        price = sh5.cell(r, 4).value
        if code is not None and price is not None:
            recursos_precios[str(code)] = float(price)
            
    partidas_precios = {}
    for r in range(2, sh6.max_row + 1):
        code = sh6.cell(r, 2).value
        price = sh6.cell(r, 6).value
        if code is not None and price is not None:
            partidas_precios[str(code)] = float(price)

    # 3. Leer los registros de la pestaña 4 (04_LOG_FIELD_ENTRIES)
    sh4 = wb['04_LOG_FIELD_ENTRIES']
    logs_json = []
    
    print(f"Leyendo {sh4.max_row - 1} filas de la pestaña 04_LOG_FIELD_ENTRIES...")
    
    for r in range(2, sh4.max_row + 1):
        id_reg = sh4.cell(r, 1).value
        fecha = sh4.cell(r, 2).value
        rol = sh4.cell(r, 3).value
        wbs = sh4.cell(r, 4).value
        code = sh4.cell(r, 5).value
        desc = sh4.cell(r, 6).value
        qty = sh4.cell(r, 7).value
        unit = sh4.cell(r, 8).value
        category = sh4.cell(r, 11).value
        orig = sh4.cell(r, 12).value
        
        if not id_reg or not code:
            continue
            
        # Calcular el P.U. y Costo en Python para evitar valores None
        pu = 0.0
        if category == 'EV_PRODUCCION':
            pu = partidas_precios.get(str(code), 0.0)
        else:
            pu = recursos_precios.get(str(code), 0.0)
            
        qty = float(qty) if qty is not None else 0.0
        costo = round(qty * pu, 2)
        
        # Formatear fecha
        if isinstance(fecha, datetime_type := type(datetime_obj := None)):
            fecha_str = fecha.strftime("%Y-%m-%d")
        else:
            fecha_str = str(fecha).split(" ")[0]

        # Estructurar registro con ambas claves (legacy y web estándar)
        record = {
            "id": id_reg,
            "id_registro": id_reg,
            "id_parte": f"PARTE-{fecha_str}-{rol}",
            "fecha": fecha_str,
            "rol": rol,
            "emisor_rol": rol,
            "wbs": wbs,
            "wbs_codigo": wbs,
            "codigoRecurso": code,
            "codigo_recurso_partida": code,
            "detalle": desc,
            "descripcion": desc,
            "cantidad": qty,
            "unidad": unit,
            "pu": pu,
            "costo_unitario": pu,
            "costo": costo,
            "costo_total_pen": costo,
            "tipo": category,
            "categoria_evm": category,
            "origen_html": orig,
            "estado_validacion": "VALIDO",
            "fecha_recepcion": None
        }
        logs_json.append(record)

    # Ordenar registros por fecha e ID
    logs_json.sort(key=lambda x: (x["fecha"], x["id"]))

    # 4. Actualizar base_datos_reportabilidad.json
    try:
        with open('data/base_datos_reportabilidad.json', 'r', encoding='utf-8') as f:
            base_json = json.load(f)
    except Exception:
        base_json = {}
        
    base_json["registros_diarios"] = logs_json
    base_json["control_calidad"] = {
        "registros_recibidos": len(logs_json),
        "registros_validos": len(logs_json),
        "registros_observados": 0,
        "regla_duplicado": "solo se considera duplicado el mismo id_registro original"
    }
    
    with open('data/base_datos_reportabilidad.json', 'w', encoding='utf-8') as f:
        json.dump(base_json, f, indent=2, ensure_ascii=False)
    print("[OK] data/base_datos_reportabilidad.json sincronizado correctamente.")
    
    # 5. Actualizar base_datos_master_sheets.json
    try:
        with open('data/base_datos_master_sheets.json', 'r', encoding='utf-8') as f:
            master_json = json.load(f)
    except Exception:
        master_json = {}
        
    master_json["registros_diarios"] = logs_json
    
    with open('data/base_datos_master_sheets.json', 'w', encoding='utf-8') as f:
        json.dump(master_json, f, indent=2, ensure_ascii=False)
    print("[OK] data/base_datos_master_sheets.json sincronizado correctamente.")

if __name__ == '__main__':
    sync()
