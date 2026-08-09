import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------
# 1. ESTRUCTURA DE DATOS MAESTRA PARA 5 DÍAS DE CAMPO
# ---------------------------------------------------------

metadata = {
    "proyecto": "Redes Sanitarias de Agua Potable y Alcantarillado - Habilitación Urbana Los Cedros",
    "codigo_proyecto": "HU-CEDROS-2026",
    "version_esquema": "3.1.0",
    "moneda": "PEN",
    "periodo_simulacion": "5 Días de Operación de Campo (2026-08-01 al 2026-08-05)",
    "descripcion": "Base de datos en vivo de reportabilidad diaria desde portales móviles hacia Google Sheets / Excel con fórmulas vivas de EVM y KPIs"
}

wbs_catalog = [
    {"codigo": "WBS-100", "nombre": "Obras Preliminares, Trazo y Movilización", "bac_pen": 8296.40, "pv_acumulado_pen": 8296.40},
    {"codigo": "WBS-200", "nombre": "Red de Alcantarillado, Zanjas y Buzones", "bac_pen": 380086.56, "pv_acumulado_pen": 45000.00},
    {"codigo": "WBS-300", "nombre": "Red de Agua Potable y Conexiones Domiciliarias", "bac_pen": 57565.20, "pv_acumulado_pen": 12000.00},
    {"codigo": "WBS-400", "nombre": "Pruebas Hidráulicas, Desinfección y Entrega", "bac_pen": 9348.06, "pv_acumulado_pen": 2500.00}
]

recursos_catalog = {
    "MO_CAPATAZ": {"desc": "CAPATAZ", "unit": "hh", "precio": 32.50, "categoria": "AC_MO"},
    "MO_OPERARIO": {"desc": "OPERARIO", "unit": "hh", "precio": 26.80, "categoria": "AC_MO"},
    "MO_OFICIAL": {"desc": "OFICIAL", "unit": "hh", "precio": 22.40, "categoria": "AC_MO"},
    "MO_PEON": {"desc": "PEON", "unit": "hh", "precio": 20.10, "categoria": "AC_MO"},
    "MAT_ESTACA": {"desc": "ESTACAS Y MADERA PARA TRAZO", "unit": "pza", "precio": 4.50, "categoria": "AC_MAT"},
    "MAT_PINTURA": {"desc": "PINTURA SPRAY MARCADOR DE ZANJA", "unit": "gla", "precio": 18.00, "categoria": "AC_MAT"},
    "MAT_CORDEL": {"desc": "CORDEL / NYLON", "unit": "m", "precio": 0.80, "categoria": "AC_MAT"},
    "MAT_ARENA_CAMAS": {"desc": "ARENA FINA SELECCIONADA PARA CAMA E=0.10M", "unit": "m3", "precio": 45.00, "categoria": "AC_MAT"},
    "MAT_TUB_PVC_200": {"desc": "TUBERIA PVC UF DN 200MM SERIE S-20 ALCANTARILLADO", "unit": "m", "precio": 42.00, "categoria": "AC_MAT"},
    "MAT_LUBRICANTE": {"desc": "LUBRICANTE PARA TUBERIA CON ESPIGA Y CAMPANA", "unit": "kg", "precio": 24.00, "categoria": "AC_MAT"},
    "MAT_MATERIAL_PR": {"desc": "MATERIAL DE PRÉSTAMO AFIRMADO PARA RELLENO COMPACTADO", "unit": "m3", "precio": 38.00, "categoria": "AC_MAT"},
    "MAT_AGUA": {"desc": "AGUA PARA COMPACTACIÓN Y PRUEBAS", "unit": "m3", "precio": 12.00, "categoria": "AC_MAT"},
    "MAT_BUZON_PREF": {"desc": "BUZON PREFABRICADO CONCRETO H=1.5-2.5M CON MARCO Y TAPA", "unit": "und", "precio": 1050.00, "categoria": "AC_MAT"},
    "MAT_CAJA_REGISTRO": {"desc": "CAJA DE REGISTRO PREFABRICADA DE CONCRETO 12X24 WITH TAPA", "unit": "und", "precio": 145.00, "categoria": "AC_MAT"},
    "MAT_TUB_PVC_160_CONEX": {"desc": "TUBERIA PVC SAL DN 160MM PARA ACOMETIDA ALCANTARILLADO", "unit": "m", "precio": 22.00, "categoria": "AC_MAT"},
    "MAT_TUB_PVC_110_AGUA": {"desc": "TUBERIA PVC C-10 DN 110MM AGUA POTABLE", "unit": "m", "precio": 26.50, "categoria": "AC_MAT"},
    "MAT_CAJA_AGUA": {"desc": "CAJA TERMOFORMADA PARA MEDIDOR + ABRAZADERA + LLAVE DE PASO", "unit": "und", "precio": 62.00, "categoria": "AC_MAT"},
    "EQ_EXCAVADORA": {"desc": "EXCAVADORA SOBRE ORUGAS 140 HP", "unit": "hm", "precio": 185.00, "categoria": "AC_EQP"},
    "EQ_RETROEXCAVADORA": {"desc": "RETROEXCAVADORA SOBRE LLANTAS 62 HP", "unit": "hm", "precio": 130.00, "categoria": "AC_EQP"},
    "EQ_PLANCHA": {"desc": "PLANCHA COMPACTADORA 7 HP", "unit": "hm", "precio": 25.00, "categoria": "AC_EQP"},
    "EQ_RODILLO_CAMINANTE": {"desc": "RODILLO VIBRATORIO CAMINANTE 1.5 TN", "unit": "hm", "precio": 45.00, "categoria": "AC_EQP"},
    "EQ_CAMION_CISTERNA": {"desc": "CAMION CISTERNA 4X2 (AGUA) 2,000 GLN", "unit": "hm", "precio": 140.00, "categoria": "AC_EQP"},
    "EQ_CAMION_GRUAN": {"desc": "CAMION GRUA 5 TN PARA MONTAJE DE BUZONES", "unit": "hm", "precio": 160.00, "categoria": "AC_EQP"},
    "EQ_ESTACION_TOTAL": {"desc": "EQUIPO DE TOPOGRAFIA ESTACION TOTAL", "unit": "hm", "precio": 18.00, "categoria": "AC_EQP"},
    "SUB_MOVILIZACION": {"desc": "SUBCONTRATO DE MOVILIZACION Y DESMOVILIZACION DE MAQUINARIA", "unit": "glb", "precio": 4000.00, "categoria": "AC_SUB"},
    "SUB_ENSAYOS_COMPAC": {"desc": "SUBCONTRATO DE ENSAYOS DE DENSIDAD DE CAMPO (PROCTOR/DENSIDAD)", "unit": "und", "precio": 80.00, "categoria": "AC_SUB"},
    "SUB_MONTAJE_BUZON": {"desc": "SUBCONTRATO SERVICIO ASENTADO Y ANCLAJE BUZONES PREFABRICADOS", "unit": "und", "precio": 250.00, "categoria": "AC_SUB"}
}

partidas_ev_catalog = {
    "01.01": {"desc": "Obras Preliminares y Trabajos Provisionales", "unit": "GLB", "pu": 8296.40},
    "01.02.01": {"desc": "Trazo, Nivelación y Replanteo de Zanjas", "unit": "M", "pu": 3.35},
    "01.02.02": {"desc": "Excavación de Zanja H=1.50m - 2.20m a Máquina (Terreno Normal)", "unit": "M3", "pu": 11.66},
    "01.02.03": {"desc": "Preparación y Colocación de Cama de Arena e=0.10m", "unit": "M", "pu": 11.12},
    "01.02.04": {"desc": "Suministro e Instalación de Tubería PVC UF DN 200mm Serie S-20 para Alcantarillado", "unit": "M", "pu": 55.51},
    "01.02.05": {"desc": "Relleno Compactado de Zanja en Capas de 0.20m con Maquinaria/Plancha", "unit": "M3", "pu": 28.54},
    "01.02.06": {"desc": "Construcción de Buzones Prefabricados de Concreto h=1.50m - 2.50m (Inc. Marco y Tapa)", "unit": "UND", "pu": 2500.73},
    "01.02.07": {"desc": "Conexiones Domiciliarias de Alcantarillado (Caja de Registro + Acometida)", "unit": "UND", "pu": 405.54},
    "01.03.01": {"desc": "Suministro e Instalación de Tubería PVC C-10 DN 110mm para Agua Potable", "unit": "M", "pu": 38.76},
    "01.03.02": {"desc": "Conexiones Domiciliarias de Agua Potable (Caja de Agua + Abrazadera + Acometida)", "unit": "UND", "pu": 156.71},
    "01.04.01": {"desc": "Pruebas Hidráulicas de Redes de Agua y Alcantarillado + Desinfección", "unit": "GLB", "pu": 9348.06}
}

# ---------------------------------------------------------
# GENERACIÓN DE 5 DÍAS DE REGISTROS DE CAMPO
# ---------------------------------------------------------

registros_diarios = []

def add_log(id_reg, fecha, rol, wbs, cod, desc, cant, unit, pu, cat, html):
    costo = round(cant * pu, 2)
    registros_diarios.append({
        "id_registro": id_reg,
        "fecha": fecha,
        "rol": rol,
        "wbs_codigo": wbs,
        "codigo_recurso_partida": cod,
        "descripcion_detalle": desc,
        "cantidad": cant,
        "unidad": unit,
        "pu_meta_pen": pu,
        "costo_total_pen": costo,
        "categoria_evm": cat,
        "origen_html": html,
        "estado_validacion": "VALIDO"
    })

# DIA 1: 2026-08-01 (Movilización, Trazo y Primeras Zanjas)
d1 = "2026-08-01"
add_log("LOG-20260801-001", d1, "Tareador (Bildin)", "WBS-100", "MO_CAPATAZ", "CAPATAZ (1 pers. x 8h)", 8.0, "hh", 32.50, "AC_MO", "tareador.html")
add_log("LOG-20260801-002", d1, "Tareador (Bildin)", "WBS-100", "MO_OPERARIO", "OPERARIO (2 pers. x 8h)", 16.0, "hh", 26.80, "AC_MO", "tareador.html")
add_log("LOG-20260801-003", d1, "Tareador (Bildin)", "WBS-100", "MO_PEON", "PEON (4 pers. x 8h)", 32.0, "hh", 20.10, "AC_MO", "tareador.html")
add_log("LOG-20260801-004", d1, "Almacenero", "WBS-100", "MAT_ESTACA", "ESTACAS Y MADERA PARA TRAZO EN AV. LOS CEDROS", 150.0, "pza", 4.50, "AC_MAT", "almacenero.html")
add_log("LOG-20260801-005", d1, "Almacenero", "WBS-100", "MAT_PINTURA", "PINTURA SPRAY MARCADOR ZANJA", 10.0, "gla", 18.00, "AC_MAT", "almacenero.html")
add_log("LOG-20260801-006", d1, "Administradora", "WBS-100", "SUB_MOVILIZACION", "MOVILIZACION DE MAQUINARIA A OBRA", 1.0, "glb", 4000.00, "AC_SUB", "administradora.html")
add_log("LOG-20260801-007", d1, "Administradora", "WBS-100", "EQ_ESTACION_TOTAL", "ESTACION TOTAL TOPOGRAFICA", 8.0, "hm", 18.00, "AC_EQP", "administradora.html")
add_log("LOG-20260801-008", d1, "Ing. de Campo", "WBS-100", "01.01", "OBRAS PRELIMINARES Y TRABAJOS PROVISIONALES (50%)", 0.5, "GLB", 8296.40, "EV_PRODUCCION", "ing_campo.html")
add_log("LOG-20260801-009", d1, "Ing. de Campo", "WBS-100", "01.02.01", "TRAZO Y REPLANTEO INICIAL TRAMO T-1", 500.0, "M", 3.35, "EV_PRODUCCION", "ing_campo.html")

# DIA 2: 2026-08-02 (Excavación de Zanja a Máquina y Cama de Arena)
d2 = "2026-08-02"
add_log("LOG-20260802-010", d2, "Tareador (Bildin)", "WBS-200", "MO_CAPATAZ", "CAPATAZ CONTROL ZANJA", 8.0, "hh", 32.50, "AC_MO", "tareador.html")
add_log("LOG-20260802-011", d2, "Tareador (Bildin)", "WBS-200", "MO_OPERARIO", "OPERARIO ENTIBADO Y CAMA", 24.0, "hh", 26.80, "AC_MO", "tareador.html")
add_log("LOG-20260802-012", d2, "Tareador (Bildin)", "WBS-200", "MO_PEON", "PEON REPELE Y CAMA (6 pers. x 8h)", 48.0, "hh", 20.10, "AC_MO", "tareador.html")
add_log("LOG-20260802-013", d2, "Almacenero", "WBS-200", "MAT_ARENA_CAMAS", "ARENA FINA SELECCIONADA PARA CAMA E=0.10M", 35.0, "m3", 45.00, "AC_MAT", "almacenero.html")
add_log("LOG-20260802-014", d2, "Administradora", "WBS-200", "EQ_EXCAVADORA", "EXCAVADORA SOBRE ORUGAS 140 HP EN ZANJA PRINCIPAL", 8.0, "hm", 185.00, "AC_EQP", "administradora.html")
add_log("LOG-20260802-015", d2, "Ing. de Campo", "WBS-200", "01.02.02", "EXCAVACION DE ZANJA H=1.5-2.2M A MAQUINA TRAMO T-1", 450.0, "M3", 11.66, "EV_PRODUCCION", "ing_campo.html")
add_log("LOG-20260802-016", d2, "Ing. de Campo", "WBS-200", "01.02.03", "COLOCACION DE CAMA DE ARENA E=0.10M TRAMO T-1", 300.0, "M", 11.12, "EV_PRODUCCION", "ing_campo.html")

# DIA 3: 2026-08-03 (Instalación de Tubería PVC 200mm y Relleno)
d3 = "2026-08-03"
add_log("LOG-20260803-017", d3, "Tareador (Bildin)", "WBS-200", "MO_OPERARIO", "OPERARIO TUBERO ALCANTARILLADO (3 pers. x 8h)", 24.0, "hh", 26.80, "AC_MO", "tareador.html")
add_log("LOG-20260803-018", d3, "Tareador (Bildin)", "WBS-200", "MO_OFICIAL", "OFICIAL COMPACTACION Y LUBRICADO", 16.0, "hh", 22.40, "AC_MO", "tareador.html")
add_log("LOG-20260803-019", d3, "Tareador (Bildin)", "WBS-200", "MO_PEON", "PEON APOYO MONTAJE TUBERIA", 40.0, "hh", 20.10, "AC_MO", "tareador.html")
add_log("LOG-20260803-020", d3, "Almacenero", "WBS-200", "MAT_TUB_PVC_200", "TUBERIA PVC UF DN 200MM S-20 ALCANTARILLADO", 280.0, "m", 42.00, "AC_MAT", "almacenero.html")
add_log("LOG-20260803-021", d3, "Almacenero", "WBS-200", "MAT_LUBRICANTE", "LUBRICANTE PARA TUBERIA PVC", 12.0, "kg", 24.00, "AC_MAT", "almacenero.html")
add_log("LOG-20260803-022", d3, "Almacenero", "WBS-200", "MAT_MATERIAL_PR", "MATERIAL AFIRMADO PRESTAMO RELLENO COMPACTADO", 60.0, "m3", 38.00, "AC_MAT", "almacenero.html")
add_log("LOG-20260803-023", d3, "Administradora", "WBS-200", "EQ_RETROEXCAVADORA", "RETROEXCAVADORA RELLENO Y LIMPIEZA", 8.0, "hm", 130.00, "AC_EQP", "administradora.html")
add_log("LOG-20260803-024", d3, "Administradora", "WBS-200", "EQ_PLANCHA", "PLANCHA COMPACTADORA 7 HP EN ZANJA", 8.0, "hm", 25.00, "AC_EQP", "administradora.html")
add_log("LOG-20260803-025", d3, "Ing. de Campo", "WBS-200", "01.02.04", "INSTALACION TUBERIA PVC UF DN 200MM ALCANTARILLADO", 250.0, "M", 55.51, "EV_PRODUCCION", "ing_campo.html")
add_log("LOG-20260803-026", d3, "Ing. de Campo", "WBS-200", "01.02.05", "RELLENO COMPACTADO DE ZANJA CAPAS 0.20M", 200.0, "M3", 28.54, "EV_PRODUCCION", "ing_campo.html")

# DIA 4: 2026-08-04 (Buzones Prefabricados y Red de Agua Potable)
d4 = "2026-08-04"
add_log("LOG-20260804-027", d4, "Tareador (Bildin)", "WBS-200", "MO_OPERARIO", "OPERARIO MONTAJE BUZONES", 16.0, "hh", 26.80, "AC_MO", "tareador.html")
add_log("LOG-20260804-028", d4, "Tareador (Bildin)", "WBS-300", "MO_OPERARIO", "OPERARIO TUBERO AGUA POTABLE", 16.0, "hh", 26.80, "AC_MO", "tareador.html")
add_log("LOG-20260804-029", d4, "Tareador (Bildin)", "WBS-300", "MO_PEON", "PEON APERTURA ZANJA AGUA", 32.0, "hh", 20.10, "AC_MO", "tareador.html")
add_log("LOG-20260804-030", d4, "Almacenero", "WBS-200", "MAT_BUZON_PREF", "BUZON PREFABRICADO DE CONCRETO H=2.0M CON MARCO/TAPA", 4.0, "und", 1050.00, "AC_MAT", "almacenero.html")
add_log("LOG-20260804-031", d4, "Almacenero", "WBS-300", "MAT_TUB_PVC_110_AGUA", "TUBERIA PVC C-10 DN 110MM AGUA POTABLE", 200.0, "m", 26.50, "AC_MAT", "almacenero.html")
add_log("LOG-20260804-032", d4, "Administradora", "WBS-200", "EQ_CAMION_GRUAN", "CAMION GRUA 5 TN MONTAJE BUZONES PREFABRICADOS", 6.0, "hm", 160.00, "AC_EQP", "administradora.html")
add_log("LOG-20260804-033", d4, "Administradora", "WBS-200", "SUB_MONTAJE_BUZON", "SUBCONTRATO ASENTADO Y SELLADO DE BUZONES", 4.0, "und", 250.00, "AC_SUB", "administradora.html")
add_log("LOG-20260804-034", d4, "Ing. de Campo", "WBS-200", "01.02.06", "CONSTRUCCION DE BUZONES PREFABRICADOS CONCRETO H=1.5-2.5M", 4.0, "UND", 2500.73, "EV_PRODUCCION", "ing_campo.html")
add_log("LOG-20260804-035", d4, "Ing. de Campo", "WBS-300", "01.03.01", "INSTALACION TUBERIA PVC C-10 DN 110MM AGUA POTABLE", 180.0, "M", 38.76, "EV_PRODUCCION", "ing_campo.html")

# DIA 5: 2026-08-05 (Conexiones Domiciliarias, Ensayos de Campo y Pruebas)
d5 = "2026-08-05"
add_log("LOG-20260805-036", d5, "Tareador (Bildin)", "WBS-200", "MO_OPERARIO", "OPERARIO CONEXION DOMICILIARIA ALCANTARILLADO", 16.0, "hh", 26.80, "AC_MO", "tareador.html")
add_log("LOG-20260805-037", d5, "Tareador (Bildin)", "WBS-300", "MO_OPERARIO", "OPERARIO CONEXION DOMICILIARIA AGUA", 16.0, "hh", 26.80, "AC_MO", "tareador.html")
add_log("LOG-20260805-038", d5, "Tareador (Bildin)", "WBS-400", "MO_OFICIAL", "OFICIAL PRUEBAS HIDRAULICAS DE RED", 8.0, "hh", 22.40, "AC_MO", "tareador.html")
add_log("LOG-20260805-039", d5, "Almacenero", "WBS-200", "MAT_CAJA_REGISTRO", "CAJA DE REGISTRO PREFABRICADA 12X24 WITH TAPA", 15.0, "und", 145.00, "AC_MAT", "almacenero.html")
add_log("LOG-20260805-040", d5, "Almacenero", "WBS-200", "MAT_TUB_PVC_160_CONEX", "TUBERIA PVC SAL DN 160MM ACOMETIDA", 90.0, "m", 22.00, "AC_MAT", "almacenero.html")
add_log("LOG-20260805-041", d5, "Almacenero", "WBS-300", "MAT_CAJA_AGUA", "CAJA TERMOFORMADA MEDIDOR + ABRAZADERA + LLAVE", 15.0, "und", 62.00, "AC_MAT", "almacenero.html")
add_log("LOG-20260805-042", d5, "Administradora", "WBS-200", "SUB_ENSAYOS_COMPAC", "SUBCONTRATO ENSAYOS DENSIDAD DE CAMPO (PROCTOR)", 10.0, "und", 80.00, "AC_SUB", "administradora.html")
add_log("LOG-20260805-043", d5, "Administradora", "WBS-400", "EQ_CAMION_CISTERNA", "CAMION CISTERNA AGUA PRUEBAS HIDRAULICAS", 8.0, "hm", 140.00, "AC_EQP", "administradora.html")
add_log("LOG-20260805-044", d5, "Ing. de Campo", "WBS-200", "01.02.07", "CONEXIONES DOMICILIARIAS ALCANTARILLADO (15 ACOMETIDAS)", 15.0, "UND", 405.54, "EV_PRODUCCION", "ing_campo.html")
add_log("LOG-20260805-045", d5, "Ing. de Campo", "WBS-300", "01.03.02", "CONEXIONES DOMICILIARIAS AGUA POTABLE (15 ACOMETIDAS)", 15.0, "UND", 156.71, "EV_PRODUCCION", "ing_campo.html")
add_log("LOG-20260805-046", d5, "Ing. de Campo", "WBS-400", "01.04.01", "PRUEBA HIDRAULICA TRAMO 1 ALCANTARILLADO Y AGUA (25%)", 0.25, "GLB", 9348.06, "EV_PRODUCCION", "ing_campo.html")

# ---------------------------------------------------------
# GUARDAR BASE DE DATOS JSON CANÓNICA CON FÓRMULAS ESPECIFICADAS
# ---------------------------------------------------------

db_json = {
    "metadata": metadata,
    "catalogos": {
        "wbs": wbs_catalog,
        "recursos": recursos_catalog,
        "partidas_ev": partidas_ev_catalog
    },
    "especificacion_formulas_vivas_excel": {
        "formula_costo_total_registro": "=ROUND(G{row}*I{row}, 2)",
        "formula_sumatoria_ev_total": "=SUMIFS(REGISTROS_DIARIOS!J:J, REGISTROS_DIARIOS!K:K, \"EV_PRODUCCION\")",
        "formula_sumatoria_ac_total": "=SUMIF(REGISTROS_DIARIOS!K:K, \"AC_*\", REGISTROS_DIARIOS!J:J)",
        "formula_cpi": "=IF(AC>0, EV/AC, 1)",
        "formula_spi": "=IF(PV>0, EV/PV, 1)",
        "formula_eac": "=IF(CPI>0, BAC/CPI, BAC)",
        "formula_ac_por_wbs": "=SUMIFS(REGISTROS_DIARIOS!J:J, REGISTROS_DIARIOS!D:D, WBS_CODE, REGISTROS_DIARIOS!K:K, \"AC_*\")",
        "formula_ev_por_wbs": "=SUMIFS(REGISTROS_DIARIOS!J:J, REGISTROS_DIARIOS!D:D, WBS_CODE, REGISTROS_DIARIOS!K:K, \"EV_PRODUCCION\")"
    },
    "registros_diarios": registros_diarios,
    "resumen_5_dias": {
        "total_registros": len(registros_diarios),
        "fechas": ["2026-08-01", "2026-08-02", "2026-08-03", "2026-08-04", "2026-08-05"],
        "roles_participantes": ["Tareador (Bildin)", "Almacenero", "Administradora", "Ing. de Campo"]
    }
}

os.makedirs("docs/data", exist_ok=True)
json_path = "docs/data/base_datos_reportabilidad_5dias.json"
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(db_json, f, ensure_ascii=False, indent=2)

print(f"[OK] Archivo JSON creado en: {json_path}")

# ---------------------------------------------------------
# 2. GENERACIÓN DEL EXCEL CON FÓRMULAS VIVAS EN OPENPYXL
# ---------------------------------------------------------

wb = openpyxl.Workbook()
# Eliminar hoja default
wb.remove(wb.active)

# Estilos de Excel
font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
font_section = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=10, bold=True)
font_regular = Font(name="Calibri", size=10)

fill_dark = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark slate
fill_blue_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid") # Royal Blue
fill_green_header = PatternFill(start_color="15803D", end_color="15803D", fill_type="solid") # Green
fill_orange_header = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid") # Orange
fill_purple_header = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid") # Purple
fill_kpi_card = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
fill_zebra = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)
double_bottom_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='double', color='1E3A8A')
)

# ---------------------------------------------------------
# HOJA 1: 01_DASHBOARD_KPI
# ---------------------------------------------------------
ws_dash = wb.create_sheet(title="01_DASHBOARD_KPI")
ws_dash.views.sheetView[0].showGridLines = True

# Banner Principal
ws_dash.merge_cells("A1:I2")
ws_dash["A1"] = "PANEL CENTRAL DE CONTROL DE RESULTADO OPERATIVO (RO) Y EVM - 5 DÍAS DE OBRA"
ws_dash["A1"].font = font_title
ws_dash["A1"].fill = fill_dark
ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Subtítulo
ws_dash["A3"] = "Habilitación Urbana Los Cedros | Redes Sanitarias | Base de Datos Viva Google Sheets / Excel"
ws_dash["A3"].font = font_section

# SECCIÓN 1: TARJETAS KPI RESUMEN (Filas 5 a 8)
kpi_titles = ["Presupuesto Meta (BAC)", "Planificado a la Fecha (PV)", "Valor Ganado Exec (EV)", "Costo Real Incurrido (AC)", "Variación Costo (CV = EV-AC)", "Índice Costo (CPI)", "Índice Plazo (SPI)", "Proyección Cierre (EAC)", "Desvío Meta (EAC - BAC)"]
kpi_formulas = [
    "=SUM('04_CATALOGO_WBS'!C4:C7)",
    "=SUM('04_CATALOGO_WBS'!D4:D7)",
    "=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!K:K, \"EV_PRODUCCION\")",
    "=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!K:K, \"AC_*\")",
    "=C6-D6",
    "=IF(D6>0, C6/D6, 1)",
    "=IF(B6>0, C6/B6, 1)",
    "=IF(F6>0, A6/F6, A6)",
    "=H6-A6"
]
kpi_formats = ['S/ #,##0.00', 'S/ #,##0.00', 'S/ #,##0.00', 'S/ #,##0.00', 'S/ #,##0.00', '0.00', '0.00', 'S/ #,##0.00', 'S/ #,##0.00']

# Fila 5: Encabezados KPI
col_idx = 1
for title in kpi_titles:
    cell = ws_dash.cell(row=5, column=col_idx, value=title)
    cell.font = font_header
    cell.fill = fill_blue_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    col_idx += 1

# Fila 6: Valores KPI (Fórmulas Vivas)
col_idx = 1
for form, fmt in zip(kpi_formulas, kpi_formats):
    cell = ws_dash.cell(row=6, column=col_idx, value=form)
    cell.font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    cell.fill = fill_kpi_card
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    col_idx += 1

# SECCIÓN 2: DESGLOSE DE COSTOS REALES (AC) POR ROL (Filas 10 a 15)
ws_dash["A10"] = "DESGLOSE DE COSTO REAL INCURRIDO ($AC$) POR COMPONENTE Y ROL DE CAMPO"
ws_dash["A10"].font = font_section

headers_ac = ["Componente Costo", "Rol Responsables", "Tipo EVM", "Fórmula Sumatoria Viva", "Monto Incurrido (PEN)", "% Incidencia AC"]
for c_idx, h in enumerate(headers_ac, start=1):
    cell = ws_dash.cell(row=11, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_dark
    cell.alignment = Alignment(horizontal="center")

ac_rows = [
    ("Mano de Obra ($HH$)", "Tareador (Bildin)", "AC_MO", "=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!K:K, \"AC_MO\")"),
    ("Materiales de Almacén", "Almacenero", "AC_MAT", "=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!K:K, \"AC_MAT\")"),
    ("Equipos y Horas Máquina", "Administradora", "AC_EQP", "=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!K:K, \"AC_EQP\")"),
    ("Subcontratos y Ensayos", "Administradora", "AC_SUB", "=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!K:K, \"AC_SUB\")")
]

r_idx = 12
for comp, rol, tipo, f_monto in ac_rows:
    ws_dash.cell(row=r_idx, column=1, value=comp).font = font_regular
    ws_dash.cell(row=r_idx, column=2, value=rol).font = font_regular
    ws_dash.cell(row=r_idx, column=3, value=tipo).font = font_bold
    
    cell_monto = ws_dash.cell(row=r_idx, column=4, value=f_monto)
    cell_monto.font = font_bold
    cell_monto.number_format = 'S/ #,##0.00'
    
    cell_val = ws_dash.cell(row=r_idx, column=5, value=f"=D{r_idx}")
    cell_val.font = font_bold
    cell_val.number_format = 'S/ #,##0.00'
    
    cell_porc = ws_dash.cell(row=r_idx, column=6, value=f"=E{r_idx}/E$16")
    cell_porc.font = font_bold
    cell_porc.number_format = '0.00%'

    for c in range(1, 7):
        ws_dash.cell(row=r_idx, column=c).border = thin_border
    r_idx += 1

# Total AC
ws_dash.cell(row=16, column=1, value="TOTAL COSTO REAL INCURRIDO (AC)").font = font_bold
ws_dash.cell(row=16, column=5, value="=SUM(E12:E15)").font = font_bold
ws_dash.cell(row=16, column=5).number_format = 'S/ #,##0.00'
ws_dash.cell(row=16, column=6, value="=SUM(F12:F15)").font = font_bold
ws_dash.cell(row=16, column=6).number_format = '0.00%'
for c in range(1, 7):
    ws_dash.cell(row=16, column=c).border = double_bottom_border

# SECCIÓN 3: EVOLUCIÓN DIARIA AC vs EV (Días 1 al 5) (Filas 18 a 26)
ws_dash["A18"] = "HISTÓRICO Y EVOLUCIÓN DIARIA DE AVANCE (DÍA 1 AL DÍA 5)"
ws_dash["A18"].font = font_section

headers_dias = ["Fecha", "Costo MO (S/)", "Costo Mat (S/)", "Costo Eqp/Sub (S/)", "Total Costo Real ($AC$)", "Valor Ganado ($EV$)", "Diferencia Diaria ($EV-AC$)", "CPI Diario"]
for c_idx, h in enumerate(headers_dias, start=1):
    cell = ws_dash.cell(row=19, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_purple_header
    cell.alignment = Alignment(horizontal="center")

fechas_list = ["2026-08-01", "2026-08-02", "2026-08-03", "2026-08-04", "2026-08-05"]
r_d = 20
for f in fechas_list:
    ws_dash.cell(row=r_d, column=1, value=f).font = font_bold
    ws_dash.cell(row=r_d, column=2, value=f"=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!B:B, A{r_d}, '02_REGISTROS_DIARIOS'!K:K, \"AC_MO\")").number_format = 'S/ #,##0.00'
    ws_dash.cell(row=r_d, column=3, value=f"=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!B:B, A{r_d}, '02_REGISTROS_DIARIOS'!K:K, \"AC_MAT\")").number_format = 'S/ #,##0.00'
    ws_dash.cell(row=r_d, column=4, value=f"=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!B:B, A{r_d}, '02_REGISTROS_DIARIOS'!K:K, \"AC_EQP\") + SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!B:B, A{r_d}, '02_REGISTROS_DIARIOS'!K:K, \"AC_SUB\")").number_format = 'S/ #,##0.00'
    ws_dash.cell(row=r_d, column=5, value=f"=SUM(B{r_d}:D{r_d})").number_format = 'S/ #,##0.00'
    ws_dash.cell(row=r_d, column=6, value=f"=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!B:B, A{r_d}, '02_REGISTROS_DIARIOS'!K:K, \"EV_PRODUCCION\")").number_format = 'S/ #,##0.00'
    ws_dash.cell(row=r_d, column=7, value=f"=F{r_d}-E{r_d}").number_format = 'S/ #,##0.00'
    ws_dash.cell(row=r_d, column=8, value=f"=IF(E{r_d}>0, F{r_d}/E{r_d}, 1)").number_format = '0.00'

    for c in range(1, 9):
        ws_dash.cell(row=r_d, column=c).border = thin_border
    r_d += 1

# Total Historico
ws_dash.cell(row=25, column=1, value="TOTAL ACUMULADO OBRA").font = font_bold
for c_i, col_let in enumerate(["B", "C", "D", "E", "F", "G"], start=2):
    cell = ws_dash.cell(row=25, column=c_i, value=f"=SUM({col_let}20:{col_let}24)")
    cell.font = font_bold
    cell.number_format = 'S/ #,##0.00'
ws_dash.cell(row=25, column=8, value="=IF(E25>0, F25/E25, 1)").font = font_bold
ws_dash.cell(row=25, column=8).number_format = '0.00'
for c in range(1, 9):
    ws_dash.cell(row=25, column=c).border = double_bottom_border


# ---------------------------------------------------------
# HOJA 2: 02_REGISTROS_DIARIOS (BASE DE DATOS VIVA DE CAMPO)
# ---------------------------------------------------------
ws_reg = wb.create_sheet(title="02_REGISTROS_DIARIOS")
ws_reg.views.sheetView[0].showGridLines = True

headers_reg = ["ID Registro", "Fecha", "Rol Responsable", "Código WBS", "Código Recurso/Partida", "Descripción / Detalle", "Cantidad", "Unidad", "P.U. Meta (S/)", "Costo Total (S/)", "Categoría EVM", "Origen HTML", "Estado Validado"]

for col_idx, h in enumerate(headers_reg, start=1):
    cell = ws_reg.cell(row=1, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue_header
    cell.alignment = Alignment(horizontal="center", vertical="center")

row_curr = 2
for reg in registros_diarios:
    ws_reg.cell(row=row_curr, column=1, value=reg["id_registro"]).font = font_regular
    ws_reg.cell(row=row_curr, column=2, value=reg["fecha"]).font = font_regular
    ws_reg.cell(row=row_curr, column=3, value=reg["rol"]).font = font_regular
    ws_reg.cell(row=row_curr, column=4, value=reg["wbs_codigo"]).font = font_bold
    ws_reg.cell(row=row_curr, column=5, value=reg["codigo_recurso_partida"]).font = font_regular
    ws_reg.cell(row=row_curr, column=6, value=reg["descripcion_detalle"]).font = font_regular
    
    cell_cant = ws_reg.cell(row=row_curr, column=7, value=reg["cantidad"])
    cell_cant.font = font_regular
    cell_cant.number_format = '#,##0.00'
    
    ws_reg.cell(row=row_curr, column=8, value=reg["unidad"]).font = font_regular
    
    cell_pu = ws_reg.cell(row=row_curr, column=9, value=reg["pu_meta_pen"])
    cell_pu.font = font_regular
    cell_pu.number_format = 'S/ #,##0.00'
    
    # FÓRMULA VIVA DE COSTO TOTAL (Cantidad * PU)
    cell_costo = ws_reg.cell(row=row_curr, column=10, value=f"=ROUND(G{row_curr}*I{row_curr}, 2)")
    cell_costo.font = font_bold
    cell_costo.number_format = 'S/ #,##0.00'
    
    ws_reg.cell(row=row_curr, column=11, value=reg["categoria_evm"]).font = font_bold
    ws_reg.cell(row=row_curr, column=12, value=reg["origen_html"]).font = font_regular
    ws_reg.cell(row=row_curr, column=13, value=reg["estado_validacion"]).font = font_regular
    
    for c in range(1, 14):
        ws_reg.cell(row=row_curr, column=c).border = thin_border
    
    row_curr += 1


# ---------------------------------------------------------
# HOJA 3: 03_EVM_POR_WBS
# ---------------------------------------------------------
ws_wbs_evm = wb.create_sheet(title="03_EVM_POR_WBS")
ws_wbs_evm.views.sheetView[0].showGridLines = True

headers_wbs_evm = [
    "Código WBS", "Descripción Frente WBS", "Presupuesto BAC Meta (S/)", "Planificado PV (S/)",
    "Costo MO (S/)", "Costo Mat (S/)", "Costo Eqp/Sub (S/)", "Total Costo Real AC (S/)",
    "Valor Ganado EV (S/)", "Diferencia (EV - AC)", "CPI (Costo)", "SPI (Plazo)", "EAC Proyectado (S/)"
]

for col_idx, h in enumerate(headers_wbs_evm, start=1):
    cell = ws_wbs_evm.cell(row=1, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_dark
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

r_wbs = 2
for wbs in wbs_catalog:
    code = wbs["codigo"]
    name = wbs["nombre"]
    
    ws_wbs_evm.cell(row=r_wbs, column=1, value=code).font = font_bold
    ws_wbs_evm.cell(row=r_wbs, column=2, value=name).font = font_regular
    
    # BAC y PV desde el catálogo
    ws_wbs_evm.cell(row=r_wbs, column=3, value=f"='04_CATALOGO_WBS'!C{r_wbs+2}").number_format = 'S/ #,##0.00'
    ws_wbs_evm.cell(row=r_wbs, column=4, value=f"='04_CATALOGO_WBS'!D{r_wbs+2}").number_format = 'S/ #,##0.00'
    
    # Fórmulas Vivas apuntando a REGISTROS_DIARIOS
    ws_wbs_evm.cell(row=r_wbs, column=5, value=f"=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!D:D, A{r_wbs}, '02_REGISTROS_DIARIOS'!K:K, \"AC_MO\")").number_format = 'S/ #,##0.00'
    ws_wbs_evm.cell(row=r_wbs, column=6, value=f"=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!D:D, A{r_wbs}, '02_REGISTROS_DIARIOS'!K:K, \"AC_MAT\")").number_format = 'S/ #,##0.00'
    ws_wbs_evm.cell(row=r_wbs, column=7, value=f"=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!D:D, A{r_wbs}, '02_REGISTROS_DIARIOS'!K:K, \"AC_EQP\") + SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!D:D, A{r_wbs}, '02_REGISTROS_DIARIOS'!K:K, \"AC_SUB\")").number_format = 'S/ #,##0.00'
    
    # Total AC
    ws_wbs_evm.cell(row=r_wbs, column=8, value=f"=SUM(E{r_wbs}:G{r_wbs})").number_format = 'S/ #,##0.00'
    
    # Total EV
    ws_wbs_evm.cell(row=r_wbs, column=9, value=f"=SUMIFS('02_REGISTROS_DIARIOS'!J:J, '02_REGISTROS_DIARIOS'!D:D, A{r_wbs}, '02_REGISTROS_DIARIOS'!K:K, \"EV_PRODUCCION\")").number_format = 'S/ #,##0.00'
    
    # EV - AC
    ws_wbs_evm.cell(row=r_wbs, column=10, value=f"=I{r_wbs}-H{r_wbs}").number_format = 'S/ #,##0.00'
    
    # CPI y SPI
    ws_wbs_evm.cell(row=r_wbs, column=11, value=f"=IF(H{r_wbs}>0, I{r_wbs}/H{r_wbs}, 1)").number_format = '0.00'
    ws_wbs_evm.cell(row=r_wbs, column=12, value=f"=IF(D{r_wbs}>0, I{r_wbs}/D{r_wbs}, 1)").number_format = '0.00'
    
    # EAC Proyectado
    ws_wbs_evm.cell(row=r_wbs, column=13, value=f"=IF(K{r_wbs}>0, C{r_wbs}/K{r_wbs}, C{r_wbs})").number_format = 'S/ #,##0.00'

    for c in range(1, 14):
        ws_wbs_evm.cell(row=r_wbs, column=c).border = thin_border
    r_wbs += 1

# Total WBS
ws_wbs_evm.cell(row=6, column=1, value="TOTAL OBRA WBS").font = font_bold
ws_wbs_evm.cell(row=6, column=2, value="Consolidado de Redes Sanitarias").font = font_regular
for c_i, col_let in enumerate(["C", "D", "E", "F", "G", "H", "I", "J"], start=3):
    cell = ws_wbs_evm.cell(row=6, column=c_i, value=f"=SUM({col_let}2:{col_let}5)")
    cell.font = font_bold
    cell.number_format = 'S/ #,##0.00'

ws_wbs_evm.cell(row=6, column=11, value="=IF(H6>0, I6/H6, 1)").font = font_bold
ws_wbs_evm.cell(row=6, column=11).number_format = '0.00'

ws_wbs_evm.cell(row=6, column=12, value="=IF(D6>0, I6/D6, 1)").font = font_bold
ws_wbs_evm.cell(row=6, column=12).number_format = '0.00'

ws_wbs_evm.cell(row=6, column=13, value="=IF(K6>0, C6/K6, C6)").font = font_bold
ws_wbs_evm.cell(row=6, column=13).number_format = 'S/ #,##0.00'

for c in range(1, 14):
    ws_wbs_evm.cell(row=6, column=c).border = double_bottom_border


# ---------------------------------------------------------
# HOJA 4: 04_CATALOGO_WBS
# ---------------------------------------------------------
ws_cat_wbs = wb.create_sheet(title="04_CATALOGO_WBS")
ws_cat_wbs.views.sheetView[0].showGridLines = True

ws_cat_wbs.append(["CÓDIGO WBS", "NOMBRE DEL FRENTE / WBS", "PRESUPUESTO META (BAC S/)", "PLANIFICADO ACUMULADO (PV S/)"])
for cell in ws_cat_wbs[1]:
    cell.font = font_header
    cell.fill = fill_blue_header
    cell.alignment = Alignment(horizontal="center")

for row in wbs_catalog:
    ws_cat_wbs.append([row["codigo"], row["nombre"], row["bac_pen"], row["pv_acumulado_pen"]])

for r in range(2, 6):
    ws_cat_wbs.cell(row=r, column=3).number_format = 'S/ #,##0.00'
    ws_cat_wbs.cell(row=r, column=4).number_format = 'S/ #,##0.00'
    for c in range(1, 5):
        ws_cat_wbs.cell(row=r, column=c).border = thin_border


# ---------------------------------------------------------
# HOJA 5: 05_CATALOGO_RECURSOS
# ---------------------------------------------------------
ws_cat_rec = wb.create_sheet(title="05_CATALOGO_RECURSOS")
ws_cat_rec.views.sheetView[0].showGridLines = True

ws_cat_rec.append(["Código Recurso", "Descripción Recurso / Insumo", "Unidad", "P.U. Meta (S/)", "Categoría EVM"])
for cell in ws_cat_rec[1]:
    cell.font = font_header
    cell.fill = fill_dark
    cell.alignment = Alignment(horizontal="center")

for cod, d in recursos_catalog.items():
    ws_cat_rec.append([cod, d["desc"], d["unit"], d["precio"], d["categoria"]])

for r in range(2, len(recursos_catalog) + 2):
    ws_cat_rec.cell(row=r, column=4).number_format = 'S/ #,##0.00'
    for c in range(1, 6):
        ws_cat_rec.cell(row=r, column=c).border = thin_border


# ---------------------------------------------------------
# AJUSTE DE ANCHO DE COLUMNAS AUTO-FIT EN TODAS LAS HOJAS
# ---------------------------------------------------------
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Ignorar celdas combinadas o sin valor
            if cell.coordinate in sheet.merged_cells:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 55)

os.makedirs("outputs", exist_ok=True)
excel_path = "outputs/Reporte_RO_Base_Viva_5Dias.xlsx"
wb.save(excel_path)

print(f"[OK] Archivo Excel con Formulas Vivas creado en: {excel_path}")
